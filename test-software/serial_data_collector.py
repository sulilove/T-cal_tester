
""" 串口/以太网数据采集与实时绘图程序（支持LAN连接 - 多物理量多Y轴版）
     温度源控制功能：通过串口 115200 波特率发送 OUTP:STAT 1/0 控制温度源开闭
功能：
1. 支持最多6台设备（串口或LAN）的实时数据采集
2. 实时绘制数据曲线（集成实时数值显示），不同物理量使用独立Y轴
3. 可设置读取间隔（ms）和读取命令
4. 数据导出为Excel文件
5. 优化UI：曲线图例名称 100% 同步设备配置名称
6. 增强解析校验：MEASure:TEMPerature?取第2值(内部温度)，MEASure:CONTrol?取第4值(加热功率)
7. 支持配置保存/导入，自动加载上次配置
8. 串口调试窗口显示详细通讯细节（发送命令、原始响应、解析结果）
9. 优化曲线显示：多Y轴自动布局，同物理量共用Y轴，不同物理量Y轴不交叠
10. 自动保存功能：停止采集时自动保存数据到本地硬盘
11. 清空缓存功能：清空未保存到硬盘的临时数据
12. LAN连接支持（端口号自动处理空值，默认8000）
13. 定时自动保存：采集过程中每45分钟自动保存数据到硬盘
14. 应急自动保存：程序意外关机或崩溃时自动保存最新数据
"""

APP_VERSION = "1.10.0"

import sys
import os
import platform
from PyQt6.QtWidgets import *
from PyQt6.QtCore import *
from PyQt6.QtGui import *
import pyqtgraph as pg
import numpy as np
import threading
import time
import socket
from collections import deque
import re
import json
import pandas as pd
from datetime import datetime, timedelta
import atexit
import signal

# 检查并导入串口库
try:
    import serial
    import serial.tools.list_ports
    SERIAL_AVAILABLE = True
except ImportError:
    print("警告: pyserial 库未安装，串口功能将不可用")
    print("请运行: pip install pyserial")
    SERIAL_AVAILABLE = False

# 设置中文字体（按平台选择系统中已安装的字体，避免 Qt 启动时反复匹配缺失字体）
if os.name == 'nt':
    font_family = "Microsoft YaHei"
elif sys.platform == 'darwin':
    # macOS 原生中文字体，避免 "WenQuanYi Micro Hei" 缺失导致的启动开销与警告
    font_family = "PingFang SC"
else:
    font_family = "WenQuanYi Micro Hei"

pg.setConfigOption('background', 'w')
pg.setConfigOption('foreground', 'k')
pg.setConfigOption('antialias', True)


# ==================== 自定义图例控件 ====================
class CustomLegendWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        # 使用 objectName 选择器，避免 QWidget 选择器的 padding 级联到
        # 内部 QTableWidget / QHeaderView，导致表头与数据列错位
        self.setObjectName("legendWidget")
        self.setStyleSheet("""
            #legendWidget {
                background-color: rgba(255, 255, 255, 0.95);
                border-radius: 5px;
            }
        """)
        self.main_layout = QVBoxLayout(self)
        self.main_layout.setContentsMargins(4, 4, 4, 4)
        self.main_layout.setSpacing(2)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        # 表格：通道名称 | 当前值 | Max | Min | Avg | Std | T0 | T1 | T2 | T3
        self.table = QTableWidget(0, 10)
        self.table.setHorizontalHeaderLabels(["通道名称", "当前值", "Max", "Min", "Avg", "Std", "T0", "T1", "T2", "T3"])
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        # 通道名称列(0)固定较宽，其余数据列平均分布（等比拉伸）
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        self.table.setColumnWidth(0, 200)
        for c in range(1, 10):
            header.setSectionResizeMode(c, QHeaderView.ResizeMode.Stretch)
        # 表头与数据行等高，且每个表头文字水平居中，与数据单元格一致
        header.setFixedHeight(34)
        header.setMinimumSectionSize(30)
        header.setStyleSheet(
            "QHeaderView::section{background:#f0f0f0;border:1px solid #dddddd;"
            "font-size:11px;font-weight:bold;padding:0px 2px;}")
        for c in range(10):
            hi = self.table.horizontalHeaderItem(c)
            if hi is not None:
                hi.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.table.setStyleSheet(
            "QTableWidget{background:#ffffff;border:1px solid #dddddd;border-radius:4px;"
            "gridline-color:#eeeeee;font-size:11px;}"
            "QTableWidget::item{padding:0px;}")
        self.table.setRowHeight(0, 34)
        self.table.verticalHeader().setDefaultSectionSize(34)
        self.main_layout.addWidget(self.table, 1)

        self.device_items = {}        # device_id -> (name_lbl, color, reset_btn, row)
        self.device_colors = {}
        self.device_containers = {}   # device_id -> row (用于显示/隐藏)
        self.device_units = {}        # device_id -> unit 单位
        self.volatility_reset_times = {}
        # 记录每次批量重置后的avg和variance
        self.post_reset_records = {}  # device_id -> {'avg': val, 'var': val, 'recorded': bool}

    def add_device(self, device_id, device_name, color, unit='°C', auto_test=False):
        if device_id in self.device_items:
            r = self.device_items[device_id][3]
            self.device_colors[device_id] = color
            if unit:
                self.device_units[device_id] = unit
            self._set_name_cell(r, device_name, color)
            return

        row = self.table.rowCount()
        self.table.insertRow(row)
        self.table.setRowHeight(row, 36)

        # 通道名称列：色点 + 名称 + 重置按钮（cellWidget）
        name_cell = QWidget()
        name_layout = QHBoxLayout(name_cell)
        name_layout.setContentsMargins(4, 2, 4, 2)
        name_layout.setSpacing(4)
        color_lbl = QLabel()
        color_lbl.setFixedSize(12, 12)
        color_lbl.setStyleSheet(f"background-color: {color}; border-radius: 3px;")
        name_layout.addWidget(color_lbl)
        name_label = QLabel(device_name)
        name_label.setObjectName("nameLabel")
        name_label.setStyleSheet(
            f'font-family:"{font_family}";font-size:13px;font-weight:bold;color:#333333;')
        name_label.setWordWrap(True)
        name_layout.addWidget(name_label, 1)
        name_layout.addStretch()
        reset_btn = QPushButton("重置")
        reset_btn.setFixedSize(38, 18)
        reset_btn.setStyleSheet(
            'QPushButton{font-size:10px;background:#f0f0f0;border:1px solid #cccccc;'
            'border-radius:3px;color:#333333;padding:1px;}'
            'QPushButton:hover{background:#e0e0e0;}')
        name_layout.addWidget(reset_btn)
        self.table.setCellWidget(row, 0, name_cell)

        self.table.setItem(row, 1, QTableWidgetItem("--"))  # 当前值
        self.table.setItem(row, 2, QTableWidgetItem("--"))  # Max
        self.table.setItem(row, 3, QTableWidgetItem("--"))  # Min
        self.table.setItem(row, 4, QTableWidgetItem("--"))  # Avg
        self.table.setItem(row, 5, QTableWidgetItem("--"))  # Std
        self.table.setItem(row, 6, QTableWidgetItem("--"))  # T0
        self.table.setItem(row, 7, QTableWidgetItem("--"))  # T1
        self.table.setItem(row, 8, QTableWidgetItem("--"))  # T2
        self.table.setItem(row, 9, QTableWidgetItem("--"))  # T3
        for c in range(1, 10):
            item = self.table.item(row, c)
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

        self.device_items[device_id] = (name_label, color, reset_btn, row)
        self.device_colors[device_id] = color
        self.device_containers[device_id] = row
        self.device_units[device_id] = unit

        reset_btn.clicked.connect(lambda checked, idx=device_id: self.reset_volatility(idx))

    def _unit_for(self, device_id, unit):
        """确定显示单位：优先用传入单位，否则用设备记录的单位"""
        if unit:
            return unit
        return self.device_units.get(device_id, '°C')

    def _set_name_cell(self, row, name, color):
        """更新通道名称单元格内容（色点+名称+重置按钮）"""
        cell = self.table.cellWidget(row, 0)
        if cell is not None:
            name_lbl = cell.findChild(QLabel, "nameLabel")
            if name_lbl is not None:
                name_lbl.setText(name)

    def update_temperature(self, device_id, temperature, unit='°C'):
        # 当前温度实时显示在"当前值"列（列1，Max 左侧），不加单位
        if device_id in self.device_items:
            row = self.device_items[device_id][3]
            if temperature is not None:
                self.table.item(row, 1).setText(f"{temperature:.2f}")

    def update_device_name(self, device_id, new_name):
        if device_id in self.device_items:
            row = self.device_items[device_id][3]
            self._set_name_cell(row, new_name, self.device_colors.get(device_id, "#333333"))

    def update_volatility(self, device_id, volatility, duration_min=None, unit='°C'):
        if device_id in self.device_items:
            row = self.device_items[device_id][3]
            if volatility is None:
                self.table.item(row, 5).setText("--")
            else:
                # 减小显示位数，单行紧凑显示（2位小数，去掉时长后缀）
                self.table.item(row, 5).setText(f"{volatility:.2f}")

    def update_stats(self, device_id, min_val=None, max_val=None, avg_val=None, unit=''):
        if device_id in self.device_items:
            row = self.device_items[device_id][3]
            self.table.item(row, 3).setText("--" if min_val is None else f"{min_val:.2f}")  # Min
            self.table.item(row, 2).setText("--" if max_val is None else f"{max_val:.2f}")  # Max
            if avg_val is not None:
                self.table.item(row, 4).setText(f"{avg_val:.2f}")  # Avg

    def reset_volatility(self, device_id):
        self.volatility_reset_times[device_id] = datetime.now()
        if device_id in self.device_items:
            row = self.device_items[device_id][3]
            self.table.item(row, 2).setText("--")  # Max
            self.table.item(row, 3).setText("--")  # Min
            self.table.item(row, 4).setText("--")  # Avg
            self.table.item(row, 5).setText("--")  # Std

    def get_reset_time(self, device_id):
        return self.volatility_reset_times.get(device_id)

    def update_post_reset_record(self, device_id, avg_val, var_val, unit='°C'):
        """更新重置后记录（表格模式下保留状态存储，不额外显示）"""
        if avg_val is not None and var_val is not None:
            self.post_reset_records[device_id] = {'avg': avg_val, 'var': var_val, 'recorded': True}
        else:
            self.post_reset_records[device_id] = {'avg': None, 'var': None, 'recorded': False}

    def update_auto_test(self, device_id, t0=None, t1=None, t2=None, t3=None,
                         vol1=None, vol2=None, avg1=None, avg2=None, t0t1=None):
        """更新自动检测信息：表格中显示 T0/T1/T2/T3 列（时间，单位 min）"""
        if device_id in self.device_items:
            row = self.device_items[device_id][3]
            self.table.item(row, 6).setText("--" if t0 is None else f"{t0:.2f}")  # T0
            if t1 is not None:
                self.table.item(row, 7).setText(f"{t1:.2f}")  # T1
            elif t0t1 is not None and t0 is not None:
                self.table.item(row, 7).setText(f"{t0 + t0t1:.2f}")
            self.table.item(row, 8).setText("--" if t2 is None else f"{t2:.2f}")  # T2
            self.table.item(row, 9).setText("--" if t3 is None else f"{t3:.2f}")  # T3

    def clear_auto_test(self):
        """清空所有自动检测信息"""
        for device_id in list(self.device_items.keys()):
            row = self.device_items[device_id][3]
            self.table.item(row, 6).setText("--")  # T0
            self.table.item(row, 7).setText("--")  # T1
            self.table.item(row, 8).setText("--")  # T2
            self.table.item(row, 9).setText("--")  # T3

    def clear_post_reset_records(self):
        """清空所有重置后记录"""
        for device_id in list(self.post_reset_records.keys()):
            self.post_reset_records[device_id] = {'avg': None, 'var': None, 'recorded': False}

    def set_post_reset_visible(self, visible):
        """兼容接口：表格模式下无独立重置记录行，空操作"""
        pass

    def set_device_visible(self, device_id, visible):
        """根据启用状态显示/隐藏指定设备的整行"""
        if device_id in self.device_containers:
            self.table.setRowHidden(self.device_containers[device_id], not visible)

    def append_terminal(self, text):
        """终端信息区域已取消，此方法为空操作（保留兼容调用）"""
        pass

    def clear_terminal(self):
        """终端信息区域已取消，此方法为空操作（保留兼容调用）"""
        pass

    def clear(self):
        self.table.setRowCount(0)
        self.device_items.clear()
        self.device_containers.clear()
        self.volatility_reset_times.clear()
        self.post_reset_records.clear()


# ==================== 串口调试窗口 ====================
class SerialDebugDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("通讯调试 - 详细通讯信息")
        self.resize(1000, 600)
        self.setWindowFlags(Qt.WindowType.Window)
        self._geom_restored = False
        layout = QVBoxLayout(self)

        control_layout = QHBoxLayout()
        self.clear_btn = QPushButton("清空显示")
        self.clear_btn.clicked.connect(self.clear_text)
        control_layout.addWidget(self.clear_btn)

        self.auto_scroll_cb = QCheckBox("自动滚动")
        self.auto_scroll_cb.setChecked(True)
        control_layout.addWidget(self.auto_scroll_cb)

        self.pause_btn = QPushButton("暂停更新")
        self.pause_btn.setCheckable(True)
        self.pause_btn.clicked.connect(self.toggle_pause)
        control_layout.addWidget(self.pause_btn)
        control_layout.addStretch()
        layout.addLayout(control_layout)

        self.text_edit = QTextEdit()
        self.text_edit.setReadOnly(True)
        self.text_edit.setFont(QFont("Courier New", 10))
        layout.addWidget(self.text_edit)
        self.paused = False

    def toggle_pause(self):
        self.paused = self.pause_btn.isChecked()
        if self.paused:
            self.pause_btn.setText("恢复更新")
        else:
            self.pause_btn.setText("暂停更新")

    def append_comm_detail(self, device_id, command, raw_response, parsed_value, success, error_msg=""):
        if self.paused:
            return
        timestamp = datetime.now().strftime("%H:%M:%S.%f")[:-3]
        device_name = f"设备{device_id+1}"
        if success:
            status_color = "#00AA00"
            status_text = "✓ 成功"
        else:
            status_color = "#FF0000"
            status_text = "✗ 失败"

        text = f"""
┌─────────────────────────────────────────────────────────────
│ [{timestamp}] <span style="color:{status_color};font-weight:bold;">{device_name} - {status_text}</span>
├─────────────────────────────────────────────────────────────
│ 发送命令: <span style="color:#0000AA;">{repr(command)}</span>
│ 原始响应: <span style="color:#AA00AA;">{repr(raw_response)}</span>
│ 解析结果: <span style="color:#00AA00;font-weight:bold;">{f"{parsed_value:.4f}" if parsed_value is not None else "无效数据"}</span>
"""
        if error_msg:
            text += f'│ 错误信息: <span style="color:#FF0000;">{error_msg}</span>\n'
        text += "└─────────────────────────────────────────────────────────────\n"
        self.text_edit.append(text)
        if self.auto_scroll_cb.isChecked():
            scrollbar = self.text_edit.verticalScrollBar()
            scrollbar.setValue(scrollbar.maximum())

    def clear_text(self):
        self.text_edit.clear()

    def showEvent(self, event):
        """窗口首次显示时恢复上次位置和大小（延迟执行，确保布局完成）"""
        super().showEvent(event)
        if not getattr(self, '_geom_restored', False):
            QTimer.singleShot(50, self._restore_geometry)
            self._geom_restored = True

    def _restore_geometry(self):
        geom_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "window_geometry.json")
        try:
            if os.path.exists(geom_file):
                with open(geom_file, 'r') as f:
                    data = json.load(f)
                if 'serial_debug' in data and isinstance(data['serial_debug'], dict):
                    g = data['serial_debug']
                    x, y, w, h = g.get('x', 0), g.get('y', 0), g.get('w', 0), g.get('h', 0)
                    if w > 0 and h > 0:
                        self.move(x, y)
                        self.resize(w, h)
        except Exception:
            pass

    def resizeEvent(self, event):
        """窗口大小/位置变化时实时保存（立即固化为默认）"""
        super().resizeEvent(event)
        geom_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "window_geometry.json")
        try:
            data = {}
            if os.path.exists(geom_file):
                with open(geom_file, 'r') as f:
                    data = json.load(f)
            g = self.geometry()
            data['serial_debug'] = {'x': g.x(), 'y': g.y(), 'w': g.width(), 'h': g.height()}
            with open(geom_file, 'w') as f:
                json.dump(data, f, indent=2)
        except Exception:
            pass

    def closeEvent(self, event):
        # 保存窗口位置与大小到 JSON 文件（直接存 x/y/w/h）
        geom_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "window_geometry.json")
        try:
            data = {}
            if os.path.exists(geom_file):
                with open(geom_file, 'r') as f:
                    data = json.load(f)
            g = self.geometry()
            data['serial_debug'] = {'x': g.x(), 'y': g.y(), 'w': g.width(), 'h': g.height()}
            with open(geom_file, 'w') as f:
                json.dump(data, f, indent=2)
        except Exception:
            pass
        event.accept()


# ==================== 串口管理器 ====================
class SharedSerialManager:
    _instances = {}
    _lock = threading.Lock()

    def __new__(cls, port, baudrate):
        key = f"{port}_{baudrate}"
        with cls._lock:
            if key not in cls._instances:
                instance = super().__new__(cls)
                instance.port = port
                instance.baudrate = int(baudrate)
                instance.serial_conn = None
                instance.connected = False
                instance.lock = threading.Lock()
                instance._reconnect_lock = threading.Lock()  # 防多线程同时重连
                instance._reconnecting = False
                cls._instances[key] = instance
            return cls._instances[key]

    def connect(self):
        with self.lock:
            if not self.connected:
                if not SERIAL_AVAILABLE:
                    return False
                if self.serial_conn:
                    try:
                        self.serial_conn.close()
                    except:
                        pass
                    self.serial_conn = None
                try:
                    self.serial_conn = serial.Serial(
                        port=self.port,
                        baudrate=self.baudrate,
                        timeout=1.0,
                        write_timeout=1.0,
                        bytesize=serial.EIGHTBITS,
                        parity=serial.PARITY_NONE,
                        stopbits=serial.STOPBITS_ONE
                    )
                    self.connected = True
                    return True
                except Exception as e:
                    print(f"串口{self.port}连接异常: {e}")
                    return False
            return True

    def disconnect(self):
        """关闭串口（不获取锁，可中断阻塞在 send_command 中的线程）"""
        serial_conn = self.serial_conn
        self.serial_conn = None
        self.connected = False
        if serial_conn:
            try:
                serial_conn.close()
            except:
                pass

    def ensure_reconnected(self, max_retries=5):
        """安全重连：使用 _reconnect_lock 防止多线程同时重连
           如果其他线程已恢复连接，立即返回成功"""
        with self._reconnect_lock:
            with self.lock:
                if self.connected and self.serial_conn:
                    try:
                        self.serial_conn.reset_input_buffer()
                        self.serial_conn.write(b"")
                    except:
                        self.connected = False
                    if self.connected:
                        return True
            for attempt in range(max_retries):
                if self.connect():
                    return True
                time.sleep(2.0)
            return False

    def flush_buffers(self):
        """轻量级清空缓冲区，不关闭串口，适合超时恢复"""
        with self.lock:
            if self.connected and self.serial_conn:
                try:
                    self.serial_conn.reset_input_buffer()
                    self.serial_conn.reset_output_buffer()
                except:
                    pass

    def test_connection(self):
        """测试串口是否仍然可用（不关闭端口）"""
        with self.lock:
            if self.connected and self.serial_conn:
                try:
                    self.serial_conn.reset_input_buffer()
                    return True
                except:
                    self.connected = False
            return False

    def send_command(self, command, timeout=2.0):
        with self.lock:
            if not self.connected or self.serial_conn is None:
                return None
            conn = self.serial_conn  # 本地引用，防止 disconnect() 并发设 None
            # 其他线程正在重连时，直接返回 None 让当前线程下次再试
            if self._reconnect_lock.locked():
                return None
            try:
                conn.reset_input_buffer()
                conn.reset_output_buffer()
                conn.write(command.encode('utf-8'))
                conn.flush()
                response = ""
                start_time = time.time()
                while time.time() - start_time < timeout:
                    if conn.in_waiting > 0:
                        time.sleep(0.05)
                        data = conn.read(conn.in_waiting)
                        response += data.decode('utf-8', errors='ignore')
                        if response:
                            time.sleep(0.05)
                            if conn.in_waiting > 0:
                                data = conn.read(conn.in_waiting)
                                response += data.decode('utf-8', errors='ignore')
                            break
                    time.sleep(0.01)
                return response.strip() if response else None
            except Exception as e:
                # 串口被强制关闭导致的错误（句柄无效）是预期行为，不打印避免干扰
                if "句柄无效" not in str(e):
                    print(f"串口{self.port}读取异常: {e}")
                return None


# ==================== 以太网管理器 ====================
class EthernetManager:
    def __init__(self, host, port):
        self.host = host
        try:
            self.port = int(port) if port and str(port).strip() else 8000
        except (ValueError, TypeError):
            self.port = 8000
        self.sock = None
        self.connected = False
        self.lock = threading.Lock()

    def connect(self):
        with self.lock:
            if not self.connected:
                try:
                    self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                    try:
                        self.sock.setsockopt(socket.IPPROTO_TCP, socket.TCP_NODELAY, 1)
                    except:
                        pass
                    self.sock.settimeout(5.0)
                    try:
                        self.sock.connect((self.host, self.port))
                        self.connected = True
                        self.sock.settimeout(5.0)
                        return True
                    except (socket.timeout, ConnectionRefusedError, OSError) as e:
                        try:
                            self.sock.close()
                            time.sleep(0.5)
                            self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                            self.sock.settimeout(5.0)
                            self.sock.connect((self.host, self.port))
                            self.connected = True
                            self.sock.settimeout(5.0)
                            return True
                        except Exception as retry_e:
                            error_type = type(retry_e).__name__
                            error_msg = str(retry_e)
                            print(f"LAN连接失败: {error_type} - {error_msg}")
                            return False
                except Exception as e:
                    error_type = type(e).__name__
                    error_msg = str(e)
                    print(f"LAN连接异常: {error_type} - {error_msg}")
                    if self.sock:
                        try:
                            self.sock.close()
                        except:
                            pass
                        self.sock = None
                    return False
            return True

    def disconnect(self):
        with self.lock:
            if self.connected and self.sock:
                try:
                    self.sock.close()
                except:
                    pass
            self.connected = False

    def send_command(self, command, timeout=3.0):
        with self.lock:
            if not self.connected:
                return None
            try:
                self.sock.settimeout(timeout)
                self.sock.sendall(command.encode('utf-8'))
                time.sleep(0.1)
                response = b""
                while True:
                    try:
                        chunk = self.sock.recv(4096)
                        if not chunk:
                            break
                        response += chunk
                        if b'\n' in chunk or len(chunk) < 4096:
                            try:
                                self.sock.settimeout(0.1)
                                chunk2 = self.sock.recv(4096)
                                if chunk2:
                                    response += chunk2
                            except:
                                pass
                            break
                    except socket.timeout:
                        break
                return response.decode('utf-8', errors='ignore').strip()
            except socket.timeout:
                return None
            except Exception as e:
                # 串口被强制关闭导致的错误（句柄无效）是预期行为，不打印避免干扰
                if "句柄无效" not in str(e):
                    print(f"串口{self.port}读取异常: {e}")
                return None


# ==================== 设备线程 ====================
class DeviceThread(QThread):
    data_received = pyqtSignal(dict, int)
    connection_status = pyqtSignal(int, bool, str)
    debug_info = pyqtSignal(int, str, str, object, bool, str)

    def __init__(self, device_id, config):
        super().__init__()
        self.device_id = device_id
        self.config = config
        self.running = False
        self.paused = False
        self.manager = None
        self.read_interval = config.get('read_interval', 1.0)

    def run(self):
        self.running = True
        conn_type = self.config.get('connection', 'serial')
        if conn_type == 'lan':
            host = self.config.get('host', '192.168.0.182')
            port = self.config.get('lan_port', 8000)
            system_platform = platform.platform()
            print(f"设备{self.device_id+1}: 开始LAN连接 - IP:{host}, 端口:{port}, 系统:{system_platform}")
            self.manager = EthernetManager(host, port)
        else:
            port = self.config.get('port', '')
            baudrate = self.config.get('baudrate', '9600')
            self.manager = SharedSerialManager(port, baudrate)

        connected = False
        for attempt in range(3):
            if not self.running:
                break
            if self.manager.connect():
                connected = True
                break
            if attempt < 2:
                print(f"设备{self.device_id+1}: 初始连接第{attempt+1}次失败，1秒后重试...")
                time.sleep(1.0)

        if not connected:
            error_msg = f"连接失败 - 系统:{platform.platform()}"
            self.connection_status.emit(self.device_id, False, error_msg)
            self.debug_info.emit(self.device_id, self.config.get('read_command', ''), "", None, False, error_msg)
            return

        self.connection_status.emit(self.device_id, True, "已连接")

        read_command = self.config.get('read_command', '')
        if read_command:
            read_command = read_command.replace('\\r', '\r').replace('\\n', '\n')
            # 自动补充 SCPI 标准换行终止符（\r\n），兼容 Fluke 9250 等严格设备
            read_command = read_command.rstrip('\r\n') + '\r\n'

        serial_exceptions = 0
        last_read_time = 0
        reconnect_count = 0  # 累计重连次数，超过上限后停止重连

        while self.running:
            if not self.paused and read_command:
                current_time = time.time()
                if current_time - last_read_time >= self.read_interval:
                    try:
                        response = self.manager.send_command(read_command, 10.0)
                        parsed_value = None
                        success = False
                        error_msg = ""

                        if response:
                            serial_exceptions = 0
                            parsed_data = self._parse_response(response)
                            if parsed_data:
                                parsed_value = parsed_data.get('value')
                                success = True
                                self.data_received.emit(parsed_data, self.device_id)
                            else:
                                error_msg = "解析失败：数据格式错误或数值超出范围"
                                self.data_received.emit({'value': None}, self.device_id)
                        else:
                            error_msg = "无响应或超时（正常情况）"
                            self.data_received.emit({'value': None}, self.device_id)

                        self.debug_info.emit(self.device_id, read_command, response, parsed_value, success, error_msg)
                        last_read_time = time.time()
                    except Exception as e:
                        serial_exceptions += 1
                        self.debug_info.emit(self.device_id, read_command, "", None, False, f"异常: {str(e)}")
                        self.data_received.emit({'value': None}, self.device_id)
                        last_read_time = time.time()

                    # 仅串口异常触发重连（无响应/超时是正常情况，不计数）
                    if serial_exceptions >= 3:
                        reconnect_count += 1
                        if reconnect_count > 3:
                            print(f"设备{self.device_id+1}: 已累计{reconnect_count-1}次重连均失败，停止重连")
                            self.connection_status.emit(self.device_id, False, "通信异常，已停止重连")
                            self.running = False
                            break
                        print(f"设备{self.device_id+1}: 第{reconnect_count}次重连（连续{serial_exceptions}次串口异常）...")
                        self.connection_status.emit(self.device_id, False, "串口异常，重连中...")
                        self.manager.disconnect()
                        if self.manager.ensure_reconnected(5):
                            serial_exceptions = 0
                            self.connection_status.emit(self.device_id, True, "已重连")
                            print(f"设备{self.device_id+1}: 第{reconnect_count}次重连成功")
                        else:
                            serial_exceptions = 0
                            self.connection_status.emit(self.device_id, False, "重连失败，继续尝试...")
                            print(f"设备{self.device_id+1}: 第{reconnect_count}次重连失败，继续尝试")
            time.sleep(0.01)

        # 不调用 manager.disconnect()——共享串口，由 stop_collection 统一清理
        self.connection_status.emit(self.device_id, False, "已断开")

    def _parse_response(self, data):
        """解析响应数据，支持ConST1210多物理量格式
        - MEASure:TEMPerature? → 取第1个逗号分隔值（控制温度）
        - MEASure:TEMPerature? 2 → 取第2个逗号分隔值（内部温度/Main）
        - MEASure:TEMPerature? 3 → 取第3个逗号分隔值（Sec传感器温度）
        - MEASure:CONTrol? 3 → 取第3个逗号分隔值（parts[3]，加热功率）
        - MEASure:CONTrol? 4 → 取第4个逗号分隔值（parts[4]，风扇功率）
        - 其他命令 → 回退到原有逻辑
        """
        try:
            data = data.strip()
            if not data:
                return None

            command = self.config.get('read_command', '')
            cmd_upper = command.upper().replace('\r', '').replace('\n', '').strip()

            if ',' in data:
                parts = data.split(',')

                # MEASure[:SCALar]:CONTrol? N → 取第N个值
                if 'CONTROL' in cmd_upper:
                    # 解析命令末尾的数字，决定取parts的第几个值
                    control_idx = 5  # 默认取parts[5]（风扇功率）
                    idx_match = re.search(r'CONTROL\S*\s+(\d+)', cmd_upper)
                    if idx_match:
                        control_idx = int(idx_match.group(1))
                    if len(parts) >= control_idx:
                        try:
                            value = float(parts[control_idx - 1])
                            if control_idx == 4:
                                return {'value': value, 'quantity_type': 'HeatingPower', 'unit': ''}
                            else:
                                return {'value': value, 'quantity_type': 'FanPower', 'unit': ''}
                        except ValueError:
                            pass

                # MEASure[:SCALar][:TEMPerature]? [N] → 取第N个逗号分隔值
                #   默认 N=1（控制温度），N=2（内部温度/Main），N=3（Sec传感器）
                elif 'TEMPERATURE' in cmd_upper:
                    temp_idx = 1  # 默认取parts[0]（控制温度）
                    idx_match = re.search(r'TEMPERATURE\S*\s+(\d+)', cmd_upper)
                    if idx_match:
                        temp_idx = int(idx_match.group(1))
                    if len(parts) >= temp_idx:
                        try:
                            value = float(parts[temp_idx - 1])
                            return {'value': value, 'quantity_type': 'Temperature', 'unit': '°C'}
                        except ValueError:
                            pass

                # SOUR:SENS:DATA? TEMP1/TEMP2 → Fluke 9250 传感器数据
                #   响应格式: +25.500 或 +25.500,status,...
                #   取第1个逗号分隔值（温度数值）
                elif 'SOUR' in cmd_upper and 'SENS' in cmd_upper and 'DATA' in cmd_upper:
                    try:
                        value = float(parts[0])
                        return {'value': value, 'quantity_type': 'Temperature', 'unit': '°C'}
                    except ValueError:
                        pass
                    # fallback: 尝试从 parts[0] 提取数值
                    numeric_match = re.search(r'[+-]?\d+\.?\d*', parts[0])
                    if numeric_match:
                        try:
                            value = float(numeric_match.group())
                            return {'value': value, 'quantity_type': 'Temperature', 'unit': '°C'}
                        except ValueError:
                            pass

                # OUTP:DATA? → 占空比数据（无范围限制）
                elif 'OUTP' in cmd_upper and 'DATA' in cmd_upper:
                    if len(parts) >= 2:
                        try:
                            value = float(parts[1])
                            return {'value': value, 'quantity_type': 'DutyCycle', 'unit': ''}
                        except ValueError:
                            pass
                    try:
                        value = float(parts[0])
                        return {'value': value, 'quantity_type': 'DutyCycle', 'unit': ''}
                    except ValueError:
                        pass

                # 默认回退：尝试取第2个值
                if len(parts) >= 2:
                    try:
                        temp = float(parts[1])
                        return {'value': temp, 'quantity_type': 'Temperature', 'unit': '°C'}
                    except ValueError:
                        pass

            # OUTP:DATA? → 占空比数据（无范围限制，纯数值响应）
            if 'OUTP' in cmd_upper and 'DATA' in cmd_upper:
                try:
                    value = float(data)
                    return {'value': value, 'quantity_type': 'DutyCycle', 'unit': ''}
                except ValueError:
                    pass
                numeric_match = re.search(r'[+-]?\d+\.?\d*', data)
                if numeric_match:
                    try:
                        value = float(numeric_match.group())
                        return {'value': value, 'quantity_type': 'DutyCycle', 'unit': ''}
                    except ValueError:
                        pass

            # 尝试解析为纯数值（无范围限制）
            try:
                temp = float(data)
                return {'value': temp, 'quantity_type': 'Temperature', 'unit': '°C'}
            except ValueError:
                pass

            # 正则匹配（无范围限制）
            numeric_match = re.search(r'[+-]?\d+\.?\d*', data)
            if numeric_match:
                try:
                    temp = float(numeric_match.group())
                    return {'value': temp, 'quantity_type': 'Temperature', 'unit': '°C'}
                except ValueError:
                    pass
        except Exception:
            pass

        return None

    def pause(self):
        self.paused = True

    def resume(self):
        self.paused = False

    def stop(self):
        self.running = False
        self.wait(1000)


# ==================== 串口下拉框（点击弹出，白底黑字） ====================
class PortComboBox(QComboBox):
    """串口选择下拉框：非可编辑模式，点击任意位置即刷新串口列表并弹出原生下拉菜单。
       显式为下拉列表设置白色背景 + 黑色文字，避免默认深色/黑色主题看不清选项。
       配置加载时若端口不在当前列表中，会加入列表首位以便显示。"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setEditable(False)
        self._apply_popup_style()

    def _apply_popup_style(self):
        """给下拉列表视图设置白色背景 + 黑色文字，保证选项清晰可见"""
        self.setStyleSheet("""
            QComboBox {
                background-color: #ffffff; color: #000000;
                border: 1px solid #cccccc; border-radius: 3px; padding: 2px 5px;
            }
            QComboBox QAbstractItemView {
                background-color: #ffffff; color: #000000;
                selection-background-color: #2196F3; selection-color: #ffffff;
                border: 1px solid #cccccc; outline: 0;
            }
            QComboBox::drop-down { border: none; width: 18px; }
        """)
        view = self.view()
        if view is not None:
            view.setStyleSheet(
                "QListView { background-color: #ffffff; color: #000000; "
                "selection-background-color: #2196F3; selection-color: #ffffff; }"
            )

    def _get_ports(self):
        """获取系统当前可用串口列表"""
        if not SERIAL_AVAILABLE:
            return []
        try:
            return [p.device for p in serial.tools.list_ports.comports()]
        except Exception:
            return []

    def _refresh_and_show(self):
        """刷新串口列表（保留当前选中值），再弹出原生下拉菜单"""
        current = self.currentText()
        port_list = self._get_ports()
        old_block = self.blockSignals(True)
        try:
            self.clear()
            self.addItems(port_list if port_list else [''])
            if current:
                idx = self.findText(current)
                if idx >= 0:
                    self.setCurrentIndex(idx)
                elif not self.isEditable():
                    # 非可编辑模式：把当前端口加到列表首位以便显示
                    self.insertItem(0, current)
                    self.setCurrentIndex(0)
                else:
                    self.setCurrentText(current)
        finally:
            self.blockSignals(old_block)
        self.showPopup()

    def mousePressEvent(self, event):
        # 点击任意位置（含编辑区/箭头）都刷新串口列表并弹出原生下拉菜单
        self._refresh_and_show()


# ==================== 温度查询线程 ====================
class TemperatureQueryThread(QThread):
    """后台查询温度源 Main/Sec 温度，不阻塞主线程

    支持:
      - Fluke 9250: SOUR:SENS:DATA? TEMP1 / TEMP2
    """
    temp1_ready = pyqtSignal(float)
    temp2_ready = pyqtSignal(float)

    def __init__(self, manager, device_type='Fluke 9250'):
        super().__init__()
        self.manager = manager
        self.device_type = device_type
        self.running = True

    def _parse_first_number(self, resp):
        """从响应中提取第一个数字，支持 +25.5 / +25.5,status / a,b,c 等格式"""
        if not resp:
            return None
        # 逗号分隔时先取段
        for part in resp.split(','):
            m = re.search(r'[+-]?\d+\.?\d*', part)
            if m:
                try:
                    return float(m.group())
                except ValueError:
                    pass
        return None

    def run(self):
        cmd1 = "SOUR:SENS:DATA? TEMP1\r\n"
        cmd2 = "SOUR:SENS:DATA? TEMP2\r\n"
        label = self.device_type
        print(f"[TempQuery {label}] 线程已启动，命令1={cmd1!r}, 命令2={cmd2!r}")

        while self.running:
            try:
                resp = self.manager.send_command(cmd1, timeout=1.0)
                # 问询响应不打印
                temp = self._parse_first_number(resp)
                if temp is not None:
                    self.temp1_ready.emit(temp)
            except Exception as e:
                print(f"[TempQuery {label}] cmd1 error: {e}")

            try:
                resp = self.manager.send_command(cmd2, timeout=1.0)
                # 问询响应不打印
                temp = self._parse_first_number(resp)
                if temp is not None:
                    self.temp2_ready.emit(temp)
            except Exception as e:
                print(f"[TempQuery {label}] cmd2 error: {e}")

            # 约5秒刷新一次
            for _ in range(50):
                if not self.running:
                    break
                time.sleep(0.1)

        print(f"[TempQuery {label}] 线程已停止")

    def stop(self):
        self.running = False
        self.wait(2000)


class Const1210QueryThread(QThread):
    """后台查询 Const 1210 控制源温度，不阻塞主线程"""
    temp_ready = pyqtSignal(float)

    def __init__(self, manager):
        super().__init__()
        self.manager = manager
        self.running = True

    def run(self):
        print("[Const1210Query] 线程已启动，开始查询温度...")
        while self.running:
            try:
                resp = self.manager.send_command("MEASure:TEMPerature?\r\n", timeout=1.0)
                if resp:
                    temp = float(resp.strip().split(',')[0])
                    self.temp_ready.emit(temp)
                else:
                    print("[Const1210Query] 未收到响应（可能超时或设备未就绪）")
            except Exception as e:
                print(f"[Const1210Query] 异常: {e}")
            for _ in range(50):  # 约5秒
                if not self.running:
                    break
                time.sleep(0.1)
        print("[Const1210Query] 线程已停止")

    def stop(self):
        self.running = False
        self.wait(2000)


# ==================== 可清空数值输入框 ====================
class NullableSpinBox(QDoubleSpinBox):
    """可清空的数值输入框：留空表示使用温度源默认值（不发送该条命令）。
       value() 在留空时返回 None。"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._is_none = False

    def value(self):
        """留空返回 None，否则返回数值"""
        return None if self._is_none else super().value()

    def setValue(self, val):
        """传入 None 或空字符串则清空（回到默认值状态）"""
        if val is None or (isinstance(val, str) and val.strip() == ''):
            self._is_none = True
            self.clear()
        else:
            self._is_none = False
            super().setValue(float(val))

    def setPlaceholderText(self, text):
        le = self.lineEdit()
        if le is not None:
            le.setPlaceholderText(text)

    def validate(self, text, pos):
        # 允许输入框内容被完全删除
        if text.strip() == '':
            return (QValidator.State.Acceptable, text, pos)
        return super().validate(text, pos)

    def valueFromText(self, text):
        self._is_none = (text.strip() == '')
        if self._is_none:
            return self.minimum()
        return super().valueFromText(text)

    def textFromValue(self, val):
        # 留空状态不回填数值文本，保持空白
        return '' if self._is_none else super().textFromValue(val)

    def stepBy(self, steps):
        if self._is_none:
            self._is_none = False
            super().setValue(self.minimum())
        super().stepBy(steps)


# ==================== 可折叠组框 ====================
class CollapsibleGroupBox(QWidget):
    """带折叠功能的组框，点击标题切换展开/折叠"""
    toggled = pyqtSignal(bool)  # collapsed 状态变化时发射

    def __init__(self, title, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        self._collapsed = False

        self.toggle_btn = QPushButton(f"▼ {title}")
        self.toggle_btn.setStyleSheet("""
            QPushButton{text-align:left;font-size:14px;font-weight:bold;
                        background:#e8e8e8;border:1px solid #ccc;
                        border-radius:4px;padding:6px 10px;}
            QPushButton:hover{background:#d0d0d0;}
        """)
        self.toggle_btn.clicked.connect(self._toggle)
        layout.addWidget(self.toggle_btn)

        self.content = QWidget()
        self.content_layout = QVBoxLayout(self.content)
        self.content_layout.setContentsMargins(4, 4, 4, 4)
        layout.addWidget(self.content)

    def contentLayout(self):
        return self.content_layout

    def _toggle(self):
        self._collapsed = not self._collapsed
        self.content.setVisible(not self._collapsed)
        arrow = "▶" if self._collapsed else "▼"
        self.toggle_btn.setText(f"{arrow} {self.toggle_btn.text()[1:].strip()}")
        self.toggled.emit(self._collapsed)

    def set_collapsed(self, collapsed):
        self._collapsed = collapsed
        self.content.setVisible(not collapsed)
        arrow = "▶" if collapsed else "▼"
        self.toggle_btn.setText(f"{arrow} {self.toggle_btn.text()[1:].strip()}")
        self.toggled.emit(self._collapsed)

    def is_collapsed(self):
        return self._collapsed


# ==================== 可拖动排序的设备行控件 ====================
class DragDropRowWidget(QWidget):
    """支持拖拽排序的设备行容器，左侧提供拖拽手柄"""
    reorder_requested = pyqtSignal(int, int)  # source_dev_id, target_dev_id

    def __init__(self, device_id, parent=None):
        super().__init__(parent)
        self.device_id = device_id
        self.setAcceptDrops(True)
        self._drag_start_pos = None

        main_layout = QHBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(2)

        # 拖拽手柄
        self.drag_handle = QLabel("⠿")
        self.drag_handle.setFixedWidth(18)
        self.drag_handle.setCursor(Qt.CursorShape.OpenHandCursor)
        self.drag_handle.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.drag_handle.setStyleSheet("color:#888;font-size:14px;")
        main_layout.addWidget(self.drag_handle)

        # 内容区域（layout 由 _create_device_row_widgets 设置）
        self.content_widget = QWidget()
        main_layout.addWidget(self.content_widget, 1)

        self.drag_handle.mousePressEvent = lambda e: self._handle_press(e)
        self.drag_handle.mouseMoveEvent = lambda e: self._handle_move(e)

    def contentWidget(self):
        return self.content_widget

    def _handle_press(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self._drag_start_pos = event.position().toPoint()
        QLabel.mousePressEvent(self.drag_handle, event)

    def _handle_move(self, event):
        if not (event.buttons() & Qt.MouseButton.LeftButton) or self._drag_start_pos is None:
            return
        if (event.position().toPoint() - self._drag_start_pos).manhattanLength() < QApplication.startDragDistance():
            return
        drag = QDrag(self)
        mime = QMimeData()
        mime.setData('application/x-device-row', str(self.device_id).encode())
        drag.setMimeData(mime)
        pixmap = self.grab()
        drag.setPixmap(pixmap)
        drag.exec(Qt.DropAction.MoveAction)

    def dragEnterEvent(self, event):
        if event.mimeData().hasFormat('application/x-device-row'):
            event.acceptProposedAction()

    def dropEvent(self, event):
        if event.mimeData().hasFormat('application/x-device-row'):
            source_id = int(event.mimeData().data('application/x-device-row').data().decode())
            if source_id != self.device_id:
                self.reorder_requested.emit(source_id, self.device_id)
            event.acceptProposedAction()


# ==================== 主程序 ====================
class DataCollectorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"T-cal_tester v{APP_VERSION}")
        self.config_file = 'serial_collector_config.json'

        self.devices = [
            {'enabled': False, 'connection': 'serial', 'port': '', 'baudrate': '9600', 'name': '设备1', 'read_command': '', 'host': '', 'lan_port': '', 'auto_test': False, 'curve_visible': True},
            {'enabled': False, 'connection': 'serial', 'port': '', 'baudrate': '9600', 'name': '设备2', 'read_command': '', 'host': '', 'lan_port': '', 'auto_test': False, 'curve_visible': True},
            {'enabled': False, 'connection': 'serial', 'port': '', 'baudrate': '9600', 'name': '设备3', 'read_command': '', 'host': '', 'lan_port': '', 'auto_test': False, 'curve_visible': True},
            {'enabled': False, 'connection': 'serial', 'port': '', 'baudrate': '9600', 'name': '设备4', 'read_command': '', 'host': '', 'lan_port': '', 'auto_test': False, 'curve_visible': True},
            {'enabled': False, 'connection': 'lan', 'port': '', 'baudrate': '9600', 'name': 'ConST1210 CH1', 'read_command': 'MEASure:TEMPerature?\r\n', 'host': '192.168.0.182', 'lan_port': '8000', 'auto_test': False, 'curve_visible': True},
            {'enabled': False, 'connection': 'lan', 'port': '', 'baudrate': '9600', 'name': 'ConST1210 CH3', 'read_command': 'MEASure:TEMPerature?\r\n', 'host': '192.168.0.182', 'lan_port': '8000', 'auto_test': False, 'curve_visible': True},
        ]

        self.device_threads = []
        self.data_buffer = {}
        self.time_buffer = {}
        self.datetime_buffer = {}
        self.max_points = None
        self.test_running = False
        self.start_time = None
        self._auto_follow = True       # 曲线是否自动跟随最新数据（手动缩放时关闭）
        self._suppress_range_signal = False  # 抑制程序内 setRange 触发的范围信号
        self.device_colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b']
        self.debug_dialog = None
        self.current_data_file = None
        self.has_unsaved_data = False
        self.auto_save_timer = None  # 每45分钟自动保存定时器

        # 顺序执行多温度测试
        self.current_test_temp = 25         # 当前测试温度（替代旧的 test_temp_spin）
        self.stability_threshold = 0.05      # 稳定性判断条件（波动判稳阈值）
        self.current_test_spec = 0.5        # 当前测试温度对应的 Spec（T0 到达偏差）
        self.sequential_running = False     # 是否正在执行顺序测试
        self.sequential_current_row = 0     # 当前执行到第几行
        self.sequential_start_time = None   # 当前行测试开始时间
        self.sequential_timer = QTimer()    # 顺序执行定时器
        self.sequential_timer.timeout.connect(self._sequential_tick)
        self.sequential_test_complete = False  # 当前行T3是否已完成
        self.sequential_mode = 'forward'    # 测试模式：forward顺序 / reverse倒序 / loop循环
        self.seq_row_order = []             # 执行的行索引序列
        self.seq_step = 0                   # 在 seq_row_order 中的当前位置

        # 定时测试
        self.sched_test_armed = False
        self.sched_test_timer = QTimer()
        self.sched_test_timer.timeout.connect(self._scheduled_test_check)

        # 自动检测状态（按需创建）
        self.auto_test_state = {}        # dev_id -> state dict
        self.auto_test_log = []          # [(时间, 通道名, T0, T1, T2, T3, Std1, Std2, Avg1, Avg2), ...]
        self.auto_test_logged = {}       # dev_id -> set of logged phases
        self.auto_test_lines = []        # 图上标注线列表
        self.auto_test_summary = None    # T3完成后的跨通道汇总 {'Main':..,'Sec':..,...}

        # 轴向测试
        self.axial_running = False       # 轴向测试是否正在计时
        self.axial_timer = None          # 轴向测试计时器
        self.axial_start_time = None     # 轴向测试开始时间
        self.axial_test_duration = 0     # 轴向测试设定时长（秒）
        self.axial_records = []          # 轴向测试记录 [(key, F_avg, M_avg, 记录时间), ...]（兼容旧字段）
        self.axial_data = {}             # 轴向测试数据字典 {key(float或str): {'F':..., 'M':..., 'time':...}}
        self.axial_keys = []             # 已添加到表格的 key 顺序列表（列顺序）。轴向=float 高度，径向=str 位置对
        self.axial_mode = 'axial'        # 当前模式：'axial' (轴向，高度/mm) 或 'radial' (径向，位置)
        self.axial_current_height = 0    # 当前测试高度
        self.axial_queue = []            # 待测试高度队列
        self.axial_queue_idx = 0         # 队列当前位置
        self.axial_columns = []          # 轴向列信息 [{'key': k, 'mode': m, 'save_btn': btn, 'col': col, 'header_widget': hw}, ...]

        # Excel保存并发控制
        self._bg_save_thread = None
        self._sequential_saving = False

        # 注册程序意外退出时的应急保存
        atexit.register(self.emergency_save_data)
        try:
            signal.signal(signal.SIGINT, self._signal_handler)
            signal.signal(signal.SIGTERM, self._signal_handler)
        except (OSError, ValueError):
            pass  # 非主线程或信号不可用时忽略

        # 多物理量多Y轴支持
        self.device_quantity_info = {}   # dev_id -> {'quantity_type', 'display_name', 'unit'}
        self.device_viewboxes = {}       # dev_id -> ViewBox
        self.extra_vb_list = []          # 额外ViewBox列表（用于清理）
        self.quantity_viewbox_map = {}   # quantity_type -> ViewBox
        self.primary_plot = None

        # 温度源设置行数
        self._ts_row_count = 3
        self._ts_row_layouts = []        # 存储行布局引用

        # 温度源设备类型（Fluke 9250 / Const 1210）
        self._ts_device_type = 'Fluke 9250'
        # 每个设备的通讯方式偏好（'serial'/'lan'），切换设备时自动恢复
        self._ts_device_conn_prefs = {'Fluke 9250': 'serial', 'Const 1210': 'lan'}
        # 每设备独立的连接管理器、连接状态、通讯参数
        self._ts_device_managers = {'Fluke 9250': None, 'Const 1210': None}
        self._ts_device_connected = {'Fluke 9250': False, 'Const 1210': False}
        self._ts_device_settings = {
            'Fluke 9250': {'port': '', 'baud': 115200, 'ip': '', 'lan_port': 8000},
            'Const 1210': {'port': '', 'baud': 115200, 'ip': '', 'lan_port': 8000},
        }
        # 每个设备类型独立保存的完整温度源参数（行设置 + 波动阈值），切换设备时自动恢复
        self._ts_device_full_settings = {
            'Fluke 9250': None, 'Const 1210': None,
        }
        self._ts_row_advanced_widgets = []  # 每行的"高级控件"列表（Const 1210 时需要隐藏）

        # 设备行列数
        self._dev_row_count = 6
        self._dev_row_containers = []    # 存储动态添加的设备行容器

        self._loading = True   # 加载标志，防止初始化时 save_config 覆盖配置文件
        self.init_ui()
        self.load_config()
        self._loading = False  # 结束加载，之后修改才会触发保存

    # ---- 温度源 per-device 代理属性：自动路由到当前设备 ----
    @property
    def temp_source_manager(self):
        """返回当前选中设备的连接管理器"""
        return self._ts_device_managers.get(self._ts_device_type)

    @temp_source_manager.setter
    def temp_source_manager(self, value):
        self._ts_device_managers[self._ts_device_type] = value

    @property
    def temp_source_connected(self):
        """返回当前选中设备的连接状态"""
        return self._ts_device_connected.get(self._ts_device_type, False)

    @temp_source_connected.setter
    def temp_source_connected(self, value):
        self._ts_device_connected[self._ts_device_type] = value

    def _redistribute_left_splitter(self):
        """当左侧可折叠组框切换时，重新分配 left_splitter 空间并更新最小尺寸"""
        if not hasattr(self, 'left_splitter') or self.left_splitter.count() < 2:
            return
        total = sum(self.left_splitter.sizes())
        if total <= 0:
            return

        top_widget = self.left_splitter.widget(0)
        legend_group = self.left_splitter.widget(1)
        spacing = 5  # 与 top_left_layout.setSpacing(5) 一致

        # 遍历 top_widget 内所有 CollapsibleGroupBox，计算最小高度
        top_min_h = 4  # 内边距余量
        gb_count = 0
        for child in top_widget.children():
            if isinstance(child, CollapsibleGroupBox):
                gb_count += 1
                if child.is_collapsed():
                    top_min_h += child.toggle_btn.sizeHint().height()
                else:
                    top_min_h += child.sizeHint().height()
        if gb_count > 1:
            top_min_h += spacing * (gb_count - 1)

        # 下部实时数据显示组框的最小高度（折叠时只保留标题栏）
        legend_min_h = legend_group.toggle_btn.sizeHint().height() + 8
        legend_group.setMinimumHeight(legend_min_h)

        top_min_h = max(top_min_h, 40, legend_min_h)

        # 设置 top_left_widget 的最小高度（拖动时不允许低于此值）
        top_widget.setMinimumHeight(top_min_h)

        new_legend_h = max(total - top_min_h, legend_min_h)
        new_top_h = total - new_legend_h
        self.left_splitter.setSizes([new_top_h, new_legend_h])

    def _update_title_datetime(self):
        """在窗口标题的版本号旁刷新当前系统日期时间"""
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.setWindowTitle(f"T-cal_tester v{APP_VERSION}    {now}")

    def init_ui(self):
        self.setGeometry(100, 100, 1800, 950)
        self._geometry_restored = False  # 标记：showEvent 中恢复一次后不再重复
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QHBoxLayout(main_widget)
        layout.setContentsMargins(10,10,10,10)
        layout.setSpacing(0)

        # 窗口标题栏：版本号旁显示实时日期时间（每秒刷新）
        self._update_title_datetime()
        self.title_time_timer = QTimer(self)
        self.title_time_timer.timeout.connect(self._update_title_datetime)
        self.title_time_timer.start(1000)

        # 主水平分割器：左侧(设备+曲线) ｜ 右侧(数据显示)
        self.main_splitter = QSplitter(Qt.Orientation.Horizontal)
        layout.addWidget(self.main_splitter)

        # 左侧面板：设备配置(上) + 实时曲线(下)
        left_panel = QWidget()
        left_panel.setStyleSheet("background-color: transparent;")
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(5)

        # 左侧内部垂直分割器：设备区域(上) ｜ 曲线(下)
        self.left_splitter = QSplitter(Qt.Orientation.Vertical)
        left_layout.addWidget(self.left_splitter)

        # 左侧顶部容器：设备配置 + 温度源控制
        top_left_widget = QWidget()
        top_left_layout = QVBoxLayout(top_left_widget)
        top_left_layout.setContentsMargins(0, 0, 0, 0)
        top_left_layout.setSpacing(5)

        # 设备配置区域（可折叠）
        self.device_group = CollapsibleGroupBox("设备配置")
        self.device_group.toggled.connect(self._redistribute_left_splitter)
        top_left_layout.addWidget(self.device_group)
        device_layout = self.device_group.contentLayout()

        # 设备表头（仅保留+/-按钮）
        header_widget = QWidget()
        header_layout = QHBoxLayout(header_widget)
        header_layout.setContentsMargins(0,0,0,0)
        header_layout.setSpacing(4)

        # 加减按钮
        dev_add_btn = QPushButton("+"); dev_add_btn.setFixedSize(22,22)
        dev_add_btn.setStyleSheet("QPushButton{background:#4CAF50;color:white;font-weight:bold;border-radius:4px;}QPushButton:hover{opacity:0.8;}")
        dev_add_btn.clicked.connect(self._dev_add_row)
        header_layout.addWidget(dev_add_btn)
        dev_remove_btn = QPushButton("-"); dev_remove_btn.setFixedSize(22,22)
        dev_remove_btn.setStyleSheet("QPushButton{background:#f44336;color:white;font-weight:bold;border-radius:4px;}QPushButton:hover{opacity:0.8;}")
        dev_remove_btn.clicked.connect(self._dev_remove_row)
        header_layout.addWidget(dev_remove_btn)
        header_layout.addStretch()

        device_layout.addWidget(header_widget)

        # 设备行区域（可滚动 + 拖拽排序）
        dev_scroll = QScrollArea()
        dev_scroll.setWidgetResizable(True)
        dev_scroll.setFrameShape(QFrame.Shape.NoFrame)
        dev_scroll.setStyleSheet("QScrollArea{background:transparent;}")
        dev_scroll.setMinimumHeight(100)
        dev_scroll.setMaximumHeight(400)
        dev_scroll_widget = QWidget()
        self._dev_rows_layout = QVBoxLayout(dev_scroll_widget)
        self._dev_rows_layout.setContentsMargins(0,0,0,0)
        self._dev_rows_layout.setSpacing(2)
        dev_scroll.setWidget(dev_scroll_widget)
        device_layout.addWidget(dev_scroll)

        self.device_widgets = []
        self._dev_row_containers = []

        self._build_device_rows()

        # 第一行：读取间隔
        auto_layout = QHBoxLayout()
        auto_layout.addWidget(QLabel("读取间隔(ms):"))
        self.interval_spin = QSpinBox()
        self.interval_spin.setRange(100,5000)
        self.interval_spin.setValue(1000)
        self.interval_spin.setSuffix(" ms")
        auto_layout.addWidget(self.interval_spin)
        auto_layout.addStretch()
        device_layout.addLayout(auto_layout)

        # 统一按钮样式
        def make_button(text, color, callback, enabled=True, width=78):
            btn = QPushButton(text)
            btn.setStyleSheet(f"""
                QPushButton {{background:{color};color:white;font-weight:bold;font-size:11px;border-radius:4px;padding:0px;}}
                QPushButton:hover {{background:{color};opacity:0.8;}}
                QPushButton:disabled {{background:#ccc;color:#888;}}
            """)
            btn.setFixedSize(width, 24)
            btn.setEnabled(enabled)
            btn.clicked.connect(callback)
            return btn

        # 按钮布局
        left_btn_layout = QHBoxLayout()
        left_btn_layout.setSpacing(6)
        for text, cb, color in [
            ("刷新串口", self.refresh_ports, "#607D8B"),
            ("保存配置", self.save_config_to_file, "#9C27B0"),
            ("导入配置", self.load_config_from_file, "#FF5722"),
            ("通讯调试", self.open_debug_window, "#3F51B5"),
        ]:
            left_btn_layout.addWidget(make_button(text, color, cb))
            left_btn_layout.setStretch(left_btn_layout.count() - 1, 1)

        btn_container = QHBoxLayout()
        btn_container.setSpacing(10)
        btn_container.addLayout(left_btn_layout)
        btn_container.addStretch()
        device_layout.addLayout(btn_container)

        # 状态栏
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_label = QLabel("就绪")
        self.status_bar.addWidget(self.status_label)

        # 温度源控制区域（可折叠）——标题改为“测试模式”
        self.temp_ctrl_group = CollapsibleGroupBox("测试模式")
        left_layout.addWidget(self.temp_ctrl_group)
        temp_ctrl_layout = self.temp_ctrl_group.contentLayout()
        temp_ctrl_layout.setSpacing(5)

        # 温度源设备选择
        device_select_layout = QHBoxLayout()
        # 稳定性判断条件（波动阈值）
        device_select_layout.addWidget(QLabel("波动阈值:"))
        self.stability_threshold_spin = QDoubleSpinBox()
        self.stability_threshold_spin.setRange(0, 9999)
        self.stability_threshold_spin.setValue(0.05)
        self.stability_threshold_spin.setSingleStep(0.01)
        self.stability_threshold_spin.setDecimals(3)
        self.stability_threshold_spin.setFixedWidth(70)
        self.stability_threshold_spin.setStyleSheet("QDoubleSpinBox{border:1px solid #999;border-radius:3px;padding:2px;}QDoubleSpinBox::up-button,QDoubleSpinBox::down-button,QDoubleSpinBox::up-arrow,QDoubleSpinBox::down-arrow{width:0px;height:0px;}")
        self.stability_threshold_spin.setToolTip("稳定性判断条件：最近窗口内温度波动标准差 <= 该值时判稳")
        self.stability_threshold_spin.valueChanged.connect(lambda v: self._on_stability_threshold_changed(v))
        device_select_layout.addWidget(self.stability_threshold_spin)
        device_select_layout.addWidget(QLabel("设备类型:"))
        self.temp_source_device_combo = QComboBox()
        self.temp_source_device_combo.addItems(['Fluke 9250', 'Const 1210'])
        self.temp_source_device_combo.setFixedWidth(130)
        self.temp_source_device_combo.currentTextChanged.connect(self._on_temp_source_device_changed)
        device_select_layout.addWidget(self.temp_source_device_combo)
        device_select_layout.addStretch()
        temp_ctrl_layout.addLayout(device_select_layout)

        # 表头（仅保留+/-按钮）
        header_layout = QHBoxLayout()
        self._ts_add_btn = QPushButton("+")
        self._ts_add_btn.setFixedSize(24, 24)
        self._ts_add_btn.setStyleSheet("QPushButton{background:#4CAF50;color:white;font-weight:bold;border-radius:4px;}QPushButton:hover{opacity:0.8;}")
        self._ts_add_btn.clicked.connect(self._ts_add_row)
        header_layout.addWidget(self._ts_add_btn)
        self._ts_remove_btn = QPushButton("-")
        self._ts_remove_btn.setFixedSize(24, 24)
        self._ts_remove_btn.setStyleSheet("QPushButton{background:#f44336;color:white;font-weight:bold;border-radius:4px;}QPushButton:hover{opacity:0.8;}")
        self._ts_remove_btn.clicked.connect(self._ts_remove_row)
        header_layout.addWidget(self._ts_remove_btn)
        header_layout.addStretch()

        # ===== 自动测试区域（外框 QGroupBox，置于 SP 表格上方）=====
        self.ts_auto_cont = QGroupBox("自动测试")
        self.ts_auto_cont.setStyleSheet(
            "QGroupBox{font-size:11px;font-weight:bold;color:#9C27B0;"
            "border:1px solid #b39ddb;border-radius:5px;margin-top:8px;padding:3px;}"
            "QGroupBox::title{subcontrol-origin:margin;left:8px;top:2px;}")
        # 自动测试框内垂直布局：按钮行 + 表头 + SP 表格
        ts_auto_layout = QVBoxLayout(self.ts_auto_cont)
        ts_auto_layout.setContentsMargins(4, 8, 4, 4)
        ts_auto_layout.setSpacing(2)
        # 按钮行（水平）：顺序/循环/倒序/间隔/定时
        self.ts_auto_btns_row = QHBoxLayout()
        self.ts_auto_btns_row.setSpacing(4)
        ts_auto_layout.addLayout(self.ts_auto_btns_row)
        self.ts_rows_layout = ts_auto_layout  # SP 行所在布局引用
        temp_ctrl_layout.addWidget(self.ts_auto_cont)

        # 表头加入自动测试框内（SP 表格上方）
        ts_auto_layout.addLayout(header_layout)

        # 行设置
        self.row_setpoint_spins = []
        self.row_setpoint_spec = []   # 每行的 Spec（到达T0的允许偏差）
        self.row_main_spins = []
        self.row_sec_spins = []
        self.row_send_btns = []       # 每行的“发送命令”按钮
        self.row_checks = []          # 每行的勾选框
        self._ts_row_layouts = []
        # 每行的 Main PID 参数
        self.row_main_pid_p = []
        self.row_main_pid_i = []
        self.row_main_pid_d = []
        # 每行的 Sec PID 参数
        self.row_sec_pid_p = []
        self.row_sec_pid_i = []
        self.row_sec_pid_d = []
        # 每行的 Weight 设置
        self.row_weights = []

        # 紧凑无箭头 SpinBox 样式
        no_arrow_style = """
            QDoubleSpinBox::up-button, QDoubleSpinBox::down-button {
                width: 0px; height: 0px;
            }
            QDoubleSpinBox {
                padding: 0px 2px;
                border: 1px solid #999999;
            }
        """

        compact_spin = lambda range_min, val, step, decimal, width: (
            lambda: (lambda s: (setattr(s, '_tmp', None), s)[1])(QDoubleSpinBox())  # placeholder
        )

        def make_spin(range_min, val, step, decimal, width, suffix=''):
            # PID 等可删除项使用可清空输入框，留空表示使用默认值（不发送）
            s = NullableSpinBox()
            s.setRange(range_min, 9999)
            s.setValue(val)
            s.setSingleStep(step)
            s.setDecimals(decimal)
            s.setFixedWidth(width)
            s.setStyleSheet(no_arrow_style)
            s.setPlaceholderText("默认")
            if suffix:
                s.setSuffix(suffix)
            return s

        def make_labeled(label_text, widget):
            """标签在框外、紧贴输入框、无底色，形成紧凑单元（如 SP 25.0）"""
            g = QWidget()
            h = QHBoxLayout(g)
            h.setContentsMargins(0, 0, 0, 0)
            h.setSpacing(0)
            lbl = QLabel(label_text)
            lbl.setStyleSheet("QLabel{color:#333333;margin:0px;padding:0px;}")
            h.addWidget(lbl)
            if isinstance(widget, QDoubleSpinBox):
                widget.setStyleSheet((widget.styleSheet() or "") + "QDoubleSpinBox{padding-left:0px;}")
            h.addWidget(widget)
            return g

        for row_idx in range(3):
            ts_container = DragDropRowWidget(row_idx)
            ts_container.reorder_requested.connect(self._ts_reorder_rows)
            row_layout = QHBoxLayout(ts_container.contentWidget())
            row_layout.setSpacing(3)
            row_layout.setContentsMargins(1, 0, 1, 0)

            chk = QCheckBox()
            chk.setChecked(True)
            chk.stateChanged.connect(lambda v, idx=row_idx: self.save_config())
            row_layout.addWidget(chk)
            self.row_checks.append(chk)

            # 发送命令按钮：放在最左侧（勾选框之后），点击后发送该行的所有参数
            send_btn = QPushButton("发送")
            send_btn.setFixedWidth(46)
            send_btn.setToolTip("发送该行的所有参数到温度源")
            send_btn.setStyleSheet(
                "QPushButton{background:#2196F3;color:white;font-weight:bold;font-size:11px;border-radius:4px;padding:0px;}"
                "QPushButton:hover{background:#1976D2;}"
                "QPushButton:disabled{background:#ccc;color:#888;}"
            )
            send_btn.clicked.connect(lambda _, btn=send_btn: self._send_row_command(self.row_send_btns.index(btn)))
            send_btn.setEnabled(False)  # 初始未通信，禁用
            row_layout.addWidget(send_btn)
            self.row_send_btns.append(send_btn)

            # Setpoint（设定温度）：标签与输入框组合为一个整体
            sp = make_spin(0, 25, 0.01, 2, 70)
            sp.valueChanged.connect(lambda v, idx=row_idx: self._on_row_setpoint_changed(idx, v))
            row_layout.addWidget(make_labeled("SP", sp))
            self.row_setpoint_spins.append(sp)

            # Spec（到达T0的允许偏差）：标签与输入框组合为一个整体
            spec = make_spin(0, 0.5, 0.1, 2, 55)
            spec.valueChanged.connect(lambda v, idx=row_idx: self._on_row_spec_changed(idx, v))
            row_layout.addWidget(make_labeled("Spec", spec))
            self.row_setpoint_spec.append(spec)

            # 以下为 Fluke 系列高级控件（Const 1210 时隐藏），同样组合为整体单元
            advanced = []

            main_spin = make_spin(0, 50, 0.01, 2, 55)
            main_spin.valueChanged.connect(lambda v, idx=row_idx: self._on_row_main_changed(idx, v))
            g = make_labeled("M", main_spin)
            row_layout.addWidget(g); advanced.append(g)
            self.row_main_spins.append(main_spin)

            mp = make_spin(0, 10, 0.1, 1, 55)
            mp.valueChanged.connect(lambda v: self.save_config())
            g = make_labeled("M-P", mp)
            row_layout.addWidget(g); advanced.append(g)
            self.row_main_pid_p.append(mp)

            mi = make_spin(0, 200, 1, 1, 55)
            mi.valueChanged.connect(lambda v: self.save_config())
            g = make_labeled("M-I", mi)
            row_layout.addWidget(g); advanced.append(g)
            self.row_main_pid_i.append(mi)

            md = make_spin(0, 50, 1, 1, 55)
            md.valueChanged.connect(lambda v: self.save_config())
            g = make_labeled("M-D", md)
            row_layout.addWidget(g); advanced.append(g)
            self.row_main_pid_d.append(md)

            sec_spin = make_spin(0, 0, 0.1, 2, 55)
            sec_spin.valueChanged.connect(lambda v, idx=row_idx: self._on_row_sec_changed(idx, v))
            g = make_labeled("S", sec_spin)
            row_layout.addWidget(g); advanced.append(g)
            self.row_sec_spins.append(sec_spin)

            sp_p = make_spin(0, 10, 0.1, 1, 55)
            sp_p.valueChanged.connect(lambda v: self.save_config())
            g = make_labeled("S-P", sp_p)
            row_layout.addWidget(g); advanced.append(g)
            self.row_sec_pid_p.append(sp_p)

            si = make_spin(0, 200, 1, 1, 55)
            si.valueChanged.connect(lambda v: self.save_config())
            g = make_labeled("S-I", si)
            row_layout.addWidget(g); advanced.append(g)
            self.row_sec_pid_i.append(si)

            sd = make_spin(0, 50, 1, 1, 55)
            sd.valueChanged.connect(lambda v: self.save_config())
            g = make_labeled("S-D", sd)
            row_layout.addWidget(g); advanced.append(g)
            self.row_sec_pid_d.append(sd)

            # Weight：已取消显示，保留空列表维持索引对齐（数据不再通过 UI 输入）
            self.row_weights.append([])

            row_layout.addStretch()  # 输入框靠左，不拉伸到右侧

            self._ts_row_advanced_widgets.append(advanced)

            # SP 行加入自动测试框内的布局
            self.ts_rows_layout.addWidget(ts_container)
            self._ts_row_layouts.append(ts_container)

        # 通用按钮样式
        btn_s = "QPushButton{color:white;font-weight:bold;font-size:11px;border-radius:4px;}"
        btn_d = "QPushButton:disabled{background:#ccc;color:#888;}"

        # 手动测试区域外框容器（QGroupBox）
        self.ts_manual_group = QGroupBox("手动测试")
        self.ts_manual_group.setStyleSheet(
            "QGroupBox{font-size:11px;font-weight:bold;color:#1f77b4;"
            "border:1px solid #90caf9;border-radius:5px;margin-top:8px;padding:3px;}"
            "QGroupBox::title{subcontrol-origin:margin;left:8px;top:2px;}")
        manual_v = QVBoxLayout(self.ts_manual_group)
        manual_v.setContentsMargins(4, 8, 4, 4)
        manual_v.setSpacing(4)

        ctrl_row = QHBoxLayout()
        ctrl_row.setSpacing(4)

        # 连接类型选择
        self.temp_source_conn_combo = QComboBox()
        self.temp_source_conn_combo.addItems(['serial', 'lan'])
        self.temp_source_conn_combo.setFixedWidth(65)
        self.temp_source_conn_combo.currentTextChanged.connect(self._on_temp_source_conn_changed)
        ctrl_row.addWidget(self.temp_source_conn_combo)

        # Serial 参数
        self.temp_source_port_combo = PortComboBox()
        self.temp_source_port_combo.setMinimumWidth(70)
        self.temp_source_port_combo.currentTextChanged.connect(lambda v: self.save_config())
        ctrl_row.addWidget(self.temp_source_port_combo)
        self.temp_source_baud_spin = QSpinBox()
        self.temp_source_baud_spin.setRange(1200, 921600)
        self.temp_source_baud_spin.setValue(115200)
        self.temp_source_baud_spin.setFixedWidth(65)
        self.temp_source_baud_spin.setStyleSheet("QSpinBox::up-button, QSpinBox::down-button { width: 0px; height: 0px; }")
        self.temp_source_baud_spin.valueChanged.connect(lambda v: self.save_config())
        ctrl_row.addWidget(self.temp_source_baud_spin)

        # LAN 参数（初始隐藏）
        self.temp_source_ip_edit = QLineEdit()
        self.temp_source_ip_edit.setPlaceholderText("IP地址")
        self.temp_source_ip_edit.setFixedWidth(110)
        self.temp_source_ip_edit.textChanged.connect(lambda v: self.save_config())
        self.temp_source_ip_edit.setVisible(False)
        ctrl_row.addWidget(self.temp_source_ip_edit)
        self.temp_source_lan_port_spin = QSpinBox()
        self.temp_source_lan_port_spin.setRange(1, 65535)
        self.temp_source_lan_port_spin.setValue(8000)
        self.temp_source_lan_port_spin.setFixedWidth(55)
        self.temp_source_lan_port_spin.valueChanged.connect(lambda v: self.save_config())
        self.temp_source_lan_port_spin.setVisible(False)
        ctrl_row.addWidget(self.temp_source_lan_port_spin)

        self.serial_btn = QPushButton()
        self.serial_btn.setFixedSize(65, 28)
        self.serial_btn.clicked.connect(self._toggle_serial)
        ctrl_row.addWidget(self.serial_btn)
        self._update_serial_btn(False)

        # SP 手动设置输入框 + 独立发送按钮（紧邻通信按钮右侧）
        sp_ctl_lbl = QLabel("SP:")
        sp_ctl_lbl.setStyleSheet("font-size:11px;font-weight:bold;color:#555555;")
        ctrl_row.addWidget(sp_ctl_lbl)
        self.ts_manual_sp_spin = QDoubleSpinBox()
        self.ts_manual_sp_spin.setRange(-100000, 100000)  # 放开范围，不再限3位
        self.ts_manual_sp_spin.setDecimals(3)
        self.ts_manual_sp_spin.setSingleStep(1.0)
        self.ts_manual_sp_spin.setValue(25.0)
        self.ts_manual_sp_spin.setFixedWidth(100)
        self.ts_manual_sp_spin.setStyleSheet(no_arrow_style)
        ctrl_row.addWidget(self.ts_manual_sp_spin)
        self.ts_manual_sp_btn = QPushButton("发送")
        self.ts_manual_sp_btn.setFixedSize(55, 28)
        self.ts_manual_sp_btn.setStyleSheet(btn_s + "QPushButton{background:#4CAF50;}" + btn_d)
        self.ts_manual_sp_btn.clicked.connect(self._send_manual_sp)
        self.ts_manual_sp_btn.setEnabled(False)  # 初始未通信，禁用
        ctrl_row.addWidget(self.ts_manual_sp_btn)

        self.ts_btn = QPushButton()
        self.ts_btn.setFixedSize(75, 28)
        self.ts_btn.clicked.connect(self._toggle_ts)
        ctrl_row.addWidget(self.ts_btn)
        self._update_ts_btn(False)
        self.ts_btn.setEnabled(False)  # 初始串口未开，禁用

        # ===== 自动测试按钮（顺序/循环/倒序/间隔/定时）加入自动测试框按钮行 =====
        ts_auto_layout = self.ts_auto_btns_row
        self.seq_btn = QPushButton()
        self.seq_btn.setFixedSize(85, 28)
        self.seq_btn.clicked.connect(self._toggle_seq)
        ts_auto_layout.addWidget(self.seq_btn)
        self._update_seq_btn()
        self.seq_btn.setEnabled(False)  # 初始串口未开，禁用

        # 循环测试按钮
        self.loop_btn = QPushButton()
        self.loop_btn.setFixedSize(85, 28)
        self.loop_btn.clicked.connect(self._toggle_loop)
        ts_auto_layout.addWidget(self.loop_btn)
        self._update_loop_btn()
        self.loop_btn.setEnabled(False)  # 初始串口未开，禁用

        # 倒序测试按钮
        self.reverse_btn = QPushButton()
        self.reverse_btn.setFixedSize(85, 28)
        self.reverse_btn.clicked.connect(self._toggle_reverse)
        ts_auto_layout.addWidget(self.reverse_btn)
        self._update_reverse_btn()
        self.reverse_btn.setEnabled(False)  # 初始串口未开，禁用

        ts_auto_layout.addWidget(QLabel("间隔:"))
        self.test_interval_spin = QDoubleSpinBox()
        self.test_interval_spin.setRange(0, 9999)
        self.test_interval_spin.setValue(0)
        self.test_interval_spin.setSingleStep(1)
        self.test_interval_spin.setDecimals(1)
        self.test_interval_spin.setFixedWidth(50)
        self.test_interval_spin.setStyleSheet("QDoubleSpinBox{border:1px solid #999;border-radius:3px;padding:2px;}QDoubleSpinBox::up-button,QDoubleSpinBox::down-button,QDoubleSpinBox::up-arrow,QDoubleSpinBox::down-arrow{width:0px;height:0px;}")
        self.test_interval_spin.setSuffix("min")
        self.test_interval_spin.setToolTip("每个setpoint完成后等待的时间")
        self.test_interval_spin.valueChanged.connect(lambda v: self.save_config())
        ts_auto_layout.addWidget(self.test_interval_spin)

        self.sched_time_edit = QDateTimeEdit()
        self.sched_time_edit.setDisplayFormat("MM-dd HH:mm")
        self.sched_time_edit.setDateTime(QDateTime.currentDateTime().addSecs(300))
        self.sched_time_edit.setFixedWidth(100)
        self.sched_time_edit.setCalendarPopup(False)
        self.sched_time_edit.setStyleSheet("QDateTimeEdit{border:1px solid #999;border-radius:3px;padding:2px;font-size:11px;}QDateTimeEdit::up-button,QDateTimeEdit::down-button,QDateTimeEdit::up-arrow,QDateTimeEdit::down-arrow,QDateTimeEdit::drop-down,QDateTimeEdit::calendar-popup{subcontrol-origin:border;subcontrol-position:right;width:0px;height:0px;border:none;image:none;margin:0;padding:0;}")
        ts_auto_layout.addWidget(self.sched_time_edit)
        self.sched_test_btn = QPushButton()
        self.sched_test_btn.setFixedSize(70, 28)
        self.sched_test_btn.clicked.connect(self._toggle_scheduled_test)
        ts_auto_layout.addWidget(self.sched_test_btn)
        self._update_sched_btn()
        self.sched_status_label = QLabel("")
        self.sched_status_label.setStyleSheet("font-size:10px;color:#607D8B;")
        ts_auto_layout.addWidget(self.sched_status_label)
        ts_auto_layout.addStretch()

        # Const 1210 温度
        self.ts_const1210_temp_label = QLabel("Temp:--°C")
        self.ts_const1210_temp_label.setStyleSheet("font-size:11px;font-weight:bold;color:#e65100;")
        self.ts_const1210_temp_label.setVisible(False)
        ctrl_row.addWidget(self.ts_const1210_temp_label)
        ctrl_row.addStretch()
        manual_v.addLayout(ctrl_row)

        # ===== 采集操作按钮行（并入手动测试区域，独立一行避免叠字）=====
        collect_row = QHBoxLayout()
        collect_row.setSpacing(6)
        self.collect_btn = QPushButton("开始测试")
        self.collect_btn.setFixedSize(88, 26)
        self.collect_btn.clicked.connect(self._toggle_collection)
        collect_row.addWidget(self.collect_btn)
        self._update_collect_btn()
        # 是否判断 T3 时刻勾选框（不勾选则连续记录，不判断 T3 结束条件）
        self.manual_check_t3_cb = QCheckBox("判断T3时刻")
        self.manual_check_t3_cb.setChecked(True)
        self.manual_check_t3_cb.setStyleSheet("font-size:11px;color:#1f77b4;")
        self.manual_check_t3_cb.stateChanged.connect(lambda v: self.save_config())
        collect_row.addWidget(self.manual_check_t3_cb)
        self.record_btn = make_button("记录数据", "#E91E63", self.record_current_data, enabled=False)
        collect_row.addWidget(self.record_btn)
        self.reset_stats_btn = make_button("重置统计", "#FF9800", self.reset_stats_all)
        collect_row.addWidget(self.reset_stats_btn)
        self.reset_plot_btn = make_button("清空图像", "#795548", self.reset_curve_display)
        collect_row.addWidget(self.reset_plot_btn)
        self.reset_view_btn = make_button("缩放重置", "#607D8B", self._reset_plot_view)
        collect_row.addWidget(self.reset_view_btn)
        self.save_btn = make_button("手动保存", "#009688", self.manual_save_data, enabled=False)
        collect_row.addWidget(self.save_btn)
        self.screenshot_btn = make_button("截图", "#673AB7", self.save_current_plot)
        collect_row.addWidget(self.screenshot_btn)
        collect_row.addStretch()
        manual_v.addLayout(collect_row)
        # 手动测试外框容器加入温度源控制布局
        temp_ctrl_layout.addWidget(self.ts_manual_group)

        # ===== 轴向测试区域（置于手动测试下方，形式参考自动测试）=====
        self.axial_group = QGroupBox("轴向/径向测试")
        self.axial_group.setStyleSheet(
            "QGroupBox{font-size:11px;font-weight:bold;color:#00796B;"
            "border:1px solid #4db6ac;border-radius:5px;margin-top:8px;padding:3px;}"
            "QGroupBox::title{subcontrol-origin:margin;left:8px;top:2px;}")
        axial_layout = QVBoxLayout(self.axial_group)
        axial_layout.setContentsMargins(4, 8, 4, 4)
        axial_layout.setSpacing(4)

        # 输入行：加减按钮（用于增加/减少高度列）
        axial_input_row = QHBoxLayout()
        # 加减按钮
        self.axial_minus_btn = QPushButton("-")
        self.axial_minus_btn.setFixedSize(26, 26)
        self.axial_minus_btn.clicked.connect(self._axial_remove_height_column)
        self.axial_minus_btn.setStyleSheet(
            "QPushButton{background:#f44336;color:white;font-weight:bold;border-radius:4px;font-size:13px;}")
        axial_input_row.addWidget(self.axial_minus_btn)
        self.axial_plus_btn = QPushButton("+")
        self.axial_plus_btn.setFixedSize(26, 26)
        # 用 lambda 拦截 clicked 信号的 bool 参数，避免 height=False 导致重复
        self.axial_plus_btn.clicked.connect(lambda: self._axial_add_height_column())
        self.axial_plus_btn.setStyleSheet(
            "QPushButton{background:#4CAF50;color:white;font-weight:bold;border-radius:4px;font-size:13px;}")
        axial_input_row.addWidget(self.axial_plus_btn)
        axial_layout.addLayout(axial_input_row)
        axial_layout.setAlignment(axial_input_row, Qt.AlignmentFlag.AlignLeft)

        # 状态显示（初始无文字，无"未开始"）
        self.axial_status_label = QLabel("")
        self.axial_status_label.setStyleSheet("font-size:11px;color:#00796B;font-weight:bold;")
        axial_layout.addWidget(self.axial_status_label)

        # 按钮行：模式选择 + 开始/记录按钮
        axial_btns_layout = QHBoxLayout()
        axial_btns_layout.setSpacing(2)
        # 模式选择下拉框
        self.axial_mode_combo = QComboBox()
        self.axial_mode_combo.addItems(["轴向", "径向"])
        self.axial_mode_combo.setFixedWidth(70)
        self.axial_mode_combo.currentIndexChanged.connect(self._axial_on_mode_changed)
        self.axial_mode_combo.setStyleSheet(
            "QComboBox{font-size:11px;font-weight:bold;padding:2px 4px;border:1px solid #4db6ac;border-radius:3px;}"
            "QComboBox QAbstractItemView{font-size:11px;}")
        axial_btns_layout.addWidget(self.axial_mode_combo)
        self.axial_start_btn = QPushButton("开始轴向测试")
        self.axial_start_btn.clicked.connect(self._toggle_axial_collection)
        self.axial_start_btn.setStyleSheet(
            "QPushButton{background:#00796B;color:white;font-weight:bold;border-radius:4px;padding:3px 8px;}"
            "QPushButton:disabled{background:#ccc;color:#888;}"
        )
        self.axial_save_btn = QPushButton("记录数据")
        self.axial_save_btn.clicked.connect(self._axial_record_data)
        self.axial_save_btn.setStyleSheet(
            "QPushButton{background:#009688;color:white;font-weight:bold;border-radius:4px;padding:3px 8px;}"
        )
        # 开始测试和保存Excel按钮不强制与下方对齐，按 preferred size 排列
        axial_btns_layout.addWidget(self.axial_start_btn)
        axial_btns_layout.addWidget(self.axial_save_btn)
        axial_layout.addLayout(axial_btns_layout)
        axial_layout.setAlignment(axial_btns_layout, Qt.AlignmentFlag.AlignLeft)

        # 示意图：已取消显示

        # 每列上方的标题行（保存按钮作为列标题，从第2列开始）
        self.axial_save_btns_layout = QHBoxLayout()
        self.axial_save_btns_layout.setSpacing(0)
        self.axial_save_btns_layout.setContentsMargins(0, 0, 0, 0)
        axial_layout.addLayout(self.axial_save_btns_layout)
        # 第1列上方显示列标题（轴向=高度，径向=位置）
        self.axial_col0_header = QLabel("高度")
        self.axial_col0_header.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.axial_col0_header.setFixedHeight(22)
        self.axial_col0_header.setStyleSheet(
            "background:#f0f0f0;color:#333333;font-size:11px;font-weight:bold;"
            "border:1px solid #dddddd;border-radius:3px;padding:1px;")
        self.axial_save_btns_layout.addWidget(self.axial_col0_header, stretch=1)

        # 结果表格：横坐标=高度(各列为一个高度)，纵坐标=高度/F-avg/M-avg 三行
        # 行0=高度(mm)可编辑，行1=F-avg，行2=M-avg
        self.axial_table = QTableWidget(3, 1)
        # 隐藏垂直表头与水平表头（保存按钮作为列标题）
        self.axial_table.verticalHeader().setVisible(False)
        self.axial_table.horizontalHeader().setVisible(False)
        # 第0列作为行标签列：行0留空（标题在按钮行），行1=F-avg、行2=M-avg（不可编辑）
        blank = QTableWidgetItem("")
        blank.setFlags(blank.flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.axial_table.setItem(0, 0, blank)
        for r, label in enumerate(["F-avg", "M-avg"], start=1):
            it = QTableWidgetItem(label)
            it.setFlags(it.flags() & ~Qt.ItemFlag.ItemIsEditable)
            it.setBackground(QColor("#f0f0f0"))
            it.setForeground(QColor("#333333"))
            self.axial_table.setItem(r, 0, it)
        # 表格样式：白底
        self.axial_table.setStyleSheet(
            "QTableWidget{background:#ffffff;border:1px solid #dddddd;border-radius:4px;"
            "gridline-color:#eeeeee;font-size:11px;}"
            "QTableWidget::item{padding:3px;}")
        self.axial_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.axial_table.horizontalHeader().setDefaultSectionSize(28)
        # 高度 = 3行单元格高（隐藏表头后不再加表头高度）
        self.axial_table.verticalHeader().setDefaultSectionSize(28)
        self.axial_table.setFixedHeight(
            self.axial_table.verticalHeader().defaultSectionSize() * 3 + 4)
        self.axial_table.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        # 监听高度单元格编辑，自动补 mm 单位
        self.axial_table.cellChanged.connect(self._axial_on_height_edited)
        axial_layout.addWidget(self.axial_table)

        # 轴向列数据：[{height: float, save_btn: QPushButton, col: int}]
        self.axial_columns = []
        # 默认4列（0/10/20/30mm），每列上方都有保存按钮
        for h in [0.0, 10.0, 20.0, 30.0]:
            self._axial_add_height_column(h, add_save_btn=True)

        # 轴向测试外框容器加入测试模式布局（手动测试下方）
        temp_ctrl_layout.addWidget(self.axial_group)

        top_left_layout.addWidget(self.temp_ctrl_group)
        self.temp_ctrl_group.toggled.connect(self._redistribute_left_splitter)

        # 温度源控制区域加入顶部容器
        # （温度源控制组的创建在 init_ui 前半段已完成，此处补充）

        # 左侧垂直分割器：上部(设备+温度源) ｜ 下部(曲线)
        self.left_splitter.addWidget(top_left_widget)
        self.left_splitter.setCollapsible(0, False)

        # 左侧垂直分割器：上部(设备+温度源) ｜ 下部(实时数据显示)
        self.left_splitter.addWidget(top_left_widget)
        self.left_splitter.setCollapsible(0, False)

        # 实时数据显示区域（可折叠，放入左侧下部）
        legend_group = CollapsibleGroupBox("实时数据显示")
        legend_group.toggled.connect(self._redistribute_left_splitter)
        legend_inner_layout = legend_group.contentLayout()
        self.legend_widget = CustomLegendWidget()
        self.legend_widget.setMinimumWidth(300)
        legend_inner_layout.addWidget(self.legend_widget)
        self.left_splitter.addWidget(legend_group)
        self.left_splitter.setCollapsible(1, False)
        self.left_splitter.setStretchFactor(0, 0)  # 上部不拉伸
        self.left_splitter.setStretchFactor(1, 1)  # 下部拉伸

        self.ts_query_thread = None
        self.ts_const1210_query_thread = None

        # 右侧面板：实时数据曲线（全部高度）
        right_panel = QWidget()
        right_panel.setStyleSheet("background-color: transparent;")
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(5)

        plot_group = CollapsibleGroupBox("实时数据曲线")
        plot_layout = plot_group.contentLayout()

        plot_container = QWidget()
        pcl = QHBoxLayout(plot_container)
        self.plot_widget = pg.GraphicsLayoutWidget()
        self.plot_widget.setBackground('#fafafa')
        pcl.addWidget(self.plot_widget)
        plot_layout.addWidget(plot_container)
        right_layout.addWidget(plot_group)

        # 主水平分割器
        self.main_splitter.addWidget(left_panel)
        self.main_splitter.addWidget(right_panel)
        self.main_splitter.setCollapsible(0, False)
        self.main_splitter.setCollapsible(1, False)
        # 左侧不拉伸，右侧拉伸：保证窗口变大时曲线区优先增大
        self.main_splitter.setStretchFactor(0, 0)
        self.main_splitter.setStretchFactor(1, 1)
        # 设置初始大小：左侧(设备+显示)较小，右侧(曲线)较大
        self.main_splitter.setSizes([500, 1300])
        # 仅当用户手动拖动分隔条时保存比例
        self.main_splitter.splitterMoved.connect(self._on_main_splitter_moved)
        # 设置最小宽度，防止遮挡
        left_panel.setMinimumWidth(300)
        right_panel.setMinimumWidth(200)

        self.init_plots()
        self.refresh_ports()

        self.plot_timer = QTimer()
        self.plot_timer.timeout.connect(self.update_plots)
        self.plot_timer.start(200)

        # 自动保存定时器（间隔在采集开始时动态设置）
        self.save_timer = QTimer()
        self.save_timer.timeout.connect(self._on_save_timer)

    # ========== 按钮状态统一更新 ==========
    def _set_btn_style(self, btn, text, bg_color):
        """设置按钮文字和样式"""
        btn.setText(text)
        btn.setStyleSheet(f"""
            QPushButton {{background:{bg_color};color:white;font-weight:bold;font-size:11px;border-radius:4px;}}
            QPushButton:hover {{background:{bg_color};}}
            QPushButton:disabled {{background:#ccc;color:#888;}}
        """)

    def _update_collect_btn(self):
        """根据采集状态更新采集按钮"""
        if self.test_running:
            self._set_btn_style(self.collect_btn, "停止测试", "#f44336")
        else:
            self._set_btn_style(self.collect_btn, "开始测试", "#4CAF50")
        self.collect_btn.setEnabled(True)
        # 记录数据按钮：仅在采集进行中可用
        if hasattr(self, 'record_btn'):
            self.record_btn.setEnabled(self.test_running)

    def _update_serial_btn(self, connected):
        """根据串口连接状态更新串口按钮"""
        if connected:
            self._set_btn_style(self.serial_btn, "断通信", "#FF9800")
        else:
            self._set_btn_style(self.serial_btn, "通信", "#2196F3")
        self.serial_btn.setEnabled(True)

    def _update_ts_btn(self, started=False):
        """根据温度源状态更新开启/关闭控制按钮"""
        if started:
            self._set_btn_style(self.ts_btn, "关闭", "#f44336")
        else:
            self._set_btn_style(self.ts_btn, "开启", "#4CAF50")
        self.ts_btn.setEnabled(True)

    def _update_seq_btn(self):
        """根据顺序测试运行状态更新顺序测试按钮"""
        if self.sequential_running and self.sequential_mode == 'forward':
            self._set_btn_style(self.seq_btn, "停止顺序", "#f44336")
        else:
            self._set_btn_style(self.seq_btn, "顺序测试", "#9C27B0")
        self.seq_btn.setEnabled(True)

    def _update_loop_btn(self):
        """根据循环测试运行状态更新循环测试按钮"""
        if self.sequential_running and self.sequential_mode == 'loop':
            self._set_btn_style(self.loop_btn, "停止循环", "#f44336")
        else:
            self._set_btn_style(self.loop_btn, "循环测试", "#FF9800")
        self.loop_btn.setEnabled(True)

    def _update_reverse_btn(self):
        """根据倒序测试运行状态更新倒序测试按钮"""
        if self.sequential_running and self.sequential_mode == 'reverse':
            self._set_btn_style(self.reverse_btn, "停止倒序", "#f44336")
        else:
            self._set_btn_style(self.reverse_btn, "倒序测试", "#607D8B")
        self.reverse_btn.setEnabled(True)

    def _toggle_collection(self):
        """切换采集开始/停止"""
        if self.test_running:
            self.stop_collection()
        else:
            self.start_collection()

    def _toggle_axial_collection(self):
        """切换轴向/径向测试采集。文件名 sp-{axis/radial}-时间.xlsx，功能同手动测试开始测试"""
        mode_cn = "径向" if self.axial_mode == 'radial' else "轴向"
        file_tag = "radial" if self.axial_mode == 'radial' else "axis"
        if self.test_running:
            self.stop_collection()
            self.axial_start_btn.setText(f"开始{mode_cn}测试")
            self._current_mode_tag = None
        else:
            self.start_collection(mode_tag=file_tag)
            self.axial_start_btn.setText(f"停止{mode_cn}测试")

    def _axial_record_data(self):
        """轴向/径向测试的"记录数据"：保存到 test data 目录，文件名与自动保存一致。
        sp-axis-时间.xlsx 或 sp-radial-时间.xlsx。不弹文件对话框。"""
        try:
            if not self.test_running:
                QMessageBox.warning(self, "警告", "请先开始测试后再记录数据")
                return
            has_data = any(len(self.data_buffer[i]) > 0 for i in range(self._dev_row_count))
            if not has_data:
                QMessageBox.warning(self, "警告", "无数据可保存")
                return
            # 文件名：sp-axis-时间 / sp-radial-时间（与自动保存一致）
            temp_str = self.get_filename_temp_str()
            if temp_str == 'NA':
                try:
                    temp_str = f"{float(self.current_test_temp):.1f}"
                except (TypeError, ValueError):
                    temp_str = 'NA'
            file_tag = 'radial' if self.axial_mode == 'radial' else 'axis'
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = f"{temp_str}-{file_tag}-{ts}.xlsx"
            script_dir = os.path.dirname(os.path.abspath(__file__))
            save_dir = os.path.join(script_dir, "test data")
            os.makedirs(save_dir, exist_ok=True)
            filepath = os.path.join(save_dir, base_name)
            # 直接调用自动保存逻辑，但写入指定文件
            self._save_to_file(filepath)
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "错误", f"记录数据失败：{str(e)}")

    def _save_to_file(self, filepath):
        """将当前实时数据 + stability + 轴向/径向结果写入指定 Excel 文件。"""
        try:
            has_data = any(len(self.data_buffer[i]) > 0 for i in range(self._dev_row_count))
            if not has_data:
                return
            max_len = max((len(self.datetime_buffer[i]) for i in range(self._dev_row_count) if self.devices[i]['enabled']), default=0)
            df = pd.DataFrame()
            times = None
            for i in range(self._dev_row_count):
                if self.devices[i]['enabled'] and len(self.datetime_buffer[i]) > 0:
                    times = [x.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3] for x in self.datetime_buffer[i]]
                    break
            if times:
                if len(times) < max_len:
                    times += [''] * (max_len - len(times))
                else:
                    times = times[:max_len]
                df["采集时间"] = times
            for i in range(self._dev_row_count):
                if self.devices[i]['enabled']:
                    name = self.devices[i]['name']
                    unit = self.device_quantity_info.get(i, {}).get('unit', '°C')
                    col_name = f"{name} ({unit})" if unit else name
                    data = list(self.data_buffer[i])[:max_len]
                    while len(data) < max_len:
                        data.append(None)
                    df[col_name] = data
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='实时数据', index=False)
                if self.auto_test_summary:
                    pd.DataFrame([self.auto_test_summary]).to_excel(writer, sheet_name='stability', index=False)
            # 轴向/径向结果写入对应 sheet（axis / radial）
            if getattr(self, 'axial_data', None):
                self._save_axis_sheet_to(filepath)
            QMessageBox.information(self, "成功", f"数据已记录至：{filepath}")
            self.status_label.setText(f"数据已记录: {filepath}")
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "错误", f"保存失败：{str(e)}")

    def _save_axis_sheet_to(self, filepath):
        """将轴向/径向结果写入指定文件的 axis/radial sheet（不弹窗）。"""
        if not self.axial_data:
            return
        from openpyxl import load_workbook
        sheet_name = 'radial' if self.axial_mode == 'radial' else 'axis'
        data = {'指标': ['F-avg', 'M-avg']}
        for c in self.axial_columns:
            mode = c.get('mode', self.axial_mode)
            item0 = self.axial_table.item(0, c['col'])
            cell_text = str(item0.text()).strip() if item0 else ''
            if mode == 'axial':
                try:
                    key = round(float(cell_text.replace('mm', '').strip()), 1)
                except (TypeError, ValueError):
                    key = c['key']
                col_label = f"{key:.1f}mm"
            else:
                key = cell_text or c['key']
                col_label = str(key)
            rec = self.axial_data.get(c['key'], {})
            data[col_label] = [
                None if rec.get('F') is None else round(rec['F'], 4),
                None if rec.get('M') is None else round(rec['M'], 4),
            ]
        df = pd.DataFrame(data)
        wb = load_workbook(filepath)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        ws.append(list(df.columns))
        for _, row in df.iterrows():
            ws.append([None if pd.isna(v) else v for v in row.tolist()])
        wb.save(filepath)

    def _toggle_serial(self):
        """切换串口开/关"""
        if self.temp_source_connected:
            self._close_temp_source_port()
        else:
            self._open_temp_source_port()

    def _toggle_ts(self):
        """切换温度源启动/停止"""
        is_const1210 = (self._ts_device_type == 'Const 1210')
        if self.temp_source_connected and self.temp_source_manager is not None:
            if is_const1210:
                # Const 1210: 使用 TEMPerature:STATus? 查询状态 (0=测试, 1=控制)
                status = self._const1210_query_status()
                if status == '1':
                    self.temp_source_stop()
                else:
                    self.temp_source_start()
            else:
                status = self._query_outp_status()
                if status == '1':
                    self.temp_source_stop()
                else:
                    self.temp_source_start()
        else:
            self.temp_source_start()

    def _toggle_seq(self):
        """切换顺序测试开始/停止"""
        if self.sequential_running:
            self._stop_sequential_test()
        else:
            self.sequential_mode = 'forward'
            self._start_sequential_test()

    def _toggle_loop(self):
        """切换循环测试开始/停止"""
        if self.sequential_running:
            self._stop_sequential_test()
        else:
            self.sequential_mode = 'loop'
            self._start_sequential_test()

    def _toggle_reverse(self):
        """切换倒序测试开始/停止"""
        if self.sequential_running:
            self._stop_sequential_test()
        else:
            self.sequential_mode = 'reverse'
            self._start_sequential_test()

    def _update_sched_btn(self):
        """根据定时状态更新定时/取消定时按钮"""
        if self.sched_test_armed:
            self._set_btn_style(self.sched_test_btn, "取消定时", "#f44336")
            self.sched_test_btn.setEnabled(True)
        else:
            self._set_btn_style(self.sched_test_btn, "定时", "#607D8B")
            self.sched_test_btn.setEnabled(
                not self.sequential_running and not self.sched_test_armed
            )

    def _toggle_scheduled_test(self):
        """切换定时设置/取消"""
        if self.sched_test_armed:
            self._cancel_scheduled_test()
        else:
            self._arm_scheduled_test()

    # ========== 根据命令确定物理量类型 ==========
    def determine_quantity_type(self, command):
        """根据读取命令确定物理量类型，返回 (quantity_type, display_name, unit)"""
        cmd = command.upper().replace('\r', '').replace('\n', '').strip()
        if 'CONTROL' in cmd:
            # 解析命令末尾的数字，区分风扇功率、加热功率等
            idx_match = re.search(r'CONTROL\S*\s+(\d+)', cmd)
            if idx_match and int(idx_match.group(1)) == 3:
                return 'FanPower', 'Fan Power', ''
            elif idx_match and int(idx_match.group(1)) == 4:
                return 'HeatingPower', 'Heating Power', ''
            return 'FanPower', 'Fan Power', ''
        elif 'TEMPERATURE' in cmd:
            return 'Temperature', 'Temperature', '°C'
        elif 'TEMP' in cmd and 'SOUR' in cmd:
            # SOUR:SENS:DATA? TEMP1/TEMP2 → Fluke 9250 传感器温度
            return 'Temperature', 'Temperature', '°C'
        elif 'OUTP' in cmd and 'DATA' in cmd:
            return 'DutyCycle', 'Duty Cycle', ''
        elif 'ELEC' in cmd:
            # 电流/电压类测量不显示单位
            return 'Electricity', 'Electricity', ''
        else:
            # 未知命令不显示单位，避免误导（如Fan通道用非标命令时显示°C）
            return 'Unknown', 'Unknown', ''

    # ========== 设置多Y轴绘图区域 ==========
    def setup_multi_axis_plot(self):
        """设置多Y轴绘图：温度作为主Y轴（左侧），其他物理量在右侧Y轴，网格只跟随主Y轴"""
        # 先清理之前直接添加到scene的额外ViewBox
        for vb in getattr(self, 'extra_vb_list', []):
            try:
                self.plot_widget.scene().removeItem(vb)
            except:
                pass

        self.plot_widget.clear()
        self.extra_vb_list = []
        self.quantity_viewbox_map = {}
        self.device_viewboxes = {}

        # 确定各启用设备的物理量类型
        self.device_quantity_info = {}
        quantity_map = {}  # quantity_type -> {'display_name', 'unit', 'devices', 'color'}

        for i in range(self._dev_row_count):
            if self.devices[i]['enabled']:
                qt, display_name, unit = self.determine_quantity_type(self.devices[i].get('read_command', ''))
                self.device_quantity_info[i] = {
                    'quantity_type': qt,
                    'display_name': display_name,
                    'unit': unit
                }
                if qt not in quantity_map:
                    quantity_map[qt] = {
                        'display_name': display_name,
                        'unit': unit,
                        'devices': [i],
                        'color': self.device_colors[i]
                    }
                else:
                    quantity_map[qt]['devices'].append(i)

        if not quantity_map:
            quantity_map['Temperature'] = {
                'display_name': '温度', 'unit': '°C', 'devices': [], 'color': '#1f77b4'
            }

        qt_list = list(quantity_map.keys())

        # 优先将温度作为主Y轴
        if 'Temperature' in quantity_map and len(quantity_map) > 1:
            qt_list.remove('Temperature')
            qt_list.insert(0, 'Temperature')
        elif 'Power' in quantity_map and len(quantity_map) > 1:
            qt_list.remove('Power')
            qt_list.insert(0, 'Power')

        primary_qt = qt_list[0]
        primary_info = quantity_map[primary_qt]

        # ===== 创建主绘图区域（温度在左侧Y轴）=====
        self.primary_plot = self.plot_widget.addPlot()

        if primary_info['unit']:
            y_label = f"{primary_info['display_name']}({primary_info['unit']})"
        else:
            y_label = primary_info['display_name']
        self.primary_plot.setLabel('left', y_label)
        self.primary_plot.setLabel('bottom', '时间(min)')
        # 网格只跟随主Y轴变化
        self.primary_plot.showGrid(x=True, y=True, alpha=0.3)

        # 设置左侧Y轴颜色
        primary_color = primary_info['color']
        self.primary_plot.getAxis('left').setPen(pg.mkPen(primary_color, width=2))
        self.primary_plot.getAxis('left').setTextPen(pg.mkPen(primary_color))
        # 禁用SI前缀自动缩放，显示实际数值
        self.primary_plot.getAxis('left').enableAutoSIPrefix(False)
        self.primary_plot.getAxis('bottom').enableAutoSIPrefix(False)
        # 用户手动缩放/平移时关闭自动跟随
        self.primary_plot.vb.sigRangeChanged.connect(self._on_view_range_changed)

        # 初始化曲线
        self.curves = [None] * self._dev_row_count

        # 为主轴（左侧Y轴）的设备创建曲线
        for dev_id in primary_info['devices']:
            color = self.device_colors[dev_id]
            curve = self.primary_plot.plot([], [], pen=pg.mkPen(color, width=2),
                                           symbol='o', symbolSize=2.5,
                                           symbolBrush=pg.mkBrush(color),
                                           symbolPen=pg.mkPen(color, width=0.5))
            self.curves[dev_id] = curve
            self.device_viewboxes[dev_id] = self.primary_plot.vb
            # 根据曲线勾选状态设置初始可见性
            curve.setVisible(self.devices[dev_id].get('curve_visible', True))

        self.quantity_viewbox_map[primary_qt] = self.primary_plot.vb

        # ===== 为其他物理量创建右侧Y轴 =====
        for idx, qt in enumerate(qt_list[1:]):
            info = quantity_map[qt]
            axis_color = info['color']

            # 创建独立ViewBox（不显示网格，只跟随主Y轴的X网格）
            vb = pg.ViewBox()
            self.primary_plot.scene().addItem(vb)
            vb.setXLink(self.primary_plot)
            vb.enableAutoRange(axis=pg.ViewBox.YAxis)

            if idx == 0:
                # 第一个额外轴：使用PlotItem内置的右侧轴
                self.primary_plot.showAxis('right')
                axis = self.primary_plot.getAxis('right')
                if info['unit']:
                    axis_label = f"{info['display_name']}({info['unit']})"
                else:
                    axis_label = info['display_name']
                axis.setLabel(axis_label)
                axis.linkToView(vb)
                axis.setPen(pg.mkPen(axis_color, width=2))
                axis.setTextPen(pg.mkPen(axis_color))
                # 隐藏副Y轴刻度线和网格，只保留刻度数值
                axis.setStyle(tickLength=0)
                axis.setGrid(False)
                axis.enableAutoSIPrefix(False)
                axis.enableAutoSIPrefix(False)
            else:
                # 第3个及以后的物理量：创建新的AxisItem并添加到布局右侧（偏移避免交叠）
                axis = pg.AxisItem('right')
                if info['unit']:
                    axis_label = f"{info['display_name']}({info['unit']})"
                else:
                    axis_label = info['display_name']
                axis.setLabel(axis_label)
                axis.setPen(pg.mkPen(axis_color, width=2))
                axis.setTextPen(pg.mkPen(axis_color))
                # 隐藏副Y轴刻度线和网格，只保留刻度数值
                axis.setStyle(tickLength=0)
                axis.setGrid(False)
                axis.enableAutoSIPrefix(False)

                # 在PlotItem布局中添加新列（向右偏移，避免交叠）
                col = 3 + (idx - 1)
                self.primary_plot.layout.addItem(axis, 2, col)
                axis.linkToView(vb)
                axis.setWidth(60)

            self.extra_vb_list.append(vb)
            self.quantity_viewbox_map[qt] = vb

            # 为此轴的设备创建曲线
            for dev_id in info['devices']:
                color = self.device_colors[dev_id]
                curve = pg.PlotDataItem(pen=pg.mkPen(color, width=2),
                                        symbol='o', symbolSize=2.5,
                                        symbolBrush=pg.mkBrush(color),
                                        symbolPen=pg.mkPen(color, width=0.5))
                vb.addItem(curve)
                self.curves[dev_id] = curve
                self.device_viewboxes[dev_id] = vb
                # 根据曲线勾选状态设置初始可见性
                curve.setVisible(self.devices[dev_id].get('curve_visible', True))

        # 处理ViewBox大小同步，确保多Y轴对齐且不交叠
        def update_views():
            for vb in self.extra_vb_list:
                vb.setGeometry(self.primary_plot.vb.sceneBoundingRect())

        if self.extra_vb_list:
            self.primary_plot.vb.sigResized.connect(update_views)
            update_views()

        # 更新图例
        self.legend_widget.clear()
        for i in range(self._dev_row_count):
            color = self.device_colors[i]
            unit = self.device_quantity_info.get(i, {}).get('unit', '°C')
            self.legend_widget.add_device(i, self.devices[i]['name'], color, unit, self.devices[i].get('auto_test', False))
        # 根据启用状态同步图例显示：未启用的通道（设备）不显示在右侧实时数据中
        for i in range(self._dev_row_count):
            self.legend_widget.set_device_visible(i, self.devices[i]['enabled'])

    def on_connection_changed(self, dev_id, conn_type):
        if hasattr(self, '_loading') and self._loading:
            return
        w = self.device_widgets[dev_id]
        if conn_type == 'serial':
            w['port_ip'].setCurrentIndex(0)
            self._populate_port_combo(w['port_combo'], self.devices[dev_id].get('port', ''))
            w['baud_port'].setPlaceholderText("波特率")
            w['baud_port'].setText(self.devices[dev_id].get('baudrate', '9600'))
        else:
            w['port_ip'].setCurrentIndex(1)
            w['ip_edit'].setText(self.devices[dev_id].get('host', ''))
            w['baud_port'].setPlaceholderText("端口号")
            w['baud_port'].setText(str(self.devices[dev_id].get('lan_port', '8000')))
        self.update_device_config(dev_id)

    def update_device_config(self, dev_id):
        w = self.device_widgets[dev_id]
        enabled = w['enable'].isChecked()
        was_enabled = self.devices[dev_id]['enabled']
        conn = w['connection'].currentText()
        self.devices[dev_id]['enabled'] = enabled
        self.devices[dev_id]['connection'] = conn
        # 实时显示/隐藏图例
        if hasattr(self, 'legend_widget'):
            self.legend_widget.set_device_visible(dev_id, enabled)
        # 实时同步曲线显示（启用状态变化立即反映到曲线）
        self._update_curve_visibility()
        # 在采集过程中，动态启停设备线程
        if self.test_running and was_enabled != enabled:
            if enabled:
                # 启用设备 → 启动采集线程
                d = self.devices[dev_id]
                interval = self.interval_spin.value() / 1000.0
                d['read_interval'] = interval
                t = DeviceThread(dev_id, d)
                t.data_received.connect(self.on_data_received)
                t.connection_status.connect(self.on_connection_status)
                t.debug_info.connect(self.on_debug_info)
                t.start()
                self.device_threads.append(t)
            else:
                # 禁用设备 → 停止对应线程
                for t in self.device_threads[:]:
                    if t.device_id == dev_id:
                        t.stop()
                        self.device_threads.remove(t)
        self.devices[dev_id]['name'] = w['name'].text()
        if conn == 'serial':
            self.devices[dev_id]['port'] = w['port_combo'].currentText()
            self.devices[dev_id]['baudrate'] = w['baud_port'].text()
            self.devices[dev_id]['host'] = ''
            self.devices[dev_id]['lan_port'] = ''
        else:
            self.devices[dev_id]['host'] = w['ip_edit'].text()
            port_str = w['baud_port'].text()
            if not port_str.strip():
                port_str = '8000'
                w['baud_port'].setText('8000')
            self.devices[dev_id]['lan_port'] = port_str
            self.devices[dev_id]['port'] = ''
            self.devices[dev_id]['baudrate'] = ''
        self.devices[dev_id]['read_command'] = w['cmd'].text()
        # auto_test 由 _on_auto_test_toggled 单独更新，但此处做兜底
        if 'auto_test' in w and not (hasattr(self, '_loading') and self._loading):
            self.devices[dev_id]['auto_test'] = w['auto_test'].isChecked()
        # 命令变更时更新图例单位
        if hasattr(self, 'legend_widget'):
            _, _, unit = self.determine_quantity_type(self.devices[dev_id]['read_command'])
            self.legend_widget.update_temperature(dev_id, None, unit)
        self.save_config()

    def _get_device_manager(self, dev_id):
        """获取指定设备的串口管理器"""
        d = self.devices[dev_id]
        if d['connection'] == 'lan':
            return None
        port = d.get('port', '')
        baudrate = d.get('baudrate', '9600')
        if not port:
            return None
        return SharedSerialManager(port, baudrate)

    def _sync_device_status(self, port, connected):
        """同步同一串口下所有设备的状态显示，同步温度源控制串口按钮"""
        for i in range(self._dev_row_count):
            d = self.devices[i]
            if d['connection'] == 'serial' and d.get('port', '') == port:
                if connected:
                    self.device_widgets[i]['status'].setText("已连接")
                    self.device_widgets[i]['status'].setStyleSheet("color:green;font-weight:bold;")
                else:
                    self.device_widgets[i]['status'].setText("未连接")
                    self.device_widgets[i]['status'].setStyleSheet("color:#888888;")
        # 同步温度源串口按钮状态
        if port and hasattr(self, 'temp_source_port_combo'):
            ts_port = self.temp_source_port_combo.currentText().strip()
            if port.lower() == ts_port.lower():
                self._update_serial_button_state(connected)

    def _open_device_serial(self, dev_id):
        """单独打开指定设备的串口（同步同端口所有设备）"""
        if not SERIAL_AVAILABLE:
            QMessageBox.warning(self, "错误", "串口库不可用")
            return
        d = self.devices[dev_id]
        if d['connection'] != 'serial':
            self.status_label.setText(f"设备{dev_id+1}: LAN连接不支持单独开关")
            return
        manager = self._get_device_manager(dev_id)
        if manager is None:
            QMessageBox.warning(self, "错误", f"设备{dev_id+1}串口参数不完整")
            return
        port = d.get('port', '')
        if manager.connect():
            self._sync_device_status(port, True)
            self.status_label.setText(f"串口 {port} 已打开")
        else:
            QMessageBox.warning(self, "错误", f"串口 {port} 打开失败")

    def _close_device_serial(self, dev_id):
        """单独关闭指定设备的串口（同步同端口所有设备）"""
        d = self.devices[dev_id]
        if d['connection'] != 'serial':
            return
        manager = self._get_device_manager(dev_id)
        port = d.get('port', '')
        if manager is not None:
            manager.disconnect()
        self._sync_device_status(port, False)
        self.status_label.setText(f"串口 {port} 已关闭")

    def refresh_ports(self):
        if SERIAL_AVAILABLE:
            port_list = [p.device for p in serial.tools.list_ports.comports()]
            current = self.temp_source_port_combo.currentText()
            self.temp_source_port_combo.clear()
            self.temp_source_port_combo.addItems(port_list if port_list else [''])
            if current:
                idx = self.temp_source_port_combo.findText(current)
                if idx >= 0:
                    self.temp_source_port_combo.setCurrentIndex(idx)
                else:
                    # 非可编辑模式 setCurrentText 失效，把端口加到列表首位
                    if not self.temp_source_port_combo.isEditable():
                        self.temp_source_port_combo.insertItem(0, current)
                        self.temp_source_port_combo.setCurrentIndex(0)
                    else:
                        self.temp_source_port_combo.setCurrentText(current)
            self.status_label.setText(f"检测到串口: {', '.join(port_list) if port_list else '无'}")
        else:
            self.status_label.setText("串口功能不可用，请安装pyserial库")

    @staticmethod
    def _populate_port_combo(combo, desired_port=''):
        """向 QComboBox 填入系统实际可用串口列表，并尝试选中 desired_port
           非可编辑模式下：若 desired_port 不在列表中，也加入列表首位以便显示"""
        combo.clear()
        port_list = []
        if SERIAL_AVAILABLE:
            try:
                port_list = [p.device for p in serial.tools.list_ports.comports()]
            except:
                pass
        if port_list:
            combo.addItems(port_list)
        if desired_port:
            idx = combo.findText(desired_port)
            if idx >= 0:
                combo.setCurrentIndex(idx)
            else:
                # 非可编辑模式 setCurrentText 失效，把端口加到列表首位
                if not combo.isEditable() and desired_port:
                    combo.insertItem(0, desired_port)
                    combo.setCurrentIndex(0)
                else:
                    combo.setCurrentText(desired_port)

    def _find_tc_main_device(self):
        """查找 TC_Main_20mm-SIC 设备，返回 (port, baudrate)
           优先按设备名查找，再按端口名(com11)查找"""
        # 方式1: 按设备名查找
        for d in self.devices:
            if 'TC_Main' in d['name'] or 'TC_Main_20mm-SIC' in d['name']:
                return d.get('port', ''), d.get('baudrate', '115200')
        # 方式2: 按端口名查找
        for d in self.devices:
            if d.get('port', '').lower() == 'com11':
                return d.get('port', ''), d.get('baudrate', '115200')
        return '', '115200'

    def _send_temp_source_cmd(self, cmd):
        """通过 SharedSerialManager 发送命令（共享串口，避免端口冲突）"""
        if self.temp_source_manager is None:
            return False
        try:
            cmd_str = cmd.decode() if isinstance(cmd, bytes) else cmd
            self.temp_source_manager.send_command(cmd_str, timeout=2.0)
            time.sleep(0.1)
            return True
        except Exception as e:
            print(f"发送命令失败: {e}")
            return False

    def _read_temp_source_response(self, timeout=1.0):
        """SharedSerialManager.send_command 已内置读取，此处不需要额外读取"""
        return ''

    def _send_and_verify_cmd(self, set_cmd, query_cmd, expected_value, max_retries=3, label=''):
        """发送设置命令 → 回读验证 → 不一致则重试，返回 True 表示成功
           验证逻辑: TARGet? 返回值 / 100 后与期望温度对比（设备返回的是缩放值如2500=25.00°C）"""
        if self.temp_source_manager is None:
            return False
        exp_val_f = float(expected_value)
        for attempt in range(max_retries):
            cmd_str = set_cmd.decode() if isinstance(set_cmd, bytes) else set_cmd
            self.temp_source_manager.send_command(cmd_str, timeout=2.0)
            time.sleep(0.3)
            q_cmd = query_cmd.decode() if isinstance(query_cmd, bytes) else query_cmd
            response = self.temp_source_manager.send_command(q_cmd, timeout=2.0)
            # DEBUG: 打印回读原始响应
            from datetime import datetime
            ts = datetime.now().strftime('%H:%M:%S.%f')[:-3]
            print(f"[{ts}] _send_and_verify_cmd[{label}][try{attempt+1}] 原始响应={repr(response)}")
            if response:
                try:
                    raw = response.strip()
                    # 设备返回可能含非数字前缀/单位，如 'Pm 0 Tm 33.0002 R 112.8344'
                    # 提取响应中所有浮点数字，逐个与期望值比较（直接匹配或 /100 缩放）
                    tokens = re.findall(r'-?\d+\.?\d*', raw)
                    matched = False
                    for tk in tokens:
                        try:
                            read_val = float(tk)
                        except ValueError:
                            continue
                        if abs(read_val - exp_val_f) < 0.01:
                            print(f"  → 直接匹配成功 (token={read_val}, 期望={exp_val_f})")
                            matched = True
                            break
                        if abs(read_val / 100.0 - exp_val_f) < 0.01:
                            print(f"  → /100匹配成功 (token={read_val}, 期望={exp_val_f})")
                            matched = True
                            break
                    if matched:
                        return True
                    print(f"  → 均不匹配: 响应数字={tokens}, 期望={exp_val_f}")
                except (ValueError, IndexError) as e:
                    print(f"  → 解析异常: {e}")
                    read_val = None
            else:
                print(f"  → 响应为空")
            self.status_label.setText(f"{label}设置不一致(期望={exp_val_f}, 读取={response or '空'}), 重试({attempt+1}/{max_retries})")
            QApplication.processEvents()
            time.sleep(0.2)
        QMessageBox.warning(self, "警告", f"{label}设置验证失败: 已重试{max_retries}次仍未读取到正确值")
        return False

    def _send_and_verify_weight(self, weight_str, row_idx):
        """发送并回读验证 Weight 参数（字符串比较，已统一为英文逗号）"""
        if self.temp_source_manager is None:
            return False
        label = f"行{row_idx+1} Weight"
        # 统一中文逗号 -> 英文逗号
        weight_str = weight_str.replace('，', ',')
        for attempt in range(3):
            self.temp_source_manager.send_command(f"SOUR:WEIGht {weight_str}\r\n", timeout=2.0)
            time.sleep(0.3)
            resp = self.temp_source_manager.send_command("SOUR:WEIGht?\r\n", timeout=2.0)
            if resp:
                # 统一设备返回的逗号风格再比较
                resp_normalized = resp.strip().replace('，', ',')
                if resp_normalized == weight_str:
                    return True
            self.status_label.setText(f"{label}不一致(设={weight_str}, 读={resp or '空'}), 重试({attempt+1}/3)")
            QApplication.processEvents()
            time.sleep(0.2)
        QMessageBox.warning(self, "警告", f"{label}设置验证失败: 已重试3次仍未读取到正确值")
        return False

    def _on_temp_source_conn_changed(self, conn_type):
        """温度源连接类型切换时，显示/隐藏对应参数输入框"""
        is_serial = (conn_type == 'serial')
        self.temp_source_port_combo.setVisible(is_serial)
        self.temp_source_baud_spin.setVisible(is_serial)
        self.temp_source_ip_edit.setVisible(not is_serial)
        self.temp_source_lan_port_spin.setVisible(not is_serial)
        # 记住当前设备的通讯方式偏好
        if self._ts_device_type:
            self._ts_device_conn_prefs[self._ts_device_type] = conn_type
        self.save_config()

    def _save_current_ts_settings(self):
        """抓取当前设备类型的完整温度源参数（行设置 + 波动阈值）到 _ts_device_full_settings"""
        n = self._ts_row_count
        settings = {
            'row_count': n,
            'stability_threshold': self.stability_threshold_spin.value(),
            'row_setpoints': [self.row_setpoint_spins[i].value() for i in range(n)],
            'row_specs': [self.row_setpoint_spec[i].value() for i in range(n)],
            'row_mains': [self.row_main_spins[i].value() for i in range(n)],
            'row_secs': [self.row_sec_spins[i].value() for i in range(n)],
            'row_checks': [self.row_checks[i].isChecked() for i in range(n)],
            'row_main_pid_p': [self.row_main_pid_p[i].value() for i in range(n)],
            'row_main_pid_i': [self.row_main_pid_i[i].value() for i in range(n)],
            'row_main_pid_d': [self.row_main_pid_d[i].value() for i in range(n)],
            'row_sec_pid_p': [self.row_sec_pid_p[i].value() for i in range(n)],
            'row_sec_pid_i': [self.row_sec_pid_i[i].value() for i in range(n)],
            'row_sec_pid_d': [self.row_sec_pid_d[i].value() for i in range(n)],
            'row_weights': [','.join(w.text() for w in self.row_weights[i]) for i in range(n)],
        }
        self._ts_device_full_settings[self._ts_device_type] = settings

    def _apply_ts_settings(self, settings):
        """将某设备类型的完整温度源参数恢复到 UI"""
        if not settings:
            return
        n = settings.get('row_count', 3)
        # 调整行数：需要更多行则添加，需要更少行则删除
        while self._ts_row_count < n:
            self._ts_add_row()
        while self._ts_row_count > n and self._ts_row_count > 1:
            self._ts_remove_row(silent=True)
        n = self._ts_row_count
        if 'stability_threshold' in settings:
            self.stability_threshold_spin.setValue(settings['stability_threshold'])
        self._set_row_lists('row_setpoints', settings.get('row_setpoints'))
        self._set_row_lists('row_specs', settings.get('row_specs'))
        self._set_row_lists('row_mains', settings.get('row_mains'))
        self._set_row_lists('row_secs', settings.get('row_secs'))
        self._set_row_lists('row_checks', settings.get('row_checks'), is_check=True)
        self._set_row_lists('row_main_pid_p', settings.get('row_main_pid_p'))
        self._set_row_lists('row_main_pid_i', settings.get('row_main_pid_i'))
        self._set_row_lists('row_main_pid_d', settings.get('row_main_pid_d'))
        self._set_row_lists('row_sec_pid_p', settings.get('row_sec_pid_p'))
        self._set_row_lists('row_sec_pid_i', settings.get('row_sec_pid_i'))
        self._set_row_lists('row_sec_pid_d', settings.get('row_sec_pid_d'))
        if 'row_weights' in settings:
            for i, v in enumerate(settings['row_weights']):
                if i >= n:
                    break
                parts = v.split(',')
                for j in range(min(len(parts), len(self.row_weights[i]))):
                    self.row_weights[i][j].setText(parts[j])

    def _set_row_lists(self, attr, values, is_check=False):
        """按列表恢复某行控件序列的值"""
        if not values:
            return
        arr = getattr(self, attr, [])
        for i, v in enumerate(values):
            if i >= len(arr):
                break
            if is_check:
                arr[i].setChecked(bool(v))
            else:
                try:
                    arr[i].setValue(float(v))
                except (ValueError, TypeError):
                    pass

    def _on_stability_threshold_changed(self, value):
        """稳定性判断条件（波动阈值）变更回调"""
        self.stability_threshold = value
        if not self._loading:
            self.save_config()

    def _on_temp_source_device_changed(self, device_type):
        """温度源设备类型切换回调：保存旧设备全部设置，恢复新设备全部设置"""
        # 保存当前设备的完整参数（连接参数 + 行设置 + 波动阈值）到旧设备
        old_device = self._ts_device_type
        if old_device:
            self._ts_device_conn_prefs[old_device] = self.temp_source_conn_combo.currentText()
            self._ts_device_settings[old_device] = {
                'port': self.temp_source_port_combo.currentText(),
                'baud': self.temp_source_baud_spin.value(),
                'ip': self.temp_source_ip_edit.text(),
                'lan_port': self.temp_source_lan_port_spin.value(),
            }
            self._save_current_ts_settings()
        # 更新当前设备类型
        self._ts_device_type = device_type
        # 恢复新设备的连接参数
        pref_conn = self._ts_device_conn_prefs.get(device_type, 'serial')
        self.temp_source_conn_combo.setCurrentText(pref_conn)
        settings = self._ts_device_settings.get(device_type, {})
        if 'port' in settings:
            port_text = settings['port']
            idx = self.temp_source_port_combo.findText(port_text)
            if idx >= 0:
                self.temp_source_port_combo.setCurrentIndex(idx)
            else:
                if not self.temp_source_port_combo.isEditable() and port_text:
                    self.temp_source_port_combo.insertItem(0, port_text)
                    self.temp_source_port_combo.setCurrentIndex(0)
                else:
                    self.temp_source_port_combo.setCurrentText(port_text)
        if 'baud' in settings:
            self.temp_source_baud_spin.setValue(settings['baud'])
        if 'ip' in settings:
            self.temp_source_ip_edit.setText(settings['ip'])
        if 'lan_port' in settings:
            self.temp_source_lan_port_spin.setValue(settings['lan_port'])
        # 恢复新设备的完整行设置 + 波动阈值（每个设备独立记忆）
        self._apply_ts_settings(self._ts_device_full_settings.get(device_type))
        # 切换连接类型时同步控件可见性
        self._on_temp_source_conn_changed(pref_conn)
        self.save_config()
    def _get_temp_source_manager(self, conn_type=None):
        """根据连接类型创建并返回串口/LAN管理器"""
        if conn_type is None:
            conn_type = self.temp_source_conn_combo.currentText()
        if conn_type == 'lan':
            ip = self.temp_source_ip_edit.text().strip()
            port = self.temp_source_lan_port_spin.value()
            if not ip:
                QMessageBox.warning(self, "错误", "请先输入IP地址")
                return None
            return EthernetManager(ip, port)
        else:
            if not SERIAL_AVAILABLE:
                QMessageBox.warning(self, "错误", "串口库不可用")
                return None
            port = self.temp_source_port_combo.currentText().strip()
            baudrate = str(self.temp_source_baud_spin.value())
            if not port:
                QMessageBox.warning(self, "错误", "请先输入串口号")
                return None
            return SharedSerialManager(port, baudrate)

    def _open_temp_source_port(self):
        """打开温度源连接（按钮回调），连接后立刻开始读取 Main/Sec 读数并更新启停按钮"""
        if self._open_temp_source_serial():
            # 连接成功后立即启动后台读数线程（不依赖"启动控制"），开始读取 Main/Sec
            self._update_serial_button_state(True)
            # 查询实际输出状态，仅用于动态切换启停按钮外观
            # Fluke 9250: OUTP:STAT? / Const 1210: TEMPerature:STATus?
            self._update_temp_source_ui_state()
            self.status_label.setText("温度源连接已打开，开始读取 Main/Sec")

            # 主动诊断一次 Main 读数，立即在状态栏反馈命令/响应是否正确
            def _diagnostic_main_read():
                if not self.temp_source_connected or self.temp_source_manager is None:
                    return
                if self._ts_device_type == 'Const 1210':
                    return
                cmd = "SOUR:SENS:DATA? TEMP1\r\n"
                try:
                    resp = self.temp_source_manager.send_command(cmd, timeout=1.0)
                    print(f"[TS diag] {cmd!r} -> {resp!r}")
                    if resp:
                        m = re.search(r'[+-]?\d+\.?\d*', resp)
                        if m:
                            val = float(m.group())
                            self.status_label.setText(f"温度源 Main 读数: {val:.2f}°C，开始后台轮询")
                        else:
                            self.status_label.setText(f"温度源响应无法解析: {resp!r}")
                    else:
                        self.status_label.setText("温度源无读数返回，请确认设备类型和命令是否正确")
                except Exception as e:
                    self.status_label.setText(f"温度源诊断读数异常: {e}")
            QTimer.singleShot(600, _diagnostic_main_read)

    def _close_temp_source_port(self):
        """关闭温度源串口（按钮回调）"""
        # 取消定时测试（如有）
        if self.sched_test_armed:
            self.sched_test_armed = False
            self.sched_test_timer.stop()
        self.sched_status_label.setText("")
        self._update_sched_btn()
        self.sched_time_edit.setEnabled(True)
        # 如果正在顺序测试，先停止
        if self.sequential_running:
            self.sequential_timer.stop()
            self.sequential_running = False
        # 先停止输出
        if self.temp_source_manager is not None:
            if self._ts_device_type != 'Const 1210':
                self._send_temp_source_cmd(b"OUTP:STAT 0\r\n")
        self._close_temp_source_serial()
        self._update_serial_btn(False)
        self._update_ts_btn(False)
        self._update_seq_btn()
        self._update_loop_btn()
        self._update_reverse_btn()
        self.status_label.setText("温度源连接已关闭")

    def _update_serial_button_state(self, connected):
        """更新温度源串口控制按钮状态，同时管理当前设备的温度读数线程"""
        is_const1210 = (self._ts_device_type == 'Const 1210')
        self._update_serial_btn(connected)
        self.ts_btn.setEnabled(connected and not self.sequential_running)
        self.seq_btn.setEnabled(connected and not self.sequential_running)
        self.loop_btn.setEnabled(connected and not self.sequential_running)
        self.reverse_btn.setEnabled(connected and not self.sequential_running)
        # 发送按钮：未通信时置灰
        if hasattr(self, 'ts_manual_sp_btn'):
            self.ts_manual_sp_btn.setEnabled(connected)
        for sb in getattr(self, 'row_send_btns', []):
            sb.setEnabled(connected)
        self._update_sched_btn()
        if connected:
            if is_const1210:
                # Const 1210: 启动温度查询线程（MEASure:TEMPerature?）
                self.ts_const1210_temp_label.setText("Temp:--°C")
                if self.ts_const1210_query_thread is None and self.temp_source_manager is not None:
                    self.ts_const1210_query_thread = Const1210QueryThread(self.temp_source_manager)
                    self.ts_const1210_query_thread.temp_ready.connect(lambda t: self.ts_const1210_temp_label.setText(f"Temp:{t:.2f}°C"))
                    self.ts_const1210_query_thread.start()
            # Fluke 9250：不显示温度读数
        else:
            # 断开：仅停止当前设备的查询线程
            if is_const1210:
                if self.ts_const1210_query_thread is not None:
                    self.ts_const1210_query_thread.stop()
                    self.ts_const1210_query_thread = None
                self.ts_const1210_temp_label.setText("Temp:--°C")
            # Fluke 9250：不显示温度读数

    def _open_temp_source_serial(self):
        """获取温度源管理器（支持串口或LAN连接）"""
        conn_type = self.temp_source_conn_combo.currentText()
        # 获取管理器
        mgr = self._get_temp_source_manager(conn_type)
        if mgr is None:
            return False
        self.temp_source_manager = mgr
        if not self.temp_source_manager.connected:
            self.temp_source_connected = self.temp_source_manager.connect()
        else:
            self.temp_source_connected = True
        if self.temp_source_connected:
            self._update_serial_button_state(True)
        return self.temp_source_connected

    def _close_temp_source_serial(self):
        """关闭共享串口连接"""
        if self.temp_source_manager is not None:
            try:
                self.temp_source_manager.disconnect()
            except:
                pass
            self.temp_source_manager = None
        self.temp_source_connected = False
        self._update_serial_button_state(False)

    def _query_outp_status(self):
        """查询 OUTP:STAT? 状态，返回 '1'/'0'/None"""
        if self.temp_source_manager is None:
            return None
        try:
            resp = self.temp_source_manager.send_command("OUTP:STAT?\r\n", timeout=2.0)
            if resp:
                resp = resp.strip()
                if resp in ('0', '1'):
                    return resp
            return None
        except:
            return None

    def _const1210_query_status(self):
        """Const 1210: 查询 TEMPerature:STATus? 状态，返回 '1'(控制)/'0'(测试)/None"""
        if self.temp_source_manager is None:
            return None
        try:
            resp = self.temp_source_manager.send_command("TEMPerature:STATus?\r\n", timeout=2.0)
            if resp:
                resp = resp.strip()
                if resp in ('0', '1'):
                    return resp
            return None
        except:
            return None

    def _update_temp_source_ui_state(self, setter_cmd=None):
        """根据回读结果更新按钮外观
           Fluke 9250: OUTP:STAT? / Const 1210: TEMPerature:STATus?
           setter_cmd: 刚发送的设置命令，用于判断期望状态
        """
        # 先短延时等待设备响应
        time.sleep(0.3)
        status = None
        is_const1210 = (self._ts_device_type == 'Const 1210')
        for _ in range(3):
            if is_const1210:
                status = self._const1210_query_status()
            else:
                status = self._query_outp_status()
            if status is not None:
                break
            time.sleep(0.2)

        if status == '1':
            self._update_ts_btn(True)
        elif status == '0':
            self._update_ts_btn(False)
        return status

    def _send_pid_params(self, row_idx):
        """发送并回读验证指定行的 MAIN、SEC PID 及 Weight 参数
           返回 True 表示全部验证通过"""
        # Const 1210 不需要 PID 参数，直接返回成功
        if self._ts_device_type == 'Const 1210':
            return True
        # 从3个输入框拼接 Weight 值（逗号分隔）
        weight_parts = [w.text().strip().replace('，', ',') for w in self.row_weights[row_idx]]
        weight_str = ','.join(weight_parts).strip(',')
        params = [
            # (命令基,  值,      标签,                 数字标志)
            ("SOUR:LCON:PBAN", self.row_main_pid_p[row_idx].value(), f"行{row_idx+1} Main P", True),
            ("SOUR:LCON:INT",  self.row_main_pid_i[row_idx].value(), f"行{row_idx+1} Main I", True),
            ("SOUR:LCON:DER",  self.row_main_pid_d[row_idx].value(), f"行{row_idx+1} Main D", True),
            ("SOUR:SEC:LCON:PBAN", self.row_sec_pid_p[row_idx].value(), f"行{row_idx+1} Sec P", True),
            ("SOUR:SEC:LCON:INT",  self.row_sec_pid_i[row_idx].value(), f"行{row_idx+1} Sec I", True),
            ("SOUR:SEC:LCON:DER",  self.row_sec_pid_d[row_idx].value(), f"行{row_idx+1} Sec D", True),
        ]
        if weight_str:
            params.append(("SOUR:WEIGht", weight_str, f"行{row_idx+1} Weight", False))

        for cmd_base, val, label, is_numeric in params:
            if is_numeric:
                # 值为 None 表示留空（使用温度源默认值），跳过该条发送
                if val is None:
                    continue
                ok = self._send_and_verify_cmd(
                    set_cmd=f"{cmd_base} {val}\r\n",
                    query_cmd=f"{cmd_base}?\r\n",
                    expected_value=val,
                    label=label
                )
            else:
                self.status_label.setText(f"{label}={val}...")
                QApplication.processEvents()
                ok = self._send_and_verify_weight(val, row_idx)
            if not ok:
                self.status_label.setText(f"{label} 设置失败")
                return False

        self.status_label.setText("PID/Weight 参数全部设置验证通过")
        return True

    def _ensure_temp_source_open(self):
        """确保温度源连接已打开（与 temp_source_start 一致），返回是否成功。"""
        is_const1210 = (self._ts_device_type == 'Const 1210')
        if not is_const1210 and not SERIAL_AVAILABLE:
            QMessageBox.warning(self, "错误", "串口库不可用")
            return False
        if self.temp_source_manager is None:
            if not self._open_temp_source_serial():
                return False
        return True

    def _send_manual_sp(self):
        """发送温度源通信区的 SP 输入框设定值（SOUR:USER:SPO）
           需先点击「通信」建立连接后才能发送"""
        if not self.temp_source_connected or self.temp_source_manager is None:
            QMessageBox.warning(self, "提示", "请先点击「通信」连接温度源后再发送")
            return
        sp = self.ts_manual_sp_spin.value()
        if self._ts_device_type == 'Const 1210':
            ok = self._send_and_verify_cmd(
                set_cmd=f"SOUR:USER:SPO {sp}\r\n",
                query_cmd="SOUR:USER:SPO?\r\n",
                expected_value=sp,
                label="SP"
            )
        else:
            # Fluke 9250 使用 SOUR:USER:SPO（SP 设置）
            ok = self._send_and_verify_cmd(
                set_cmd=f"SOUR:USER:SPO {sp}\r\n",
                query_cmd="SOUR:USER:SPO?\r\n",
                expected_value=sp,
                label="SP"
            )
        if ok:
            self.status_label.setText(f"SP 设置成功: {sp}°C")
            msg = f"SP 发送成功: {sp}°C"
            print(f"[发送] {msg}")
            if hasattr(self, 'legend_widget'):
                self.legend_widget.append_terminal(msg)
            # 手动测试 SP 发送成功弹框提示（自动测试不弹框）
            QMessageBox.information(self, "成功", f"SP 设置成功: {sp}°C")
        else:
            self.status_label.setText("SP 发送失败")
            print("[发送] SP 发送失败")
            if hasattr(self, 'legend_widget'):
                self.legend_widget.append_terminal("SP 发送失败")

    def _send_row_command(self, row_idx):
        """发送指定行的所有参数到温度源：SP、Main、Sec、PID、Weight 并启动输出。
           支持 Const 1210 与 Fluke 9250。
           注意：行勾选框仅用于是否参与自动测试，不影响本方法（手动发送不受勾选限制）。
           需先点击「通信」建立连接后才能发送。"""
        if row_idx < 0 or row_idx >= self._ts_row_count:
            return

        if not self.temp_source_connected or self.temp_source_manager is None:
            QMessageBox.warning(self, "提示", "请先点击「通信」连接温度源后再发送")
            return

        sp = self.row_setpoint_spins[row_idx].value()
        main_val = self.row_main_spins[row_idx].value()
        sec_val = self.row_sec_spins[row_idx].value()
        if sp is None:
            QMessageBox.warning(self, "提示", f"行{row_idx+1} SP 未填写，请填写后再发送")
            return
        is_const1210 = (self._ts_device_type == 'Const 1210')

        if is_const1210:
            self.status_label.setText(f"行{row_idx+1}: SP={sp}°C - 发送温度...")
            QApplication.processEvents()
            ok = self._send_and_verify_cmd(
                set_cmd=f"SOUR:USER:SPO {sp}\r\n",
                query_cmd="SOUR:USER:SPO?\r\n",
                expected_value=sp,
                label=f"行{row_idx+1} Const 1210"
            )
            if ok:
                self._update_ts_btn(True)
                self.status_label.setText(f"行{row_idx+1} 命令已发送 (SP={sp}°C)")
                msg = f"行{row_idx+1} SP命令发送成功: SP={sp}°C"
                print(f"[发送] {msg}")
                if hasattr(self, 'legend_widget'):
                    self.legend_widget.append_terminal(msg)
                # 手动点击行发送按钮成功后弹框提示
                QMessageBox.information(self, "成功", f"行{row_idx+1} SP 设置成功: {sp}°C")
            else:
                self.status_label.setText(f"行{row_idx+1} 命令发送失败")
                print(f"[发送] 行{row_idx+1} SP命令发送失败")
                if hasattr(self, 'legend_widget'):
                    self.legend_widget.append_terminal(f"行{row_idx+1} SP命令发送失败")
            return

        # ---- Fluke 9250 ----
        self.status_label.setText(f"行{row_idx+1}: 发送 PID 参数...")
        QApplication.processEvents()
        if not self._send_pid_params(row_idx):
            self.status_label.setText(f"行{row_idx+1} PID 参数发送失败")
            print(f"[发送] 行{row_idx+1} PID 参数发送失败")
            if hasattr(self, 'legend_widget'):
                self.legend_widget.append_terminal(f"行{row_idx+1} PID 参数发送失败")
            return

        self.status_label.setText(f"行{row_idx+1}: 设置温度={sp}°C, Main={main_val}, Sec={sec_val}...")
        QApplication.processEvents()
        # Main 为默认值（留空）时不发送
        if main_val is not None:
            ok = self._send_and_verify_cmd(
                set_cmd=f"SOUR:MAIN:SPO {main_val}\r\n",
                query_cmd="SOUR:MAIN:SPO?\r\n",
                expected_value=main_val,
                label=f"行{row_idx+1} Main"
            )
            if not ok:
                self.status_label.setText(f"行{row_idx+1} Main 发送失败")
                print(f"[发送] 行{row_idx+1} Main 发送失败")
                if hasattr(self, 'legend_widget'):
                    self.legend_widget.append_terminal(f"行{row_idx+1} Main 发送失败")
                return

        # Sec 为默认值（留空）时不发送
        if sec_val is not None:
            ok = self._send_and_verify_cmd(
                set_cmd=f"SOUR:SEC:SPO {sec_val}\r\n",
                query_cmd="SOUR:SEC:SPO?\r\n",
                expected_value=sec_val,
                label=f"行{row_idx+1} Sec"
            )
            if not ok:
                self.status_label.setText(f"行{row_idx+1} Sec 发送失败")
                print(f"[发送] 行{row_idx+1} Sec 发送失败")
                if hasattr(self, 'legend_widget'):
                    self.legend_widget.append_terminal(f"行{row_idx+1} Sec 发送失败")
                return
        # SP 发送按钮不关联启动：不发送 OUTP:STAT 1（启动由「开启」按钮控制）
        self._update_temp_source_ui_state()
        self.status_label.setText(f"行{row_idx+1} 命令已发送 (SP={sp}°C, M={main_val}, S={sec_val})")
        msg = f"行{row_idx+1} SP命令发送成功: SP={sp}°C"
        print(f"[发送] {msg}")
        if hasattr(self, 'legend_widget'):
            self.legend_widget.append_terminal(msg)
        # 手动点击行发送按钮成功后弹框提示
        QMessageBox.information(self, "成功", f"行{row_idx+1} SP 设置成功: {sp}°C")

    def temp_source_start(self):
        """启动温度源输出：先发送并验证 PID 参数，再发送 OUTP:STAT 1
           Const 1210: 跳过 PID，直接使用预留的控制命令"""
        is_const1210 = (self._ts_device_type == 'Const 1210')

        # Const 1210 不需要串口库
        if not is_const1210 and not SERIAL_AVAILABLE:
            QMessageBox.warning(self, "错误", "串口库不可用")
            return

        # 打开/获取连接
        if self.temp_source_manager is None:
            if not self._open_temp_source_serial():
                return

        if is_const1210:
            # Const 1210: 发送 SOUR:USER:SPO <sp> 设定温度，问询 SOUR:USER:SPO?
            # 始终用第一行作为 SP 来源，不受勾选状态影响
            pid_row = 0
            sp = self.row_setpoint_spins[pid_row].value()

            self.status_label.setText(f"Const 1210: 设定温度={sp}°C...")
            QApplication.processEvents()
            ok = self._send_and_verify_cmd(
                set_cmd=f"SOUR:USER:SPO {sp}\r\n",
                query_cmd="SOUR:USER:SPO?\r\n",
                expected_value=sp,
                label="Const 1210启动"
            )
            self._update_temp_source_ui_state()
            self.seq_btn.setEnabled(True)
            self.loop_btn.setEnabled(True)
            self.reverse_btn.setEnabled(True)
            if ok:
                self.status_label.setText("Const 1210 温度源已启动")
            else:
                self.status_label.setText("Const 1210 温度源启动失败（回读状态异常）")
            return

        # ---- Fluke 9250 原有逻辑 ----
        # 1) 发送并验证第一行的 PID 参数（不受勾选状态影响）
        pid_row = 0
        self.status_label.setText(f"设置行{pid_row+1} PID 参数...")
        QApplication.processEvents()
        if not self._send_pid_params(pid_row):
            self.status_label.setText("PID 参数设置失败，启动已取消")
            return

        # 2) 发送 OUTP:STAT 1 并验证
        self.status_label.setText("发送启动命令...")
        QApplication.processEvents()
        ok = self._send_and_verify_cmd(
            set_cmd=b"OUTP:STAT 1\r\n",
            query_cmd=b"OUTP:STAT?\r\n",
            expected_value=1,
            label="启动温度源"
        )
        # 更新按钮状态
        self._update_temp_source_ui_state()
        # 顺序测试结束后，重新启用顺序测试按钮
        self.seq_btn.setEnabled(True)
        self.loop_btn.setEnabled(True)
        self.reverse_btn.setEnabled(True)
        if ok:
            self.status_label.setText("温度源已启动")
        else:
            self.status_label.setText("温度源启动失败（回读状态异常）")

    def temp_source_stop(self):
        """仅停止温度源输出，不影响采集和顺序测试"""
        if self.temp_source_manager is not None:
            is_const1210 = (self._ts_device_type == 'Const 1210')
            if is_const1210:
                # Const 1210: TEMPerature:STATus:MEASure 停止控制
                self._send_temp_source_cmd(b"TEMPerature:STATus:MEASure\r\n")
            else:
                self._send_temp_source_cmd(b"OUTP:STAT 0\r\n")
            self._update_ts_btn(False)
            self.status_label.setText("温度源已停止")
        else:
            self.status_label.setText("温度源连接未打开")

    def _stop_sequential_test(self):
        """停止顺序测试：停止温度源 → 停止采集 → 保存数据 → 保存截图（以setpoint命名）"""
        # 先获取当前行的 setpoint，用于截图文件名
        current_sp = None
        if self.sequential_running:
            row = self.sequential_current_row
            if row < self._ts_row_count:
                current_sp = self.row_setpoint_spins[row].value()
            # 提前设置好文件名用的温度值，再停止运行标志，确保截图以setpoint命名
            if current_sp is not None and current_sp > 0:
                self.current_test_temp = current_sp

        # 先停定时器和标志，防止 _sequential_tick 在后续操作中触发 _send_and_verify_cmd 弹窗
        self.sequential_timer.stop()
        self.sequential_running = False

        # 停止温度源（定时器已停，不会再有弹窗干扰）
        self.temp_source_stop()

        # 停止采集并保存（stop_collection 内部已包含截图）
        if self.test_running:
            self.has_unsaved_data = True
            self.auto_save_data()
            self.has_unsaved_data = False
            self.stop_collection()

        # 顺序测试结束，关闭 setpoint 命名标志（必须在 stop_collection 之后）
        self._seq_naming_setpoint = False

        self._update_sched_btn()
        self.sched_time_edit.setEnabled(True)
        self._update_ts_btn(False)
        self._update_seq_btn()
        self._update_loop_btn()
        self._update_reverse_btn()
        self.status_label.setText("测试已手动停止")

    def _start_sequential_test(self):
        """开始顺序测试：逐行设置温度、采集、停止、等待间隔后继续下一行"""
        is_const1210 = (self._ts_device_type == 'Const 1210')
        if not is_const1210 and not SERIAL_AVAILABLE:
            QMessageBox.warning(self, "错误", "串口库不可用")
            return

        # 打开/获取串口
        if self.temp_source_manager is None:
            if not self._open_temp_source_serial():
                return

        # 检查是否有勾选的行
        valid_rows = [r for r in range(self._ts_row_count) if self.row_checks[r].isChecked()]
        if not valid_rows:
            QMessageBox.warning(self, "警告", "请先勾选需要测试的行")
            return

        # 根据模式构建行执行顺序
        # forward顺序：0,1,2,...   reverse倒序：n-1,n-2,...,0
        # loop循环：0,1,2,...,n-1,n-2,...,1（往返一遍）
        if self.sequential_mode == 'reverse':
            self.seq_row_order = list(reversed(valid_rows))
        elif self.sequential_mode == 'loop':
            self.seq_row_order = valid_rows + list(reversed(valid_rows[:-1]))
        else:
            self.seq_row_order = list(valid_rows)

        self.sequential_running = True
        self.sequential_current_row = self.seq_row_order[0]
        self.seq_step = 0
        self.sequential_test_complete = False
        # 顺序测试文件名采用 setpoint 温度（非实时值）
        self._seq_naming_setpoint = True

        # 切换按钮状态
        self._update_ts_btn(True)
        self._update_seq_btn()
        self._update_loop_btn()
        self._update_reverse_btn()
        self._update_sched_btn()
        self.sched_time_edit.setEnabled(False)

        # 启动顺序测试定时器（每秒检查一次）
        self.sequential_timer.start(1000)

        mode_name = {"forward": "顺序测试", "reverse": "倒序测试", "loop": "循环测试"}.get(self.sequential_mode, "测试")
        self.status_label.setText(f"{mode_name}启动，共 {len(self.seq_row_order)} 个设定点")
        if hasattr(self, 'legend_widget'):
            self.legend_widget.append_terminal(f"{mode_name}启动，共 {len(self.seq_row_order)} 个设定点")
        # 开始第一行
        self._start_current_row()

    def _arm_scheduled_test(self):
        """设定定时测试：到达指定日期时间后自动打开串口、启动温度源、开始顺序测试"""
        target_dt = self.sched_time_edit.dateTime()
        now = QDateTime.currentDateTime()
        if target_dt <= now:
            QMessageBox.warning(self, "警告", "定时时间必须晚于当前时间")
            return
        self.sched_target_time = target_dt
        self.sched_test_armed = True
        self._update_sched_btn()
        self.sched_time_edit.setEnabled(False)
        remain_secs = int(now.secsTo(target_dt))
        if remain_secs >= 3600:
            remain_str = f"{remain_secs//3600}h{(remain_secs%3600)//60}min"
        else:
            remain_str = f"{remain_secs//60}min"
        self.sched_status_label.setText(f"已设定 {target_dt.toString('MM-dd HH:mm')} ({remain_str})")
        self.sched_test_timer.start(1000)  # 每秒检查

    def _cancel_scheduled_test(self):
        """取消已设定的定时测试"""
        self.sched_test_armed = False
        self.sched_test_timer.stop()
        self._update_sched_btn()
        self.sched_time_edit.setEnabled(True)
        self.sched_status_label.setText("")

    def _scheduled_test_check(self):
        """每秒检查是否到达定时测试时间"""
        if not self.sched_test_armed:
            return
        now = QDateTime.currentDateTime()
        target = self.sched_target_time
        remain_secs = int(now.secsTo(target))
        if remain_secs <= 0:
            self.sched_test_armed = False
            self.sched_test_timer.stop()
            self.sched_status_label.setText("定时测试执行中...")
            # 定时测试使用顺序模式
            self.sequential_mode = 'forward'
            # _start_sequential_test 内部会自动处理串口连接和温度源启动
            self._start_sequential_test()
        elif remain_secs < 60:
            self.sched_status_label.setText(f"{remain_secs}s")
        elif remain_secs < 3600:
            self.sched_status_label.setText(f"{remain_secs//60}min{remain_secs%60}s")

    def _start_current_row(self):
        """启动当前行的测试（先确保温度源关闭再开始）"""
        # 安全：启动新行前先确保温度源是关闭状态
        self.temp_source_stop()

        # 确保温度源串口连接正常，若断开则自动重连
        if self.temp_source_manager is None or not self.temp_source_manager.connected:
            self.status_label.setText("温度源串口已断开，尝试重新连接...")
            QApplication.processEvents()
            if not self._open_temp_source_serial():
                self.status_label.setText("温度源串口重连失败，停止测试")
                self.sequential_running = False
                self._update_seq_btn()
                self._update_loop_btn()
                self._update_reverse_btn()
                return

        row = self.sequential_current_row
        if row >= self._ts_row_count:
            # 所有行都已完成
            self.status_label.setText("测试全部完成")
            self.temp_source_stop()
            self.sequential_running = False
            self._update_seq_btn()
            self._update_loop_btn()
            self._update_reverse_btn()
            return

        # 跳过未勾选的行
        if not self.row_checks[row].isChecked():
            if self._advance_seq_step():
                self._start_current_row()
            return

        sp = self.row_setpoint_spins[row].value()
        main_val = self.row_main_spins[row].value()
        sec_val = self.row_sec_spins[row].value()
        if sp <= 0:
            # 该行无有效设置，跳过
            if self._advance_seq_step():
                self._start_current_row()
            return

        # 设置当前测试温度和 Spec
        self.current_test_temp = sp
        self.current_test_spec = self.row_setpoint_spec[row].value()

        is_const1210 = (self._ts_device_type == 'Const 1210')

        if is_const1210:
            # Const 1210: 发送 SOUR:USER:SPO <sp> 设定温度，问询 SOUR:USER:SPO?
            self.status_label.setText(f"行{row+1}: SP={sp}°C - Const 1210 设置温度...")
            QApplication.processEvents()
            ok = self._send_and_verify_cmd(
                set_cmd=f"SOUR:USER:SPO {sp}\r\n",
                query_cmd="SOUR:USER:SPO?\r\n",
                expected_value=sp,
                label=f"行{row+1} Const 1210"
            )
            if not ok:
                self.status_label.setText(f"行{row+1} Const 1210 设置失败，跳过该行")
                self.sequential_current_row += 1
                self._start_current_row()
                return

            # 更新按钮状态
            self._update_ts_btn(True)

            self.sequential_start_time = time.time()
            self.sequential_test_complete = False

            # 启动数据采集（传入当前温度）
            self.start_collection(temp=sp)
            return

        # ---- Fluke 9250 原有逻辑 ----
        self.status_label.setText(f"行{row+1}: 设置 PID 参数...")
        QApplication.processEvents()
        if not self._send_pid_params(row):
            self.status_label.setText(f"行{row+1} PID 参数设置失败，跳过该行")
            if self._advance_seq_step():
                self._start_current_row()
            return

        self.status_label.setText(f"行{row+1}: 设置温度={sp}°C, Main={main_val}, Sec={sec_val}...")

        # 1) 发送 SP 设置并回读验证（Fluke 9250 用 SOUR:USER:SPO，与手动测试一致）
        ok = self._send_and_verify_cmd(
            set_cmd=f"SOUR:USER:SPO {sp}\r\n",
            query_cmd="SOUR:USER:SPO?\r\n",
            expected_value=sp,
            label=f"行{row+1} SP"
        )
        if not ok:
            self.status_label.setText(f"行{row+1} Main设置失败，跳过该行")
            if self._advance_seq_step():
                self._start_current_row()
            return

        # 2) 发送 sec 设置并回读验证（Sec 为默认值留空时不发送）
        if sec_val is not None:
            ok = self._send_and_verify_cmd(
                set_cmd=f"SOUR:SEC:SPO {sec_val}\r\n",
                query_cmd="SOUR:SEC:SPO?\r\n",
                expected_value=sec_val,
                label=f"行{row+1} Sec"
            )
            if not ok:
                self.status_label.setText(f"行{row+1} Sec设置失败，跳过该行")
                if self._advance_seq_step():
                    self._start_current_row()
                return

        # 3) 发送 OUTP:STAT 1 并回读验证
        ok = self._send_and_verify_cmd(
            set_cmd=b"OUTP:STAT 1\r\n",
            query_cmd=b"OUTP:STAT?\r\n",
            expected_value=1,
            label=f"行{row+1} 启动"
        )
        if not ok:
            self.status_label.setText(f"行{row+1} 启动失败，跳过该行")
            if self._advance_seq_step():
                self._start_current_row()
            return

        # 更新按钮状态
        self._update_temp_source_ui_state()

        self.sequential_start_time = time.time()
        self.sequential_test_complete = False
        self.status_label.setText(f"行{row+1}: 温度={sp}°C, Main={main_val}, Sec={sec_val} - 开始采集")

        # 启动数据采集（传入当前温度）
        self.start_collection(temp=sp)

    def _sequential_tick(self):
        """顺序执行定时器回调：检查当前行是否该结束"""
        if not self.sequential_running or self.sequential_start_time is None:
            return
        row = self.sequential_current_row
        if row >= self._ts_row_count:
            return

        elapsed = (time.time() - self.sequential_start_time) / 60.0  # 分钟

        # 检查所有 auto_test 通道是否都已完成 T3
        # 过滤掉无数据卡在 idle 的通道（避免通讯不通时阻塞顺序测试）
        all_auto_chs = list(self.auto_test_state.keys())
        auto_chs = [ch for ch in all_auto_chs
                    if self.auto_test_state[ch]['phase'] != 'idle'
                    or (ch in self.data_buffer and len(self.data_buffer[ch]) >= 2)]
        # 如果全是无数据的 idle 通道（全被过滤了），按无 auto_test 处理
        if all_auto_chs and not auto_chs:
            # 有 auto_test 通道但都没有数据 → 等待10分钟(采集启动) + 60分钟超时
            target_duration = 10 if elapsed < 10 else 60
        elif auto_chs:
            all_complete = all(
                self.auto_test_state[ch]['phase'] == 'complete' and self.auto_test_state[ch]['T3'] is not None
                for ch in auto_chs
            )
            if all_complete:
                self.sequential_test_complete = True
                # 取最晚完成的 T3 作为目标时长
                t3_duration = max(self.auto_test_state[ch]['T3'] for ch in auto_chs)
                target_duration = t3_duration
            else:
                # 有 auto_test 通道但未全部完成，继续等待（24h兜底）
                target_duration = 1440
        else:
            # 无 auto_test 通道，用60min超时避免卡死
            target_duration = 60

        if target_duration <= 0:
            return

        # 达到目标时长 → 停止当前行（即使 test_running=False 也会结束行）
        if elapsed >= target_duration:
            # 防重入：避免保存/切换过程中定时器再次触发导致跳过下一行
            if getattr(self, '_row_transitioning', False):
                return
            self._row_transitioning = True
            self.status_label.setText(f"行{row+1}采集完成({elapsed:.1f}min)，正在保存数据...")
            try:
                # 保险：在保存前主动触发一次最后的自动检测
                self._check_special_channels()

                # 0) 标记正在顺序存储，阻止 _on_save_timer 并发写文件
                self._sequential_saving = True
                # 等待可能的后台保存线程完成
                bg_thread = getattr(self, '_bg_save_thread', None)
                if bg_thread and bg_thread.is_alive():
                    bg_thread.join(timeout=10)

                # 1) 先保存数据（含avg1/avg2记录）
                self.has_unsaved_data = True
                self.auto_save_data()
                self.has_unsaved_data = False
                self._sequential_saving = False

                # 2) 停止采集线程
                self.test_running = False
                self.save_timer.stop()
                if self.auto_save_timer is not None:
                    self.auto_save_timer.stop()
                    self.auto_save_timer = None
                for t in self.device_threads:
                    t.stop()

                # 3) 停止温度源输出（先尝试回读验证，失败则直接发命令）
                is_const1210 = (self._ts_device_type == 'Const 1210')
                if is_const1210:
                    # Const 1210: 跳过 OUTP:STAT 命令（命令待定）
                    self.temp_source_stop()
                else:
                    self.status_label.setText(f"行{row+1} 发送停止命令...")
                    QApplication.processEvents()
                    ok = self._send_and_verify_cmd(
                        set_cmd=b"OUTP:STAT 0\r\n",
                        query_cmd=b"OUTP:STAT?\r\n",
                        expected_value=0,
                        label=f"行{row+1} 停止"
                    )
                    if not ok:
                        # 验证失败时直接发送停止命令确保安全
                        self._send_temp_source_cmd(b"OUTP:STAT 0\r\n")
                    self._update_temp_source_ui_state()

                # 4) 清理UI状态
                self._update_collect_btn()
                self.save_btn.setEnabled(False)
                for i in range(self._dev_row_count):
                    self.device_widgets[i]['status'].setText("未连接")
                    unit = self.device_quantity_info.get(i, {}).get('unit', '')
                    self.legend_widget.update_temperature(i, None, unit)
                    self.legend_widget.update_volatility(i, None, unit=unit)
                    self.legend_widget.update_stats(i, unit=unit)
                self._reset_auto_test_state()
                self.legend_widget.clear_auto_test()
                self.status_label.setText(f"行{row+1}采集完成，温度源已停止")
                # 行采集完成：截取整窗截图（命名为 setpoint-时间，保存到 test data）
                self._save_whole_window_screenshot('stop')
            except Exception as e:
                # 任何异常都要确保温度源关闭，但不中断顺序测试
                import traceback
                print(f"行{row+1}停止时发生异常: {e}")
                traceback.print_exc()
                try:
                    if self._ts_device_type != 'Const 1210':
                        self._send_temp_source_cmd(b"OUTP:STAT 0\r\n")
                except:
                    pass
                try:
                    self.test_running = False
                    self.save_timer.stop()
                    if self.auto_save_timer is not None:
                        self.auto_save_timer.stop()
                        self.auto_save_timer = None
                    for t in self.device_threads:
                        t.stop()
                except:
                    pass
                # 不设置 sequential_running=False，继续执行下一行

            # 5) 等待间隔后进入下一行（异常时也继续）
            try:
                interval_min = self.test_interval_spin.value()
            except:
                interval_min = 0
            if interval_min > 0 and row < self._ts_row_count - 1:
                self.status_label.setText(f"行{row+1}完成(有异常)，等待{interval_min}min后执行下一行...")
                QTimer.singleShot(int(interval_min * 60 * 1000), self._proceed_next_row)
            else:
                self._proceed_next_row()

    def _advance_seq_step(self):
        """按模式序列推进到下一行；返回 True 表示仍有下一行，False 表示全部完成"""
        if not self.sequential_running:
            return False
        self.seq_step += 1
        if self.seq_step < len(self.seq_row_order):
            self.sequential_current_row = self.seq_row_order[self.seq_step]
            return True
        return False

    def _proceed_next_row(self):
        """进入下一行测试"""
        if not self.sequential_running:
            self._row_transitioning = False
            return
        if self._advance_seq_step():
            self._row_transitioning = False
            self._start_current_row()
        else:
            self._row_transitioning = False
            mode_name = {"forward": "顺序测试", "reverse": "倒序测试", "loop": "循环测试"}.get(self.sequential_mode, "测试")
            self.status_label.setText(f"{mode_name}全部完成")
            if hasattr(self, 'legend_widget'):
                self.legend_widget.append_terminal(f"{mode_name}全部完成")
            self.temp_source_stop()
            self.sequential_running = False
            self._update_seq_btn()
            self._update_loop_btn()
            self._update_reverse_btn()
            self._update_sched_btn()
            self.sched_time_edit.setEnabled(True)

    def get_filename_temp_str(self):
        """文件名用的温度字符串（统一使用 SP 设定值，保留一位小数）。

        - 优先使用 current_test_temp（SP 设定值）。
        - 若无 SP（current_test_temp 无效），回退到第一个启用且有数据的通道实时值。
        """
        # SP 设定值：current_test_temp
        try:
            if self.current_test_temp is not None:
                return f"{float(self.current_test_temp):.1f}"
        except (TypeError, ValueError):
            pass
        # 回退：第一个启用且有数据的通道实时值
        first_temp = None
        for i in range(self._dev_row_count):
            if not self.devices[i]['enabled']:
                continue
            buf = self.data_buffer.get(i) if hasattr(self, 'data_buffer') else None
            if buf:
                first_temp = buf[-1]
                break
        if first_temp is None:
            return 'NA'
        try:
            return f"{float(first_temp):.1f}"
        except (TypeError, ValueError):
            return 'NA'

    def start_collection(self, temp=None, mode_tag=None):
        enabled = [d for d in self.devices if d['enabled']]
        if not enabled:
            QMessageBox.warning(self, "警告", "请至少启用一个设备")
            return

        if temp is not None:
            self.current_test_temp = temp
        elif hasattr(self, 'ts_manual_sp_spin'):
            # 未指定 SP 时，用温度源控制的手动 SP 值作为文件名温度
            self.current_test_temp = self.ts_manual_sp_spin.value()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # 文件名温度：统一使用 SP 设定值（current_test_temp）
        temp_str = self.get_filename_temp_str()
        if temp_str == 'NA':
            try:
                temp_str = f"{float(self.current_test_temp):.1f}"
            except (TypeError, ValueError):
                temp_str = 'NA'
        # 区分自动测试(auto)与手动测试(manual)：顺序/循环/倒序/定时测试为自动测试
        if mode_tag is None:
            mode_tag = 'auto' if getattr(self, 'sequential_running', False) else 'manual'
        self._current_mode_tag = mode_tag
        base_name = f"{temp_str}-{mode_tag}-{timestamp}.xlsx"
        # 保存到 test data 文件夹（与记录数据同一目录）
        script_dir = os.path.dirname(os.path.abspath(__file__))
        save_dir = os.path.join(script_dir, "test data")
        os.makedirs(save_dir, exist_ok=True)
        self.current_data_file = os.path.join(save_dir, base_name)
        # 重置手动测试完成标记（开始测试时）
        self._manual_test_done = False
        # 一开始采集就立即生成 Excel 文件（含空 sheet 结构）
        try:
            from openpyxl import Workbook
            book = Workbook()
            if 'Sheet' in book.sheetnames:
                del book['Sheet']
            book.create_sheet('实时数据')
            book.create_sheet('stability')
            if mode_tag == 'axis':
                book.create_sheet('axis')
            elif mode_tag == 'radial':
                book.create_sheet('radial')
            book.save(self.current_data_file)
            print(f"[start_collection] 已创建 Excel: {self.current_data_file}")
        except Exception as e:
            print(f"[start_collection] 创建 Excel 失败: {e}")
        # 标记采集文件尚未用“第一个通道实时值”重命名（首个数据点到达后再重命名）
        self._data_file_renamed = False
        self.data_buffer = {i: deque(maxlen=self.max_points) for i in range(self._dev_row_count)}
        self.time_buffer = {i: deque(maxlen=self.max_points) for i in range(self._dev_row_count)}
        self.datetime_buffer = {i: deque(maxlen=self.max_points) for i in range(self._dev_row_count)}
        self.start_time = time.time()
        self.test_running = True
        self.has_unsaved_data = False
        self._plot_dirty = True
        self._auto_follow = True
        self._suppress_range_signal = False
        self._reset_auto_test_state()

        # 输出采集开始程序信息到终端区
        if hasattr(self, 'legend_widget'):
            self.legend_widget.append_terminal(f"开始采集，数据文件:{os.path.basename(self.current_data_file)}")

        # 设置多Y轴绘图（根据启用的设备命令自动判断物理量）
        self.setup_multi_axis_plot()

        self.device_threads = []
        interval = self.interval_spin.value() / 1000.0

        for i, d in enumerate(self.devices):
            if d['enabled']:
                d['read_interval'] = interval
                t = DeviceThread(i, d)
                t.data_received.connect(self.on_data_received)
                t.connection_status.connect(self.on_connection_status)
                t.debug_info.connect(self.on_debug_info)
                t.start()
                self.device_threads.append(t)

        # 启动每45分钟自动保存定时器
        if self.auto_save_timer is not None:
            self.auto_save_timer.stop()
        self.auto_save_timer = QTimer()
        self.auto_save_timer.timeout.connect(self.periodic_auto_save)
        self.auto_save_timer.start(45 * 60 * 1000)  # 45分钟

        self._update_collect_btn()
        self.save_btn.setEnabled(False)
        self.status_label.setText(f"正在采集数据... 保存文件: {self.current_data_file}")
        self._save_tick_count = 0
        self.save_timer.start(5000)

    def _save_whole_window_screenshot(self, suffix):
        """截取整个软件窗口并保存到 test data 文件夹。
        文件名：{温度}-{auto/manual}-{suffix}-{时间}.png
        温度策略：统一使用 SP 设定值，并根据测试模式加入 auto/manual 标识。
        """
        try:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            temp_str = self.get_filename_temp_str()
            mode_tag = getattr(self, '_current_mode_tag', None)
            if mode_tag is None:
                mode_tag = 'auto' if getattr(self, 'sequential_running', False) else 'manual'
            script_dir = os.path.dirname(os.path.abspath(__file__))
            save_dir = os.path.join(script_dir, "test data")
            os.makedirs(save_dir, exist_ok=True)
            if mode_tag == 'axis':
                # 轴向测试截图文件名：sp-axis-时间.png
                img_path = os.path.join(save_dir, f"{temp_str}-axis-{ts}.png")
            elif mode_tag == 'radial':
                # 径向测试截图文件名：sp-radial-时间.png
                img_path = os.path.join(save_dir, f"{temp_str}-radial-{ts}.png")
            else:
                img_path = os.path.join(save_dir, f"{temp_str}-{mode_tag}-{suffix}-{ts}.png")
            self._save_screenshot_with_overlay(img_path, ts)
            print(f"[screenshot] 已保存整窗截图: {img_path}")
        except Exception as e:
            print(f"[screenshot] 截图失败: {e}")

    def _save_screenshot_with_overlay(self, filepath, timestamp, jpg=False):
        """截取窗口并保存，在图像顶部绘制程序版本号和当前系统时间（半透明黑底白字）"""
        try:
            pixmap = self.grab()
            painter = QPainter(pixmap)
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            text = f"T-cal_tester v{APP_VERSION}    {now_str}"
            font = QFont(font_family, 16)
            font.setBold(True)
            painter.setFont(font)
            fm = QFontMetrics(font)
            tw = fm.horizontalAdvance(text)
            th = fm.height()
            # 半透明黑底条
            painter.fillRect(0, 0, tw + 20, th + 10, QColor(0, 0, 0, 140))
            painter.setPen(QColor(255, 255, 255))
            painter.drawText(10, th + 2, text)
            painter.end()
            if jpg:
                pixmap.save(filepath, 'JPG', 95)
            else:
                pixmap.save(filepath, 'PNG')
        except Exception as e:
            print(f"[截图] 保存失败: {e}")

    def stop_collection(self):
        self.test_running = False
        self.save_timer.stop()
        if hasattr(self, 'legend_widget'):
            self.legend_widget.append_terminal("停止采集")
        # 停止45分钟自动保存定时器
        if self.auto_save_timer is not None:
            self.auto_save_timer.stop()
            self.auto_save_timer = None
        for t in self.device_threads:
            t.stop()
        # 强制关闭所有 SharedSerialManager 实例，中断阻塞在 send_command 上的线程
        for mgr in list(SharedSerialManager._instances.values()):
            mgr.disconnect()
        if self.has_unsaved_data:
            self.auto_save_data()
        else:
            # 兜底：即使标志未设置，但缓冲区有数据也保存
            has_data = any(len(self.data_buffer[i]) > 0 for i in range(self._dev_row_count))
            if has_data:
                self.auto_save_data()
        # 停止采集后截取整个软件窗口（文件名温度策略：顺序测试用 setpoint，否则第一个通道实时值）
        self._save_whole_window_screenshot('stop')
        # 如果温度源控制仍在使用，需立即重连共享串口
        if self.sequential_running:
            port, baudrate = self._find_tc_main_device()
            if port:
                self.temp_source_manager = SharedSerialManager(port, baudrate)
                if not self.temp_source_manager.connected:
                    self.temp_source_manager.connect()
        self._update_collect_btn()
        self.save_btn.setEnabled(False)
        for i in range(self._dev_row_count):
            self.device_widgets[i]['status'].setText("未连接")
            unit = self.device_quantity_info.get(i, {}).get('unit', '')
            self.legend_widget.update_temperature(i, None, unit)
            self.legend_widget.update_volatility(i, None, unit=unit)
            self.legend_widget.update_stats(i, unit=unit)
        self._reset_auto_test_state()
        self.legend_widget.clear_auto_test()
        self.status_label.setText("采集已停止，数据已自动保存")

    def auto_save_data(self):
        """保存当前所有数据到Excel（主线程调用，等待后台保存完成以防竞态）"""
        # 等待可能的后台保存线程完成，避免旧快照覆盖新数据
        bg_thread = getattr(self, '_bg_save_thread', None)
        if bg_thread and bg_thread.is_alive():
            bg_thread.join(timeout=10)
        try:
            has_data = any(len(self.data_buffer[i]) > 0 for i in range(self._dev_row_count))
            if not has_data:
                self.status_label.setText("无数据可保存")
                return

            max_len = max((len(self.datetime_buffer[i]) for i in range(self._dev_row_count) if self.devices[i]['enabled']), default=0)
            df = pd.DataFrame()
            times = None
            for i in range(self._dev_row_count):
                if self.devices[i]['enabled'] and len(self.datetime_buffer[i]) > 0:
                    times = [x.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3] for x in self.datetime_buffer[i]]
                    break
            if times:
                if len(times) < max_len:
                    times += [''] * (max_len - len(times))
                else:
                    times = times[:max_len]
                df["采集时间"] = times
            for i in range(self._dev_row_count):
                if self.devices[i]['enabled']:
                    name = self.devices[i]['name']
                    unit = self.device_quantity_info.get(i, {}).get('unit', '°C')
                    if unit:
                        col_name = f"{name} ({unit})"
                    else:
                        col_name = name
                    data = list(self.data_buffer[i])[:max_len]
                    while len(data) < max_len:
                        data.append(None)
                    df[col_name] = data
            # 写入实时数据 sheet
            import os
            full_path = os.path.abspath(self.current_data_file)
            with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='实时数据', index=False)
                if self.auto_test_summary:
                    pd.DataFrame([self.auto_test_summary]).to_excel(writer, sheet_name='stability', index=False)
            # axis/radial 模式且已有轴向结果时，补写对应 sheet（ExcelWriter 重写会清掉它）
            if getattr(self, '_current_mode_tag', '') in ('axis', 'radial') and getattr(self, 'axial_data', None):
                self._save_axis_sheet()
            self.status_label.setText(f"数据已自动保存至: {full_path}")
            print(f"[auto_save_data] Excel已保存: {full_path}")
        except Exception as e:
            import traceback
            print(f"[auto_save_data] 保存失败: {e}")
            traceback.print_exc()
            self.status_label.setText(f"自动保存失败: {str(e)}")
            QMessageBox.critical(self, "错误", f"自动保存失败: {str(e)}")

    def _save_loop_range(self):
        """返回当前设备数的 range，供各保存方法使用"""
        return range(self._dev_row_count)

    def periodic_auto_save(self):
        """每45分钟自动保存数据到硬盘"""
        if self.test_running and self.has_unsaved_data:
            self.auto_save_data()
            self.status_label.setText(f"定时自动保存完成: {self.current_data_file}  (每45分钟保存一次)")

    def emergency_save_data(self):
        """程序意外关闭时的应急保存（atexit回调）"""
        if hasattr(self, 'test_running') and self.test_running and self.has_unsaved_data:
            try:
                has_data = any(len(self.data_buffer[i]) > 0 for i in range(self._dev_row_count))
                if not has_data:
                    return

                max_len = max((len(self.datetime_buffer[i]) for i in range(self._dev_row_count) if self.devices[i]['enabled']), default=0)
                df = pd.DataFrame()
                times = None
                for i in range(self._dev_row_count):
                    if self.devices[i]['enabled'] and len(self.datetime_buffer[i]) > 0:
                        times = [x.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3] for x in self.datetime_buffer[i]]
                        break
                if times:
                    if len(times) < max_len:
                        times += [''] * (max_len - len(times))
                    else:
                        times = times[:max_len]
                    df["采集时间"] = times
                for i in range(self._dev_row_count):
                    if self.devices[i]['enabled']:
                        name = self.devices[i]['name']
                        unit = self.device_quantity_info.get(i, {}).get('unit', '°C')
                        if unit:
                            col_name = f"{name} ({unit})"
                        else:
                            col_name = name
                        data = list(self.data_buffer[i])[:max_len]
                        while len(data) < max_len:
                            data.append(None)
                        df[col_name] = data
                import os
                full_path = os.path.abspath(self.current_data_file)
                with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='实时数据', index=False)
                    if self.auto_test_summary:
                        pd.DataFrame([self.auto_test_summary]).to_excel(writer, sheet_name='stability', index=False)
                print(f"[应急保存] 数据已保存至: {full_path}")
            except Exception as e:
                import traceback
                print(f"[应急保存] 保存失败: {e}")
                traceback.print_exc()

    def _log_startup_info(self):
        """程序启动后向终端区输出程序信息"""
        try:
            py_ver = f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}"
            self.legend_widget.clear_terminal()
            self.legend_widget.append_terminal(f"T-cal_tester v{APP_VERSION} 启动完成")
            self.legend_widget.append_terminal(f"Python {py_ver} | 系统 {platform.platform()}")
            self.legend_widget.append_terminal(f"配置文件: {self.config_file} | 设备数: {len(self.devices)}")
        except Exception as e:
            print(f"[启动信息] 输出失败: {e}")

    def _save_window_geometry(self):
        """将窗口位置和大小保存到 JSON 文件（直接存 x/y/w/h 整数）"""
        geom_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "window_geometry.json")
        try:
            data = {}
            if os.path.exists(geom_file):
                with open(geom_file, 'r') as f:
                    data = json.load(f)
            g = self.geometry()
            data['main'] = {'x': g.x(), 'y': g.y(), 'w': g.width(), 'h': g.height()}
            # 保存温度源控制组框的折叠状态
            if hasattr(self, 'temp_ctrl_group'):
                data['temp_ctrl'] = {'collapsed': self.temp_ctrl_group.is_collapsed()}
            with open(geom_file, 'w') as f:
                json.dump(data, f, indent=2)
        except Exception:
            pass

    def showEvent(self, event):
        """窗口首次显示时恢复上次位置和大小（延迟执行，确保布局完成）"""
        super().showEvent(event)
        if not getattr(self, '_geometry_restored', False):
            QTimer.singleShot(50, self._restore_window_geometry)
            self._geometry_restored = True

    def _restore_window_geometry(self):
        """实际执行窗口几何信息恢复（在 showEvent 后延迟调用）"""
        geom_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "window_geometry.json")
        try:
            if os.path.exists(geom_file):
                with open(geom_file, 'r') as f:
                    data = json.load(f)
                # 恢复主窗口位置大小
                if 'main' in data and isinstance(data['main'], dict):
                    g = data['main']
                    x, y, w, h = g.get('x', 0), g.get('y', 0), g.get('w', 0), g.get('h', 0)
                    if w > 0 and h > 0:
                        self.move(x, y)
                        self.resize(w, h)
                # 恢复温度源控制组框的折叠状态（不设固定高度，避免间距异常）
                if hasattr(self, 'temp_ctrl_group') and 'temp_ctrl' in data:
                    tc = data['temp_ctrl']
                    if isinstance(tc, dict):
                        collapsed = tc.get('collapsed', False)
                        self.temp_ctrl_group.set_collapsed(collapsed)
                # 恢复主分割器左右比例（左侧/右侧）
                if hasattr(self, 'main_splitter') and 'main_split' in data:
                    sizes = data['main_split']
                    if isinstance(sizes, list) and len(sizes) == 2 and all(s > 0 for s in sizes):
                        self.main_splitter.setSizes(sizes)
            else:
                # 首次启动：把当前打开的窗口位置作为默认大小写入
                self._save_window_geometry()
        except Exception:
            pass

    def resizeEvent(self, event):
        """窗口大小变化时保存几何信息"""
        super().resizeEvent(event)
        self._save_window_geometry()

    def moveEvent(self, event):
        """窗口位置变化时保存几何信息（拖动即固化为默认）"""
        super().moveEvent(event)
        self._save_window_geometry()

    def _on_main_splitter_moved(self, pos, index):
        """用户手动拖动主分割器时，保存左右比例到 JSON"""
        geom_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "window_geometry.json")
        try:
            data = {}
            if os.path.exists(geom_file):
                with open(geom_file, 'r') as f:
                    data = json.load(f)
            data['main_split'] = self.main_splitter.sizes()
            with open(geom_file, 'w') as f:
                json.dump(data, f, indent=2)
        except Exception:
            pass

    def _signal_handler(self, signum, frame):
        """信号处理（SIGINT/SIGTERM）：先停温度源 → 应急保存 → 退出"""
        print(f"\n收到退出信号({signum})，正在停止温度源并保存数据...")
        try:
            if self._ts_device_type != 'Const 1210':
                self._send_temp_source_cmd(b"OUTP:STAT 0\r\n")
        except:
            pass
        self.emergency_save_data()
        self._save_window_geometry()
        sys.exit(0)

    def closeEvent(self, event):
        """窗口关闭时：先停温度源 → 保存数据 → 关串口（断开两台设备）"""
        # 停止两台设备的温度输出，防止意外关闭后继续加热
        for dev_type in ('Fluke 9250', 'Const 1210'):
            mgr = self._ts_device_managers.get(dev_type)
            if mgr is not None:
                try:
                    if dev_type != 'Const 1210':
                        mgr.send_command("OUTP:STAT 0\r\n", timeout=1.0)
                    else:
                        mgr.send_command("TEMPerature:STATus:MEASure\r\n", timeout=1.0)
                except:
                    pass
        self.save_config()
        if self.sequential_running:
            self.sequential_timer.stop()
        if self.test_running:
            self.stop_collection()
        # 断开两台设备
        for dev_type in ('Fluke 9250', 'Const 1210'):
            mgr = self._ts_device_managers.get(dev_type)
            if mgr is not None:
                try:
                    mgr.disconnect()
                except:
                    pass
            self._ts_device_managers[dev_type] = None
            self._ts_device_connected[dev_type] = False
        # 确保温度查询线程退出
        if self.ts_query_thread is not None:
            self.ts_query_thread.stop()
            self.ts_query_thread = None
        if self.ts_const1210_query_thread is not None:
            self.ts_const1210_query_thread.stop()
            self.ts_const1210_query_thread = None
        # 保存窗口位置与大小，下次启动恢复
        self._save_window_geometry()
        event.accept()

    def manual_save_data(self):
        # 文件名温度统一使用 SP 值（get_filename_temp_str），手动测试加 manual 标识
        temp_label = self.get_filename_temp_str()
        path, _ = QFileDialog.getSaveFileName(self, "保存数据", f"{temp_label}-manual-{datetime.now():%Y%m%d_%H%M%S}.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        try:
            has_data = any(len(self.data_buffer[i]) > 0 for i in range(self._dev_row_count))
            if not has_data:
                QMessageBox.warning(self, "警告", "无数据可保存")
                return
            max_len = max((len(self.datetime_buffer[i]) for i in range(self._dev_row_count) if self.devices[i]['enabled']), default=0)
            df = pd.DataFrame()
            times = None
            for i in range(self._dev_row_count):
                if self.devices[i]['enabled'] and len(self.datetime_buffer[i]) > 0:
                    times = [x.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3] for x in self.datetime_buffer[i]]
                    break
            if times:
                if len(times) < max_len:
                    times += [''] * (max_len - len(times))
                else:
                    times = times[:max_len]
                df["采集时间"] = times
            for i in range(self._dev_row_count):
                if self.devices[i]['enabled']:
                    name = self.devices[i]['name']
                    unit = self.device_quantity_info.get(i, {}).get('unit', '°C')
                    if unit:
                        col_name = f"{name} ({unit})"
                    else:
                        col_name = name
                    data = list(self.data_buffer[i])[:max_len]
                    while len(data) < max_len:
                        data.append(None)
                    df[col_name] = data
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='实时数据', index=False)
                if self.auto_test_summary:
                    pd.DataFrame([self.auto_test_summary]).to_excel(writer, sheet_name='stability', index=False)
            QMessageBox.information(self, "成功", f"数据已保存至：{path}")
            self.status_label.setText(f"数据已手动保存至: {path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存失败：{str(e)}")

    def record_current_data(self):
        """记录当前实时数据窗口中各启用通道的实时值/波动/Min/Max/Avg 到独立 Excel 文件，并截图"""
        try:
            from openpyxl import Workbook
            if not self.test_running or not self.current_data_file:
                QMessageBox.warning(self, "警告", "请先开始采集后再记录数据")
                return
            now = datetime.now()
            ts = now.strftime("%Y%m%d_%H%M%S")
            # 文件名温度统一策略：第一个启用通道实时值，保留一位小数
            temp_str = self.get_filename_temp_str()

            # 统一保存到脚本所在目录下的 “test data” 文件夹
            script_dir = os.path.dirname(os.path.abspath(__file__))
            dir_path = os.path.join(script_dir, "test data")
            os.makedirs(dir_path, exist_ok=True)
            record_base = f"{temp_str}-manual-{ts}"
            path = os.path.join(dir_path, f"{record_base}.xlsx")
            img_path = os.path.join(dir_path, f"{record_base}.png")

            # 收集各启用通道当前实时数据
            rows = []
            for i in range(self._dev_row_count):
                if not self.devices[i]['enabled']:
                    continue
                name = self.devices[i].get('name', f'设备{i+1}')
                unit = self.device_quantity_info.get(i, {}).get('unit', '°C')
                buf = self.data_buffer[i]
                cur = buf[-1] if buf else None
                if cur is None:
                    volatility = min_val = max_val = avg_val = None
                else:
                    volatility, _, min_val, max_val, avg_val = self.calculate_volatility(i)
                rows.append([
                    i + 1,
                    name,
                    round(cur, 4) if cur is not None else None,
                    round(volatility, 4) if volatility is not None else None,
                    round(min_val, 4) if min_val is not None else None,
                    round(max_val, 4) if max_val is not None else None,
                    round(avg_val, 4) if avg_val is not None else None,
                    now.strftime('%Y-%m-%d %H:%M:%S'),
                ])
            if not rows:
                QMessageBox.warning(self, "警告", "没有启用的通道或尚无数据")
                return

            # 截图整个软件主窗口（文件：温度-record-时间.png，顶部叠加版本号+系统时间）
            img_ok = False
            try:
                self._save_screenshot_with_overlay(img_path, ts)
                img_ok = os.path.exists(img_path)
            except Exception as e:
                img_ok = False
                print(f"[record_current_data] 截图失败: {e}")

            # 写入独立的 Excel 文件（新文件，不依赖采集 Excel）
            book = Workbook()
            if 'Sheet' in book.sheetnames:
                del book['Sheet']
            ws = book.create_sheet('实时数据记录')
            headers = ['通道', '设备名称', '实时值(°C)', '波动', 'Min', 'Max', 'Avg', '记录时间']
            ws.append(headers)
            for r in rows:
                ws.append(r)
            book.save(path)

            msg = f"记录已保存至：{path}"
            if img_ok:
                msg += f"\n截图已保存：{img_path}"
            else:
                msg += "\n（截图未生成）"
            QMessageBox.information(self, "成功", msg)
            self.status_label.setText(f"记录数据已保存: {path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"记录失败：{str(e)}")

    def reset_stats_all(self):
        """重置统计：将所有启用通道的波动/Min/Max/Avg 统计清零（不影响曲线显示与数据记录）"""
        for i in range(self._dev_row_count):
            if not self.devices[i]['enabled']:
                continue
            self.legend_widget.reset_volatility(i)
            unit = self.device_quantity_info.get(i, {}).get('unit', '')
            self.legend_widget.update_stats(i, unit=unit)
        self.status_label.setText("统计（波动/Min/Max/Avg）已重置")

    def reset_curve_display(self):
        """清空图像：重新绘图（重建 plot），保留数据缓冲区，随后从缓冲区重绘曲线"""
        # 保存现有数据缓冲区，重建图像结构
        try:
            self.init_plots()
        except Exception as e:
            print(f"[清空图像] 重建绘图失败: {e}")
        # 恢复自动跟随视图
        self._auto_follow = True
        # 强制下一帧从已有缓冲区重新绘制曲线
        self._plot_dirty = True
        self.status_label.setText("图像已清空并重新绘图")

    def _reset_plot_view(self):
        """重置图形缩放视图，恢复自动缩放显示全部数据"""
        if not hasattr(self, 'primary_plot') or self.primary_plot is None:
            return
        # 恢复自动跟随：曲线会持续绘制并始终显示全部数据
        self._auto_follow = True
        self._suppress_range_signal = True
        self.primary_plot.vb.enableAutoRange(axis=pg.ViewBox.YAxis, enable=True)
        self.primary_plot.vb.enableAutoRange(axis=pg.ViewBox.XAxis, enable=False)
        # 重置额外Y轴视图
        for vb in getattr(self, 'extra_vb_list', []):
            vb.enableAutoRange(axis=pg.ViewBox.YAxis, enable=True)
        self._suppress_range_signal = False
        self.status_label.setText("图形视图已重置，自动跟随最新数据")

    def _on_view_range_changed(self, *args):
        """用户手动缩放/平移时关闭自动跟随，避免被程序自动范围覆盖"""
        if getattr(self, '_suppress_range_signal', False):
            return
        if self.primary_plot is not None and not self.primary_plot.vb.autoRangeEnabled():
            self._auto_follow = False

    def save_current_plot(self, auto_save=False):
        """保存当前图像曲线为JPG文件"""
        try:
            # 文件名温度统一使用 SP 值（get_filename_temp_str），如 100.0-时间
            temp_label = self.get_filename_temp_str()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{temp_label}-manual-{timestamp}.jpg"
            # 保存到 test data 文件夹
            script_dir = os.path.dirname(os.path.abspath(__file__))
            save_dir = os.path.join(script_dir, "test data")
            os.makedirs(save_dir, exist_ok=True)
            filepath = os.path.join(save_dir, filename)
            # 截取整个主窗口画面，并在顶部叠加版本号+系统时间
            self._save_screenshot_with_overlay(filepath, timestamp, jpg=True)
            if auto_save:
                self.status_label.setText(f"曲线已自动保存: {filename}")
            else:
                QMessageBox.information(self, "成功", f"截图已保存至：{filepath}")
                self.status_label.setText(f"曲线已保存: {filename}")
        except Exception as e:
            self.status_label.setText(f"保存曲线失败: {str(e)}")

    # ========== 轴向测试 ==========
    def _axial_on_height_edited(self, row, col):
        """key 单元格（第0行、第1列及之后）编辑后处理。
        轴向模式：自动补 mm 单位（输入 9 -> 显示 9.0mm）
        径向模式：保留用户输入的字符串位置标识。"""
        if row != 0 or col == 0:
            return
        try:
            item = self.axial_table.item(row, col)
            if item is None:
                return
            text = str(item.text()).strip()
            if text == '':
                return
            # 找到该列
            target = None
            for c in self.axial_columns:
                if c['col'] == col:
                    target = c
                    break
            if target is None:
                return
            mode = target.get('mode', self.axial_mode)
            if mode == 'axial':
                # 轴向模式：自动补 mm
                cleaned = text.replace('mm', '').strip()
                try:
                    val = float(cleaned)
                except (TypeError, ValueError):
                    return
                new_key = round(val, 1)
                new_text = f"{new_key:.1f}mm"
                if item.text() != new_text:
                    self.axial_table.blockSignals(True)
                    item.setText(new_text)
                    self.axial_table.blockSignals(False)
            else:
                # 径向模式：直接用用户输入作为 key
                new_key = text
                new_text = text
            # 同步缓存的 key
            old_key = target['key']
            if new_key != old_key:
                if old_key in self.axial_keys:
                    self.axial_keys.remove(old_key)
                    self.axial_data.pop(old_key, None)
                target['key'] = new_key
                if new_key not in self.axial_keys:
                    self.axial_keys.append(new_key)
        except Exception as e:
            print(f"[axial edit] 处理失败: {e}")

    def _axial_add_height_column(self, value=None, add_save_btn=True):
        """"+"按钮：新增一列。轴向模式=高度(float mm)，径向模式=位置字符串（如 a-c）。
        value 为 None 时自动取下一值：
          - 轴向模式：最后一列高度 + 10mm
          - 径向模式：按默认列表循环 [a-c, b-c, d-c, c-d, e-c, ...]
        add_save_btn=False 时不创建该列的保存按钮（用于第1列）。"""
        try:
            mode = self.axial_mode
            # 自动生成 key
            if value is None:
                if mode == 'radial':
                    # 径向模式：从默认列表依次取，超出后拼接 e-c, f-c, ...
                    default_keys = ['a-c', 'b-c', 'd-c', 'c-d']
                    if len(self.axial_keys) < len(default_keys):
                        value = default_keys[len(self.axial_keys)]
                    else:
                        # 超过4个则用字母递增，首字母用 chr(ord('a')+idx)
                        idx = len(self.axial_keys) - len(default_keys) + 4
                        first_char = chr(ord('a') + idx)
                        value = f"{first_char}-c"
                else:
                    # 轴向模式：最后一列 + 10mm
                    if self.axial_keys:
                        try:
                            value = round(float(max(self.axial_keys)) + 10.0, 1)
                        except (TypeError, ValueError):
                            value = 0.0
                    else:
                        value = 0.0
            # 标准化 key（轴向转 float，径向保留 str）
            if mode == 'axial':
                if isinstance(value, bool):
                    value = None
                if value is None:
                    value = 0.0
                key = round(float(value), 1)
                display = f"{key:.1f}mm"
            else:
                key = str(value).strip()
                display = key
            # 重复检查
            if key in self.axial_keys:
                QMessageBox.information(
                    self, "提示",
                    f"该列已存在: {display}\n现有列: {self.axial_keys}")
                return
            col = len(self.axial_columns) + 1
            self.axial_table.setColumnCount(col + 1)
            # 该列第0行：key 显示（可编辑）
            key_item = QTableWidgetItem(display)
            key_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.axial_table.setItem(0, col, key_item)
            self.axial_table.setItem(1, col, QTableWidgetItem("--"))
            self.axial_table.setItem(2, col, QTableWidgetItem("--"))
            save_btn = None
            header_widget = None
            if add_save_btn:
                save_btn = QPushButton("保存")
                save_btn.setFixedHeight(22)
                save_btn.setStyleSheet(
                    "QPushButton{background:#E91E63;color:white;font-weight:bold;border-radius:3px;"
                    "padding:1px 4px;font-size:10px;}"
                    "QPushButton:hover{background:#F06292;}")
                # 使用 lambda 捕获当前 col 引用
                save_btn.clicked.connect(lambda checked=False, c=col: self._axial_save_height(c))
                header_widget = save_btn
            else:
                header_widget = QLabel("")
                header_widget.setFixedHeight(22)
            self.axial_save_btns_layout.addWidget(header_widget, stretch=1)
            self.axial_columns.append({
                'key': key,
                'mode': mode,
                'save_btn': save_btn,
                'col': col,
                'header_widget': header_widget,
            })
            self.axial_keys.append(key)
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "错误", f"添加列失败：{str(e)}")

    def _axial_clear_all_columns(self):
        """切换模式时清空所有列和缓存数据"""
        # 移除所有列上的保存按钮 / 占位
        for c in list(self.axial_columns):
            hw = c.get('header_widget')
            if hw is not None:
                self.axial_save_btns_layout.removeWidget(hw)
                hw.deleteLater()
        self.axial_columns.clear()
        self.axial_keys.clear()
        self.axial_data.clear()
        self.axial_records.clear()
        # 表格重置为 1 列（仅行标签列）
        self.axial_table.blockSignals(True)
        self.axial_table.setColumnCount(1)
        # 行0保留空白 item
        if self.axial_table.item(0, 0) is None:
            blank = QTableWidgetItem("")
            blank.setFlags(blank.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.axial_table.setItem(0, 0, blank)
        self.axial_table.blockSignals(False)

    def _axial_on_mode_changed(self, idx):
        """QComboBox 切换轴向/径向模式：清空现有数据并按新模式重新填充默认列"""
        new_mode = 'radial' if idx == 1 else 'axial'
        if new_mode == self.axial_mode:
            return
        self.axial_mode = new_mode
        # 更新列0 标题
        if hasattr(self, 'axial_col0_header'):
            self.axial_col0_header.setText("位置" if new_mode == 'radial' else "高度")
        # 同步开始按钮文字（未在采集中时）
        if hasattr(self, 'axial_start_btn') and not self.test_running:
            mode_cn = "径向" if new_mode == 'radial' else "轴向"
            self.axial_start_btn.setText(f"开始{mode_cn}测试")
        # 清空旧列与缓存
        self._axial_clear_all_columns()
        # 按新模式填充默认列
        if new_mode == 'radial':
            for key in ['a-c', 'b-c', 'd-c', 'c-d']:
                self._axial_add_height_column(key, add_save_btn=True)
        else:
            for h in [0.0, 10.0, 20.0, 30.0]:
                self._axial_add_height_column(h, add_save_btn=True)
        # 状态文字
        try:
            if hasattr(self, 'status_label'):
                self.status_label.setText(
                    f"已切换到{'径向' if new_mode == 'radial' else '轴向'}测试模式")
        except Exception:
            pass

    def _axial_remove_height_column(self):
        """"-"按钮：删除最后一列"""
        if not self.axial_columns:
            return
        col_info = self.axial_columns.pop()
        hw = col_info.get('header_widget')
        if hw is not None:
            self.axial_save_btns_layout.removeWidget(hw)
            hw.deleteLater()
        key = col_info.get('key')
        if key in self.axial_keys:
            self.axial_keys.remove(key)
            self.axial_data.pop(key, None)
        col_count = self.axial_table.columnCount()
        if col_count > 1:
            self.axial_table.setColumnCount(col_count - 1)

    def _axial_save_height(self, col):
        """点击某列的"保存"按钮：将当前通道1/2的avg写入该列单元格。
        轴向模式 key 为 float（高度/mm），径向模式 key 为 str（位置对）。"""
        try:
            target = None
            for c in self.axial_columns:
                if c['col'] == col:
                    target = c
                    break
            if target is None:
                return
            mode = target.get('mode', self.axial_mode)
            # 从该列第0行单元格读取 key（用户可编辑）
            key_item = self.axial_table.item(0, col)
            cell_text = str(key_item.text()).strip() if key_item else ''
            if mode == 'axial':
                try:
                    key = round(float(cell_text.replace('mm', '').strip()), 1)
                    display = f"{key:.1f}mm"
                except (TypeError, ValueError):
                    key = target['key']
                    display = f"{key:.1f}mm"
            else:
                key = cell_text
                display = key
                if not key:
                    key = target['key']
                    display = key
            # 更新缓存的 key
            if key != target['key']:
                if target['key'] in self.axial_keys:
                    self.axial_keys.remove(target['key'])
                    self.axial_data.pop(target['key'], None)
                target['key'] = key
                self.axial_keys.append(key)
                # 规范化单元格显示
                self.axial_table.blockSignals(True)
                self.axial_table.setItem(0, col, QTableWidgetItem(display))
                self.axial_table.blockSignals(False)
            f_avg = self._get_channel_avg(0)
            m_avg = self._get_channel_avg(1)
            if f_avg is None and m_avg is None:
                QMessageBox.warning(self, "警告", "通道1/通道2 暂无数据，无法保存")
                return
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.axial_table.setItem(1, col, QTableWidgetItem("--" if f_avg is None else f"{f_avg:.4f}"))
            self.axial_table.setItem(2, col, QTableWidgetItem("--" if m_avg is None else f"{m_avg:.4f}"))
            self.axial_data[key] = {'F': f_avg, 'M': m_avg, 'time': now_str}
            self.axial_records.append((key, f_avg, m_avg, now_str))
            self._save_axis_sheet()
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "错误", f"保存失败：{str(e)}")

    def _get_channel_avg(self, dev_id):
        """获取指定通道的窗口平均温度（avg），无数据返回 None"""
        if not (0 <= dev_id < self._dev_row_count):
            return None
        if not self.devices[dev_id]['enabled']:
            return None
        try:
            result = self.calculate_volatility(dev_id)
            if result and len(result) >= 5:
                return result[4]  # avg
        except Exception:
            pass
        return None

    def _axial_save_excel(self):
        """将轴向/径向测试记录保存为Excel（行=指标，列=各列 key）"""
        if not self.axial_data:
            QMessageBox.warning(self, "警告", "暂无轴向/径向测试记录可保存")
            return
        try:
            import os
            import pandas as pd
            script_dir = os.path.dirname(os.path.abspath(__file__))
            save_dir = os.path.join(script_dir, "test data")
            os.makedirs(save_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            mode_tag = 'radial' if self.axial_mode == 'radial' else 'axial'
            filename = f"{mode_tag}-{timestamp}.xlsx"
            filepath = os.path.join(save_dir, filename)
            data = {'指标': ['F-avg', 'M-avg']}
            for c in self.axial_columns:
                mode = c.get('mode', self.axial_mode)
                # 从单元格读取最新 key
                item0 = self.axial_table.item(0, c['col'])
                cell_text = str(item0.text()).strip() if item0 else ''
                if mode == 'axial':
                    try:
                        key = round(float(cell_text.replace('mm', '').strip()), 1)
                    except (TypeError, ValueError):
                        key = c['key']
                    col_label = f"{key:.1f}mm"
                else:
                    key = cell_text or c['key']
                    col_label = str(key)
                rec = self.axial_data.get(c['key'], {})
                data[col_label] = [
                    None if rec.get('F') is None else round(rec['F'], 4),
                    None if rec.get('M') is None else round(rec['M'], 4),
                ]
            df = pd.DataFrame(data)
            sheet_name = 'radial' if self.axial_mode == 'radial' else 'axis'
            df.to_excel(filepath, sheet_name=sheet_name, index=False)
            QMessageBox.information(self, "成功", f"测试数据已保存至：{filepath}")
            self.status_label.setText(f"测试数据已保存: {filepath}")
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "错误", f"保存失败：{str(e)}")

    def _save_axis_sheet(self):
        """将轴向/径向测试结果写入当前采集文件的 'axis' sheet。
        表结构：行为指标(F-avg / M-avg)，列为各 key。
        轴向模式：key 为 float 高度，列名带 mm
        径向模式：key 为 str 位置对，列名为位置字符串。"""
        if not self.axial_data:
            return
        if not self.current_data_file or not os.path.exists(self.current_data_file):
            QMessageBox.warning(self, "警告", "当前无采集文件，无法写入 axis sheet")
            return
        sheet_name = 'radial' if self.axial_mode == 'radial' else 'axis'
        try:
            from openpyxl import load_workbook
            data = {'指标': ['F-avg', 'M-avg']}
            for c in self.axial_columns:
                mode = c.get('mode', self.axial_mode)
                item0 = self.axial_table.item(0, c['col'])
                cell_text = str(item0.text()).strip() if item0 else ''
                if mode == 'axial':
                    try:
                        key = round(float(cell_text.replace('mm', '').strip()), 1)
                    except (TypeError, ValueError):
                        key = c['key']
                    col_label = f"{key:.1f}mm"
                else:
                    key = cell_text or c['key']
                    col_label = str(key)
                rec = self.axial_data.get(c['key'], {})
                data[col_label] = [
                    None if rec.get('F') is None else round(rec['F'], 4),
                    None if rec.get('M') is None else round(rec['M'], 4),
                ]
            df = pd.DataFrame(data)
            full_path = os.path.abspath(self.current_data_file)
            # sheet 名：轴向=axis，径向=radial
            sheet_name = 'radial' if self.axial_mode == 'radial' else 'axis'
            wb = load_workbook(full_path)
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws = wb.create_sheet(sheet_name)
            ws.append(list(df.columns))
            for _, row in df.iterrows():
                ws.append([None if pd.isna(v) else v for v in row.tolist()])
            wb.save(full_path)
            mode_cn = '径向' if self.axial_mode == 'radial' else '轴向'
            self.status_label.setText(f"{mode_cn}测试结果已写入 {sheet_name} sheet: {full_path}")
            print(f"[{sheet_name}] 已写入 sheet '{sheet_name}' -> {full_path}")
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "错误", f"{sheet_name} sheet 写入失败：{str(e)}")

    def on_connection_status(self, dev_id, ok, msg):
        s = self.device_widgets[dev_id]['status']
        if ok:
            s.setText("已连接")
            s.setStyleSheet("color:green;font-weight:bold;")
        else:
            s.setText(msg)
            s.setStyleSheet("color:red;font-weight:bold;")

    def _rename_collection_file_if_needed(self, temp):
        """首个启用通道收到第一个数据点时，用 SP 设定值（而非实时值）重命名采集 Excel 文件。"""
        if getattr(self, '_data_file_renamed', True):
            return
        if not self.current_data_file or not os.path.exists(self.current_data_file):
            return
        # 文件名温度统一使用 SP 设定值（current_test_temp），不用实时 main 温度
        try:
            temp_str = f"{float(self.current_test_temp):.1f}"
        except (TypeError, ValueError):
            try:
                temp_str = f"{float(temp):.1f}"
            except (TypeError, ValueError):
                return
        # 仅当文件名仍为回退默认温度时才重命名（避免覆盖真实命名）
        old_path = self.current_data_file
        dir_path = os.path.dirname(old_path)
        ts = os.path.basename(old_path).split('-', 1)[1] if '-' in os.path.basename(old_path) else datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
        new_path = os.path.join(dir_path, f"{temp_str}-{ts}")
        if new_path == old_path:
            self._data_file_renamed = True
            return
        try:
            os.replace(old_path, new_path)
            self.current_data_file = new_path
            self._data_file_renamed = True
            print(f"[rename] 采集文件已重命名为: {new_path}")
        except Exception as e:
            print(f"[rename] 重命名失败: {e}")

    def on_data_received(self, data, dev_id):
        if not self.test_running:
            return
        t = time.time() - self.start_time
        dt = datetime.now()
        temp = data.get('value')
        if temp is not None:
            self.data_buffer[dev_id].append(temp)
            self.time_buffer[dev_id].append(t)
            self.datetime_buffer[dev_id].append(dt)
            self._plot_dirty = True

            # 文件名统一使用 SP 设定值（start_collection 已按 SP 命名），不再用实时值重命名
            first_enabled = next((i for i in range(self._dev_row_count) if self.devices[i]['enabled']), None)
            if (first_enabled is not None and dev_id == first_enabled
                    and not getattr(self, '_data_file_renamed', True)):
                self._rename_collection_file_if_needed(temp)

            unit = self.device_quantity_info.get(dev_id, {}).get('unit', '')
            self.legend_widget.update_temperature(dev_id, temp, unit)

            result = self.calculate_volatility(dev_id)
            if result:
                volatility, duration_min, min_val, max_val, avg_val = result
                self.legend_widget.update_volatility(dev_id, volatility, duration_min, unit)
                self.legend_widget.update_stats(dev_id, min_val, max_val, avg_val, unit)

            if not self.has_unsaved_data:
                self.has_unsaved_data = True
                self.save_btn.setEnabled(True)
        else:
            unit = self.device_quantity_info.get(dev_id, {}).get('unit', '')
            self.legend_widget.update_temperature(dev_id, None, unit)
            # 不清除波动和统计量——保持上次有效值

    def calculate_volatility(self, dev_id):
        if not self.devices[dev_id]['enabled']:
            return None
        # 增量缓存：缓存键为 (缓冲长度, 最后一点时间戳)。
        # 未满窗口时长度变化会触发重算；满窗口时长度恒定但末尾时间戳变化也会触发重算，
        # 从而保证结果始终正确，同时避免同一帧内重复全量计算。
        cache = getattr(self, '_vol_cache', None)
        if cache is None:
            cache = self._vol_cache = {}
        dts = self.datetime_buffer[dev_id]
        n = len(dts)
        if n == 0:
            cache[dev_id] = (0, None)
            return None
        key = (n, int(np.asarray(dts, dtype='datetime64[us]')[-1].astype('int64')))
        c = cache.get(dev_id)
        if c is not None and c[0] == key:
            return c[1]

        now = datetime.now()
        # 所有通道波动计算窗口最大 30min，滚动更新
        # Stability 通道：窗口起点取重置时刻与 now-30min 的较大者
        # 非 Stability 通道：固定取最近 30min 滚动窗口
        is_stability = bool(self.devices[dev_id].get('auto_test', False))
        if is_stability:
            reset_time = self.legend_widget.get_reset_time(dev_id)
            if reset_time:
                start_time = max(reset_time, now - timedelta(minutes=30))
            else:
                start_time = now - timedelta(minutes=30)
        else:
            start_time = now - timedelta(minutes=30)  # 非 Stability 也限 30min 滚动窗口

        # 用 numpy 从末尾切片取窗口内数据，避免 Python 级逐点循环
        temps_arr = np.asarray(self.data_buffer[dev_id], dtype=float)
        dts_arr = np.asarray(dts, dtype='datetime64[us]')
        if start_time is None:
            mask = np.ones(len(dts_arr), dtype=bool)
        else:
            start_ns = np.datetime64(start_time)
            mask = dts_arr >= start_ns
        if not mask.any():
            result = None
        else:
            temps = temps_arr[mask]
            earliest_dt = dts_arr[mask][0]
            if len(temps) < 2:
                result = None
            else:
                volatility = float(np.std(temps))
                min_val = float(np.min(temps))
                max_val = float(np.max(temps))
                avg_val = float(np.mean(temps))
                if earliest_dt:
                    data_duration_min = (now - earliest_dt.astype('O')).total_seconds() / 60.0
                else:
                    data_duration_min = 0
                result = (volatility, data_duration_min, min_val, max_val, avg_val)
        cache[dev_id] = (key, result)
        return result

    def _filter_outliers_for_display(self, values):
        """将统计离群点替换为 np.nan，避免单个跳变点撑大曲线 Y 轴比例。
        原始数据仍保留在 data_buffer 中，不影响 Excel 保存和统计计算。
        使用 IQR 方法（1.5 倍 IQR 为常见阈值，这里放宽到 3 倍以减少误伤）。
        输入为 numpy 数组，返回 numpy 数组（pyqtgraph 可直接绘制）。
        """
        n = len(values)
        if n < 10:
            return values
        arr = np.asarray(values, dtype=float)
        q1, q3 = np.percentile(arr, [25, 75])
        iqr = q3 - q1
        if iqr <= 0:
            return arr
        lower = q1 - 3.0 * iqr
        upper = q3 + 3.0 * iqr
        # 用 nan 代替异常点，pyqtgraph 绘制时会跳过，Y 轴自动范围也不会被撑大
        return np.where((arr >= lower) & (arr <= upper), arr, np.nan)

    def update_plots(self):
        if not self.test_running:
            return
        self._check_special_channels()
        # 手动测试模式：检查 stability 通道 T3 是否全部完成
        # 轴向/径向测试（axis/radial 模式）不参与 T3 结束条件，始终记录，直到手动停止
        # 手动测试下若未勾选"判断T3时刻"则连续记录，不判断 T3 结束条件
        if (self.test_running and not self.sequential_running
                and getattr(self, '_current_mode_tag', None) not in ('axis', 'radial')
                and getattr(self, 'manual_check_t3_cb', None)
                and self.manual_check_t3_cb.isChecked()
                and not getattr(self, '_manual_test_done', False)):
            self._check_manual_test_completion()
        # 仅在有新数据时更新曲线
        if getattr(self, '_plot_dirty', False):
            self._plot_dirty = False
            for i in range(self._dev_row_count):
                if self.curves[i] is not None:
                    try:
                        if (self.devices[i]['enabled'] and self.devices[i].get('curve_visible', True)
                                and len(self.time_buffer[i]) > 0 and len(self.data_buffer[i]) > 0):
                            # 用 numpy 向量化，避免每帧 Python 级列表推导
                            times_min = np.asarray(self.time_buffer[i], dtype=float) / 60.0
                            # 过滤异常显示点，原始数据仍保留在 buffer 中
                            plot_vals = self._filter_outliers_for_display(np.asarray(self.data_buffer[i], dtype=float))
                            self.curves[i].setData(times_min, plot_vals)
                            self.curves[i].setVisible(True)
                        else:
                            self.curves[i].setVisible(False)
                    except RuntimeError:
                        self.curves[i] = None
            # 自动跟随：保持 X 轴显示全部数据，曲线持续绘制；手动缩放时不覆盖用户视图
            if getattr(self, '_auto_follow', True) and self.primary_plot is not None:
                self.primary_plot.vb.enableAutoRange(axis=pg.ViewBox.YAxis, enable=True)
                self.primary_plot.vb.enableAutoRange(axis=pg.ViewBox.XAxis, enable=False)
                tmin = tmax = None
                for i in range(self._dev_row_count):
                    if self.devices[i]['enabled'] and self.time_buffer[i]:
                        tb = self.time_buffer[i]
                        lo, hi = tb[0] / 60.0, tb[-1] / 60.0
                        if tmin is None or lo < tmin:
                            tmin = lo
                        if tmax is None or hi > tmax:
                            tmax = hi
                if tmin is not None and tmax is not None:
                    if tmax <= tmin:
                        tmax = tmin + 0.01
                    self._suppress_range_signal = True
                    self.primary_plot.vb.setXRange(tmin, tmax, padding=0.02)
                    self._suppress_range_signal = False
        self._save_tick_count += 1

    def _on_save_timer(self):
        """定时器回调：每5秒一次，仅在有新数据时后台写Excel（每60秒一次）"""
        if not self.test_running:
            return
        if not hasattr(self, '_save_tick_count'):
            self._save_tick_count = 0
        # 如果顺序测试正在结束当前行，跳过后台保存避免竞态
        if getattr(self, '_sequential_saving', False):
            return
        if self._save_tick_count % 12 == 0 and self.has_unsaved_data:
            # 等待上次后台保存完成（最多等5秒）
            bg_thread = getattr(self, '_bg_save_thread', None)
            if bg_thread and bg_thread.is_alive():
                bg_thread.join(timeout=5)
            # 始终包含所有已启用设备（即使某个通道暂时无数据也要保留列）
            enabled_ids = [i for i in range(self._dev_row_count) if self.devices[i]['enabled']]
            if not enabled_ids:
                return
            # 计算所有通道的最大数据长度作为统一行数
            all_max = max((len(self.data_buffer[i]) for i in enabled_ids), default=0)
            buf_copy = {}
            for i in enabled_ids:
                data = list(self.data_buffer[i])
                dt_list = list(self.datetime_buffer[i])
                # 补齐到统一长度（不足部分填None在后面）
                while len(data) < all_max:
                    data.append(None)
                while len(dt_list) < all_max:
                    dt_list.append(None)
                buf_copy[i] = (data, dt_list, self.devices[i]['name'])
            if buf_copy:
                filepath = self.current_data_file
                auto_log = list(self.auto_test_log)
                self._bg_save_thread = threading.Thread(target=self._save_excel_bg,
                                     args=(filepath, buf_copy, auto_log, all_max),
                                     daemon=True)
                self._bg_save_thread.start()
                self.status_label.setText(f"数据已后台保存: {filepath}")

    def _save_excel_bg(self, filepath, buf_copy, auto_log, row_count):
        """后台线程保存Excel，不阻塞UI"""
        try:
            if row_count <= 0:
                return
            df = pd.DataFrame()
            # 构建统一的时间列：取第一个有值的设备的时间
            times = None
            for dev_id, (data_list, dt_list, name) in buf_copy.items():
                if dt_list and dt_list[0] is not None:
                    times = [x.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3] if x else '' for x in dt_list]
                    break
            if times:
                while len(times) < row_count:
                    times.append('')
                df["采集时间"] = times[:row_count]
            # 写入所有设备的列（包括非stability通道）
            for dev_id, (data_list, dt_list, name) in buf_copy.items():
                data = data_list[:row_count]
                df[name] = data
            import os
            full_path = os.path.abspath(filepath)
            with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='实时数据', index=False)
                if getattr(self, 'auto_test_summary', None):
                    pd.DataFrame([self.auto_test_summary]).to_excel(writer, sheet_name='stability', index=False)
        except Exception as e:
            import traceback
            print(f"[_save_excel_bg] 后台保存Excel失败: {e}")
            traceback.print_exc()

    def _on_row_setpoint_changed(self, row_idx, value):
        """温度源设置行 setpoint 变更"""
        if not self._loading:
            self.save_config()

    def _on_row_spec_changed(self, row_idx, value):
        """温度源设置行 Spc（Spec）变更"""
        if not self._loading:
            self.save_config()

    def _on_row_main_changed(self, row_idx, value):
        """温度源设置行 main 变更"""
        if not self._loading:
            self.save_config()

    def _on_row_sec_changed(self, row_idx, value):
        """温度源设置行 sec 变更"""
        if not self._loading:
            self.save_config()

    def _reset_auto_test_state(self):
        """重置所有勾选了 auto_test 通道的自动检测状态"""
        blank_state = {'T0': None, 'T1': None, 'T2': None, 'T3': None,
                       'std1': None, 'std2': None, 'avg1': None, 'avg2': None,
                       'phase': 'idle', 'T0_time': None, 'T2_time': None}
        self.auto_test_state.clear()
        self.auto_test_logged.clear()
        self.auto_test_summary = None
        for d in range(self._dev_row_count):
            if self.devices[d].get('auto_test', False):
                self.auto_test_state[d] = dict(blank_state)
                self.auto_test_logged[d] = set()
        self.auto_test_log.clear()
        self.auto_test_lines.clear()

    def _build_device_rows(self):
        """重建所有设备行（可拖动排序）"""
        # 清除旧行
        for container in self._dev_row_containers:
            container.setParent(None)
            container.deleteLater()
        self._dev_row_containers.clear()
        self.device_widgets.clear()

        dev_layout = self._dev_rows_layout
        # 清除旧行控件
        while dev_layout.count():
            item = dev_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        for dev_id in range(self._dev_row_count):
            container = DragDropRowWidget(dev_id)
            container.reorder_requested.connect(self._on_device_row_reorder)
            row_w = self._create_device_row_widgets(dev_id, container.contentWidget())
            dev_layout.addWidget(container)
            self.device_widgets.append(row_w)
            self._dev_row_containers.append(container)

    def _on_device_row_reorder(self, source_dev_id, target_dev_id):
        """设备行拖动排序回调"""
        src_idx = self.devices.index(self.devices[source_dev_id])
        tgt_idx = self.devices.index(self.devices[target_dev_id])
        if src_idx == tgt_idx:
            return
        # 移动设备
        dev = self.devices.pop(src_idx)
        self.devices.insert(tgt_idx, dev)
        # 同步颜色
        col = self.device_colors.pop(src_idx)
        self.device_colors.insert(tgt_idx, col)
        self._rebuild_device_after_change()

    def _create_device_row_widgets(self, dev_id, parent):
        """创建设备行控件（不含拖拽柄），返回 widget dict"""
        row_layout = QHBoxLayout(parent)
        row_layout.setContentsMargins(0, 0, 0, 2)
        row_layout.setSpacing(4)
        w = {}

        enable_check = QCheckBox()
        enable_check.setChecked(self.devices[dev_id]['enabled'])
        row_layout.addWidget(enable_check)
        w['enable'] = enable_check

        conn_combo = QComboBox()
        conn_combo.addItems(['serial', 'lan'])
        conn_combo.setCurrentText(self.devices[dev_id]['connection'])
        conn_combo.setMinimumWidth(70)
        row_layout.addWidget(conn_combo)
        w['connection'] = conn_combo

        name_edit = QLineEdit(self.devices[dev_id]['name'])
        name_edit.setFixedWidth(110)
        row_layout.addWidget(name_edit)
        w['name'] = name_edit

        # 串口 / IP 地址 双层切换：串口显示 PortComboBox（点击弹对话框），LAN显示QLineEdit
        port_ip_stack = QStackedWidget()
        port_combo = PortComboBox()
        port_combo.setMinimumWidth(110)
        ip_edit = QLineEdit()
        ip_edit.setPlaceholderText("IP地址")
        port_ip_stack.addWidget(port_combo)   # index 0: serial
        port_ip_stack.addWidget(ip_edit)       # index 1: lan
        if self.devices[dev_id]['connection'] == 'serial':
            port_ip_stack.setCurrentIndex(0)
            self._populate_port_combo(port_combo, self.devices[dev_id].get('port', ''))
        else:
            port_ip_stack.setCurrentIndex(1)
            ip_edit.setText(self.devices[dev_id]['host'])
        port_ip_stack.setFixedWidth(120)
        row_layout.addWidget(port_ip_stack)
        w['port_ip'] = port_ip_stack
        w['port_combo'] = port_combo
        w['ip_edit'] = ip_edit

        baud_port_edit = QLineEdit()
        if self.devices[dev_id]['connection'] == 'serial':
            baud_port_edit.setPlaceholderText("波特率")
            baud_port_edit.setText(self.devices[dev_id]['baudrate'])
        else:
            baud_port_edit.setPlaceholderText("端口号")
            baud_port_edit.setText(str(self.devices[dev_id]['lan_port']))
        baud_port_edit.setMinimumWidth(70)
        row_layout.addWidget(baud_port_edit)
        w['baud_port'] = baud_port_edit

        cmd_edit = QLineEdit(self.devices[dev_id]['read_command'])
        cmd_edit.setPlaceholderText("如：READ?\\r\\n")
        cmd_edit.setMinimumWidth(220)
        row_layout.addWidget(cmd_edit, 1)
        w['cmd'] = cmd_edit

        # 自动检测勾选
        auto_chk = QCheckBox("Stability")
        auto_chk.setChecked(self.devices[dev_id].get('auto_test', False))
        auto_chk.setToolTip("勾选后此通道参与T0/T1/T2/T3/Std1/Std2/Avg1/Avg2检测")
        auto_chk.stateChanged.connect(lambda s, idx=dev_id: self._on_auto_test_toggled(idx, s))
        row_layout.addWidget(auto_chk)
        w['auto_test'] = auto_chk

        # 曲线显示勾选框：选中(开启)才显示该通道实时数据曲线，风格与 Stability 一致
        curve_btn = QCheckBox("曲线")
        curve_btn.setChecked(self.devices[dev_id].get('curve_visible', True))
        curve_btn.setToolTip("开启后显示该通道实时数据曲线")
        curve_btn.stateChanged.connect(lambda s, idx=dev_id: self._on_curve_btn_toggled(idx, bool(s)))
        row_layout.addWidget(curve_btn)
        w['curve_btn'] = curve_btn

        status_label = QLabel("未连接")
        status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        row_layout.addWidget(status_label)
        w['status'] = status_label

        name_edit.textChanged.connect(lambda text, idx=dev_id: self.on_device_name_changed(idx, text))
        enable_check.stateChanged.connect(lambda s, idx=dev_id: self.update_device_config(idx))
        conn_combo.currentTextChanged.connect(lambda text, idx=dev_id: self.on_connection_changed(idx, text))
        port_combo.currentTextChanged.connect(lambda text, idx=dev_id: self.update_device_config(idx))
        ip_edit.textChanged.connect(lambda text, idx=dev_id: self.update_device_config(idx))
        baud_port_edit.textChanged.connect(lambda text, idx=dev_id: self.update_device_config(idx))
        cmd_edit.textChanged.connect(lambda text, idx=dev_id: self.update_device_config(idx))
        return w

    def _on_auto_test_toggled(self, dev_id, state):
        """auto_test 勾选状态变更"""
        if hasattr(self, '_loading') and self._loading:
            return
        self.devices[dev_id]['auto_test'] = bool(state)
        self.update_device_config(dev_id)

    def _on_curve_btn_toggled(self, dev_id, checked):
        """曲线按钮切换：选中才显示该通道实时数据曲线"""
        if hasattr(self, '_loading') and self._loading:
            return
        self.devices[dev_id]['curve_visible'] = bool(checked)
        self.update_device_config(dev_id)
        # 立即刷新曲线显示
        self._update_curve_visibility()

    def _update_curve_visibility(self):
        """根据曲线按钮状态立即刷新各通道曲线显示"""
        if not hasattr(self, 'curves'):
            return
        # 用实际长度保护，避免 curves 未初始化或长度不足时索引越界
        n = min(self._dev_row_count, len(self.curves))
        for i in range(n):
            if self.curves[i] is not None:
                try:
                    visible = (self.devices[i].get('enabled', False)
                               and self.devices[i].get('curve_visible', True))
                    self.curves[i].setVisible(visible)
                except RuntimeError:
                    self.curves[i] = None

    def _dev_add_row(self):
        """设备增加一行"""
        if self._dev_row_count >= 20:
            return
        new_dev = {'enabled': False, 'connection': 'serial', 'port': '', 'baudrate': '9600',
                   'name': f'设备{self._dev_row_count+1}', 'read_command': '',
                   'host': '', 'lan_port': '', 'auto_test': False, 'curve_visible': True}
        self.devices.append(new_dev)
        self._dev_row_count += 1
        # 确保颜色列表足够
        extra_palette = ['#17becf', '#bcbd22', '#7f7f7f', '#e377c2', '#8c564b', '#9467bd',
                         '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b',
                         '#393b79', '#637939', '#8c6d31', '#843c39', '#7b4173', '#5254a3']
        while len(self.device_colors) < self._dev_row_count:
            self.device_colors.append(extra_palette[len(self.device_colors) % len(extra_palette)])
        self._rebuild_device_after_change()

    def _dev_remove_row(self):
        """设备减少一行"""
        if self._dev_row_count <= 1:
            return
        self.devices.pop()
        self._dev_row_count -= 1
        self._rebuild_device_after_change()

    def _rebuild_device_after_change(self):
        """重建设备行 UI 并保存配置"""
        self._build_device_rows()
        self.save_config()
        # 重新连接所有设备（如果正在采集则停止后重新开始）
        if self.test_running:
            self.stop_collection()
        # 重新初始化图和图例（init_plots 会根据 _dev_row_count 添加对应数量的图例）
        self.init_plots()
        # 由于 _build_device_rows 中重建复选框会触发 stateChanged 信号，
        # 信号会在事件队列中延迟执行，覆盖 init_plots 创建的图例显示状态。
        # 这里强制将所有设备的显示状态与 enable 状态同步。
        for i in range(self._dev_row_count):
            if hasattr(self, 'legend_widget'):
                self.legend_widget.set_device_visible(i, self.devices[i]['enabled'])

    def _ts_add_row(self):
        """温度源增加一行设置（单行紧凑布局）"""
        if self._ts_row_count >= 10:
            return
        row_idx = self._ts_row_count

        no_arrow_style = """
            QDoubleSpinBox::up-button, QDoubleSpinBox::down-button {
                width: 0px; height: 0px;
            }
            QDoubleSpinBox { padding: 0px 2px; border: 1px solid #999999; }
        """

        def mk_spin(range_min, val, step, decimal, width, suffix='', nullable=False):
            # nullable=True 时使用可清空输入框，留空表示使用默认值（不发送）。
            # 新添加的行 PID 参数默认留空（用温度源默认值），而非沿用固定数值
            s = NullableSpinBox() if nullable else QDoubleSpinBox()
            s.setRange(range_min, 9999)
            if nullable:
                s.setValue(None)  # 留空：使用默认值
            else:
                s.setValue(val)
            s.setSingleStep(step); s.setDecimals(decimal)
            s.setFixedWidth(width); s.setStyleSheet(no_arrow_style)
            if nullable:
                s.setPlaceholderText("默认")
            if suffix: s.setSuffix(suffix)
            return s

        ts_container = DragDropRowWidget(row_idx)
        ts_container.reorder_requested.connect(self._ts_reorder_rows)
        row_layout = QHBoxLayout(ts_container.contentWidget())
        row_layout.setSpacing(2)
        row_layout.setContentsMargins(1, 0, 1, 0)

        chk = QCheckBox()
        chk.setChecked(True)
        chk.stateChanged.connect(lambda v, idx=row_idx: self.save_config())
        row_layout.addWidget(chk)
        self.row_checks.append(chk)

        # 发送命令按钮：放在最左侧（勾选框之后），点击后发送该行的所有参数
        send_btn = QPushButton("发送")
        send_btn.setFixedWidth(46)
        send_btn.setToolTip("发送该行的所有参数到温度源")
        send_btn.setStyleSheet(
            "QPushButton{background:#2196F3;color:white;font-weight:bold;font-size:11px;border-radius:4px;padding:0px;}"
            "QPushButton:hover{background:#1976D2;}"
            "QPushButton:disabled{background:#ccc;color:#888;}"
        )
        # 点击时按当前列表位置解析行号，保证拖动排序后仍对应正确行
        send_btn.clicked.connect(lambda _, btn=send_btn: self._send_row_command(self.row_send_btns.index(btn)))
        send_btn.setEnabled(self.temp_source_connected)  # 未通信时禁用
        row_layout.addWidget(send_btn)
        self.row_send_btns.append(send_btn)

        row_layout.addWidget(QLabel("SP"))
        sp = mk_spin(0, 25, 1, 1, 70, nullable=True)  # 新行留空，使用默认值
        sp.valueChanged.connect(lambda v, idx=row_idx: self._on_row_setpoint_changed(idx, v))
        row_layout.addWidget(sp)
        self.row_setpoint_spins.append(sp)

        row_layout.addWidget(QLabel("Spec"))
        spec = mk_spin(0, 0.5, 0.1, 2, 55, nullable=True)  # 新行留空，使用默认值
        spec.valueChanged.connect(lambda v, idx=row_idx: self._on_row_spec_changed(idx, v))
        row_layout.addWidget(spec)
        self.row_setpoint_spec.append(spec)

        row_layout.addWidget(QLabel("M"))
        main_spin = mk_spin(0, 50, 0.01, 2, 55, nullable=True)  # 新行留空，使用默认值
        main_spin.valueChanged.connect(lambda v, idx=row_idx: self._on_row_main_changed(idx, v))
        row_layout.addWidget(main_spin)
        self.row_main_spins.append(main_spin)

        row_layout.addWidget(QLabel("M-P"))
        mp = mk_spin(0, 10, 0.1, 1, 55, nullable=True)
        mp.valueChanged.connect(lambda v: self.save_config())
        row_layout.addWidget(mp)
        self.row_main_pid_p.append(mp)

        row_layout.addWidget(QLabel("M-I"))
        mi = mk_spin(0, 200, 1, 1, 55, nullable=True)
        mi.valueChanged.connect(lambda v: self.save_config())
        row_layout.addWidget(mi)
        self.row_main_pid_i.append(mi)

        row_layout.addWidget(QLabel("M-D"))
        md = mk_spin(0, 50, 1, 1, 55, nullable=True)
        md.valueChanged.connect(lambda v: self.save_config())
        row_layout.addWidget(md)
        self.row_main_pid_d.append(md)

        row_layout.addWidget(QLabel("S"))
        sec_spin = mk_spin(0, 0, 1, 1, 55, nullable=True)  # 新行留空，使用默认值
        sec_spin.valueChanged.connect(lambda v, idx=row_idx: self._on_row_sec_changed(idx, v))
        row_layout.addWidget(sec_spin)
        self.row_sec_spins.append(sec_spin)

        row_layout.addWidget(QLabel("S-P"))
        sp_p = mk_spin(0, 10, 0.1, 1, 55, nullable=True)
        sp_p.valueChanged.connect(lambda v: self.save_config())
        row_layout.addWidget(sp_p)
        self.row_sec_pid_p.append(sp_p)

        row_layout.addWidget(QLabel("S-I"))
        si = mk_spin(0, 200, 1, 1, 55, nullable=True)
        si.valueChanged.connect(lambda v: self.save_config())
        row_layout.addWidget(si)
        self.row_sec_pid_i.append(si)

        row_layout.addWidget(QLabel("S-D"))
        sd = mk_spin(0, 50, 1, 1, 55, nullable=True)
        sd.valueChanged.connect(lambda v: self.save_config())
        row_layout.addWidget(sd)
        self.row_sec_pid_d.append(sd)

        # Weight：已取消显示，保留空列表维持索引对齐（数据不再通过 UI 输入）
        self.row_weights.append([])

        row_layout.addStretch()  # 输入框靠左，不拉伸到右侧

        # 收集该行的"高级控件"（Const 1210 时需要隐藏，从 index 5 开始"
        advanced = []
        for idx in range(5, row_layout.count()):
            item = row_layout.itemAt(idx)
            if item and item.widget():
                advanced.append(item.widget())
        self._ts_row_advanced_widgets.append(advanced)

        # 插入到自动测试框的 SP 行布局末尾
        temp_ctrl_layout = getattr(self, 'ts_rows_layout', None)
        if temp_ctrl_layout is not None:
            temp_ctrl_layout.addWidget(ts_container)
        self._ts_row_layouts.append(ts_container)

        self._ts_row_count += 1

        # 确保新行的可见性与当前设备类型一致
        if self._ts_device_type == 'Const 1210':
            for w in advanced:
                w.setVisible(False)

        self.save_config()

    def _ts_remove_row(self, silent=False):
        """温度源减少一行设置"""
        if self._ts_row_count <= 1:
            return
        if not getattr(self, 'ts_rows_layout', None):
            return
        # 移除最后一行
        if self._ts_row_layouts:
            last_container = self._ts_row_layouts.pop()
            last_container.setParent(None)
            last_container.deleteLater()
        if self.row_checks:
            self.row_checks.pop()
        if self.row_setpoint_spins:
            self.row_setpoint_spins.pop()
        if self.row_setpoint_spec:
            self.row_setpoint_spec.pop()
        if self.row_main_spins:
            self.row_main_spins.pop()
        if self.row_sec_spins:
            self.row_sec_spins.pop()
        if self.row_send_btns:
            self.row_send_btns.pop()
        if self.row_main_pid_p:
            self.row_main_pid_p.pop()
        if self.row_main_pid_i:
            self.row_main_pid_i.pop()
        if self.row_main_pid_d:
            self.row_main_pid_d.pop()
        if self.row_sec_pid_p:
            self.row_sec_pid_p.pop()
        if self.row_sec_pid_i:
            self.row_sec_pid_i.pop()
        if self.row_sec_pid_d:
            self.row_sec_pid_d.pop()
        if self.row_weights:
            self.row_weights.pop()
        if self._ts_row_advanced_widgets:
            self._ts_row_advanced_widgets.pop()
        self._ts_row_count -= 1
        if not silent:
            self.save_config()

    def _ts_reorder_rows(self, source_dev_id, target_dev_id):
        """温度源行拖动排序回调"""
        src = source_dev_id
        tgt = target_dev_id
        if src == tgt:
            return
        temp_ctrl_layout = getattr(self, 'ts_rows_layout', None)
        if not temp_ctrl_layout:
            return
        # 找到实际索引
        src_idx = None
        tgt_idx = None
        for i, c in enumerate(self._ts_row_layouts):
            if c.device_id == src:
                src_idx = i
            if c.device_id == tgt:
                tgt_idx = i
        if src_idx is None or tgt_idx is None or src_idx == tgt_idx:
            return

        def move_item(lst, s, t):
            item = lst.pop(s)
            lst.insert(t, item)

        for lst in [self.row_checks, self.row_setpoint_spins, self.row_setpoint_spec,
                    self.row_main_spins, self.row_sec_spins, self.row_main_pid_p,
                    self.row_main_pid_d, self.row_sec_pid_p, self.row_sec_pid_i,
                    self.row_sec_pid_d, self.row_weights, self.row_send_btns,
                    self._ts_row_layouts, self._ts_row_advanced_widgets]:
            move_item(lst, src_idx, tgt_idx)

        # 重建自动测试框内的 SP 行布局：保存表头（index 0），重排其余行
        header_item = temp_ctrl_layout.takeAt(0)
        while temp_ctrl_layout.count():
            item = temp_ctrl_layout.takeAt(0)
            if item.widget():
                item.widget().setParent(None)
        # 重新加入：表头 + 新顺序行
        temp_ctrl_layout.addItem(header_item)
        for c in self._ts_row_layouts:
            temp_ctrl_layout.addWidget(c)
        self.save_config()

    def _check_manual_test_completion(self):
        """手动测试模式：所有勾选 Stability 通道的 T3 到达后，保存 excel+截图并提示。
           温度源保持运行，继续工作。"""
        try:
            # 勾选了 Stability 且启用的通道
            auto_chs = [d for d in range(self._dev_row_count)
                        if self.devices[d].get('auto_test', False) and self.devices[d].get('enabled', False)]
            if not auto_chs:
                return
            # 全部通道 T3 完成
            all_done = all(
                d in self.auto_test_state
                and self.auto_test_state[d].get('phase') == 'complete'
                and self.auto_test_state[d].get('T3') is not None
                for d in auto_chs
            )
            if not all_done:
                return
            self._manual_test_done = True
            # 构建 T3 汇总（通道1-5 的 avg/std/max/min + t0/t1）
            self.auto_test_summary = self._build_auto_test_summary()
            # 1) 保存 stability（写入 current_data_file）
            stability_ok = False
            try:
                self.auto_save_data()
                stability_ok = True
            except Exception as e:
                print(f"[手动测试完成] Excel保存失败: {e}")
            # 2) 计算 accuracy：通道5(User=dev4)和通道1(Fix=dev0)在 T3 时刻倒数 10min 的 max/avg/min
            accuracy_rows = self._build_accuracy_rows(window_minutes=10)
            # 3) 保存 accuracy sheet 到同一 Excel
            accuracy_ok = False
            try:
                if accuracy_rows:
                    self._save_accuracy_sheet(accuracy_rows)
                    accuracy_ok = True
            except Exception as e:
                print(f"[手动测试完成] accuracy保存失败: {e}")
            # 4) 保存截图
            try:
                self._save_whole_window_screenshot('T3')
            except Exception as e:
                print(f"[手动测试完成] 截图失败: {e}")
            msg = "数据截图已保存"
            if stability_ok and accuracy_ok:
                msg += "\nstability 与 accuracy 全部保存完成"
            elif stability_ok:
                msg += "\n注意：accuracy 写入失败，仅 stability 已保存"
            elif accuracy_ok:
                msg += "\n注意：stability 写入失败，仅 accuracy 已保存"
            else:
                msg += "\n注意：stability 与 accuracy 均写入失败"
            QMessageBox.information(self, "完成", msg)
            self.status_label.setText("测试完成，数据与截图已保存，停止采集")
            # T3 到达后停止采集（确保 stability 和 accuracy 都保存完成后再停）
            try:
                self.stop_collection()
            except Exception as e:
                print(f"[手动测试完成] 停止采集失败: {e}")
        except Exception as e:
            print(f"[手动测试完成] 异常: {e}")

    def _build_accuracy_rows(self, window_minutes=10):
        """计算通道5(User=dev4)和通道1(Fix=dev0)在 T3 时刻倒数 window_minutes 分钟内的 max/avg/min。
        返回 [('No.', row_label, ...)] 列表形如 [{'No.': 'User', 'Max': ..., 'Avg': ..., 'Min': ...}, ...]。"""
        rows = []
        # 通道1(Fix)与通道5(User)对应 dev_id
        targets = [
            (4, 'User'),   # 通道5
            (0, 'Fix'),    # 通道1
        ]
        # T3 相对 start_time 的分钟数取通道4(User)或通道1(Fix)的 T3（任一已完成即可）
        t3_min = None
        for d in (4, 0):
            st = self.auto_test_state.get(d)
            if st and st.get('phase') == 'complete' and st.get('T3') is not None:
                t3_min = float(st['T3'])
                break
        if t3_min is None:
            print("[accuracy] 未找到有效的 T3 时刻")
            return rows
        win_sec = float(window_minutes) * 60.0
        t3_sec = t3_min * 60.0
        win_start_sec = max(0.0, t3_sec - win_sec)
        for dev_id, label in targets:
            entry = {'No.': label, 'Max': None, 'Avg': None, 'Min': None}
            try:
                if not (0 <= dev_id < self._dev_row_count):
                    rows.append(entry)
                    continue
                if not self.devices[dev_id].get('enabled', False):
                    rows.append(entry)
                    continue
                tb = self.time_buffer.get(dev_id)
                db = self.data_buffer.get(dev_id)
                if not tb or not db:
                    rows.append(entry)
                    continue
                vals = []
                for i, v in enumerate(db):
                    if i < len(tb):
                        ts = float(tb[i])
                        if win_start_sec <= ts <= t3_sec:
                            try:
                                vals.append(float(v))
                            except (TypeError, ValueError):
                                continue
                if vals:
                    entry['Max'] = float(np.max(vals))
                    entry['Avg'] = float(np.mean(vals))
                    entry['Min'] = float(np.min(vals))
            except Exception as e:
                print(f"[accuracy] 计算通道{dev_id+1}({label})失败: {e}")
            rows.append(entry)
        return rows

    def _save_accuracy_sheet(self, accuracy_rows):
        """将 accuracy_rows 写入 current_data_file 的 'accuracy' sheet。
        若文件已存在则追加/覆盖该 sheet；否则新建文件。
        与 stability 写入分离：先调 auto_save_data() 后再调本函数，确保两个 sheet 共存。"""
        if not accuracy_rows:
            return
        import os
        from openpyxl import load_workbook
        full_path = os.path.abspath(self.current_data_file)
        # 计算列顺序：'No.' 在第1列，其他数值列保留插入顺序
        cols = ['No.', 'Max', 'Avg', 'Min']
        df = pd.DataFrame(accuracy_rows, columns=cols)
        # 先写入临时文件，然后用 load_workbook 合并到 full_path
        tmp_path = full_path + '.accuracy.tmp'
        try:
            with pd.ExcelWriter(tmp_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='accuracy', index=False)
            if os.path.exists(full_path):
                wb = load_workbook(full_path)
                if 'accuracy' in wb.sheetnames:
                    del wb['accuracy']
                # 从临时文件读取 accuracy 并追加
                wb_tmp = load_workbook(tmp_path)
                src = wb_tmp['accuracy']
                dst = wb.create_sheet('accuracy')
                for row in src.iter_rows(values_only=True):
                    dst.append(row)
                wb_tmp.close()
                wb.save(full_path)
            else:
                # 文件不存在时直接把临时文件改名
                if os.path.exists(tmp_path):
                    os.replace(tmp_path, full_path)
            self.status_label.setText(f"accuracy 数据已保存至: {full_path}")
            print(f"[accuracy] 已写入 sheet 'accuracy' -> {full_path}")
        finally:
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except Exception:
                pass

    def _build_auto_test_summary(self):
        """T3 完成后构建跨通道汇总。
        通道编号映射 dev_id：通道1=dev0, 通道2=dev1, 通道3=dev2, 通道4=dev3, 通道5=dev4。
        Main=通道3 avg, Sec=通道4 avg, User=通道5 avg, U-Std=通道5 std,
        F-avg=通道1 avg, M-avg=通道2 avg, F-std=通道1 std,
        F-Max/F-Min=通道1 T3窗口极值, t0/t1=通道1 T0/T1。
        对于未勾选 Stability 的通道，auto_test_state 中无 avg2/std2，
        这里 fallback 使用 calculate_volatility 当前窗口的 avg/std，避免漏记。"""
        def _safe(d, key):
            try:
                st = self.auto_test_state.get(d, {})
                return st.get(key)
            except Exception:
                return None

        def _val_or_calc(d, key):
            # 优先用 auto_test_state（如有 T3 检测结果），否则用当前窗口 calculate_volatility
            v = _safe(d, key)
            if v is not None:
                return v
            if not (0 <= d < self._dev_row_count):
                return None
            if not self.devices[d].get('enabled', False):
                return None
            try:
                res = self.calculate_volatility(d)
                if res and len(res) >= 5:
                    if key == 'avg2':
                        return res[4]
                    if key == 'std2':
                        return res[0]
            except Exception:
                pass
            return None

        # 通道1（dev0）F-Max/F-Min：T3 时刻所在 30min 窗口的极值
        f_max = f_min = None
        try:
            if 0 < len(self.data_buffer) and len(self.data_buffer[0]) > 0:
                now_ts = time.time()
                start_ts = now_ts - 30 * 60
                vals = []
                tb = self.time_buffer[0]
                db = self.data_buffer[0]
                for i, v in enumerate(db):
                    if i < len(tb) and tb[i] >= start_ts:
                        vals.append(v)
                if vals:
                    f_max = float(np.max(vals))
                    f_min = float(np.min(vals))
        except Exception:
            f_max = f_min = None

        summary = {
            'Main': _val_or_calc(2, 'avg2'),
            'Sec': _val_or_calc(3, 'avg2'),
            'User': _val_or_calc(4, 'avg2'),
            'U-Std': _val_or_calc(4, 'std2'),
            'F-avg': _val_or_calc(0, 'avg2'),
            'M-avg': _val_or_calc(1, 'avg2'),
            'F-std': _val_or_calc(0, 'std2'),
            'F-Max': f_max,
            'F-Min': f_min,
            't0': _safe(0, 'T0'),
            't1': _safe(0, 'T1'),
        }
        return summary

    def _check_special_channels(self):
        """所有勾选了 Stability 的通道自动检测：T0/T1/T2/T3/Std1/Std2/Avg1/Avg2"""
        for dev_id in range(self._dev_row_count):
            if not self.devices[dev_id].get('auto_test', False):
                continue
            if dev_id not in self.auto_test_state:
                continue
            if not self.devices[dev_id]['enabled'] or len(self.data_buffer[dev_id]) < 2:
                continue
            state = self.auto_test_state[dev_id]
            set_temp = self.current_test_temp
            elapsed_min = (time.time() - self.start_time) / 60.0 if self.start_time else 0
            # 获取当前最新温度
            latest_temp = self.data_buffer[dev_id][-1]
            if state['phase'] == 'idle':
                # 根据 Spec 判断首次到达设定温度阈值
                spec = self.current_test_spec
                # 升温方向：当前温度低于设定值 → 到达阈值 = setpoint - spec
                # 降温方向：当前温度高于设定值 → 到达阈值 = setpoint + spec
                if latest_temp < set_temp:
                    # 升温：到达 setpoint - spec 即认为到达
                    t0_threshold = set_temp - spec
                    t0_condition = latest_temp >= t0_threshold
                elif latest_temp > set_temp:
                    # 降温：到达 setpoint + spec 即认为到达
                    t0_threshold = set_temp + spec
                    t0_condition = latest_temp <= t0_threshold
                else:
                    # 正好等于设定值
                    t0_condition = True
                if t0_condition:
                    state['T0'] = elapsed_min
                    state['T0_time'] = time.time()
                    state['phase'] = 'T0_reached'
                    # 重置该通道
                    self.legend_widget.volatility_reset_times[dev_id] = datetime.now()
                    self.legend_widget.reset_volatility(dev_id)
                    self.legend_widget.update_auto_test(dev_id, t0=state['T0'])
                    self._update_plot_annotations()
                    self.status_label.setText(f"通道{dev_id+1}到达{set_temp}°C, T0={state['T0']:.1f}min")
                    # 记录T0
                    if 'T0' not in self.auto_test_logged.get(dev_id, set()):
                        ch_name = self.devices[dev_id]['name']
                        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        self.auto_test_log.append((now_str, ch_name, state['T0'], None, None, None, None, None, None, None, None))
                        self.auto_test_logged.setdefault(dev_id, set()).add('T0')
            elif state['phase'] == 'T0_reached':
                if state['T0_time'] is None:
                    continue
                elapsed_since_T0 = (time.time() - state['T0_time']) / 60.0
                if elapsed_since_T0 < 30:
                    continue
                # 获取当前波动（从重置时间起算）
                result = self.calculate_volatility(dev_id)
                if result is None:
                    continue
                volatility = result[0]
                avg_val = result[4] if len(result) >= 5 else None
                if volatility <= self.stability_threshold:
                    state['T2'] = elapsed_min
                    state['T2_time'] = time.time()
                    state['T1'] = state['T2'] - 30.0
                    state['std1'] = volatility
                    state['avg1'] = avg_val
                    state['phase'] = 'T2_found'
                    self.legend_widget.update_auto_test(dev_id, t0=state['T0'], t1=state['T1'],
                                                        t2=state['T2'], vol1=state['std1'],
                                                        avg1=state['avg1'])
                    self._update_plot_annotations()
                    self.status_label.setText(f"通道{dev_id+1}稳定, T2={state['T2']:.1f}min, std1={state['std1']:.4f}")
                    # 记录T2/Std1
                    if 'T2' not in self.auto_test_logged.get(dev_id, set()):
                        ch_name = self.devices[dev_id]['name']
                        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        t0t1 = state['T1'] - state['T0'] if state['T1'] is not None and state['T0'] is not None else None
                        self.auto_test_log.append((now_str, ch_name, state['T0'], t0t1, state['T1'],
                                                    state['T2'], None,
                                                    state['std1'], None,
                                                    state['avg1'], None))
                        self.auto_test_logged.setdefault(dev_id, set()).add('T2')
            elif state['phase'] == 'T2_found':
                if state['T2_time'] is None:
                    continue
                elapsed_since_T2 = (time.time() - state['T2_time']) / 60.0
                if elapsed_since_T2 >= 30:
                    state['T3'] = elapsed_min
                    result = self.calculate_volatility(dev_id)
                    state['std2'] = result[0] if result else None
                    state['avg2'] = result[4] if result and len(result) >= 5 else None
                    state['phase'] = 'complete'
                    self.legend_widget.update_auto_test(dev_id, t0=state['T0'], t1=state['T1'],
                                                        t2=state['T2'], t3=state['T3'],
                                                        vol1=state['std1'], vol2=state['std2'],
                                                        avg1=state['avg1'], avg2=state['avg2'])
                    # 记录到日志缓冲区
                    if 'complete' not in self.auto_test_logged.get(dev_id, set()):
                        ch_name = self.devices[dev_id]['name']
                        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        t0t1 = state['T1'] - state['T0'] if state['T1'] is not None and state['T0'] is not None else None
                        self.auto_test_log.append((now_str, ch_name, state['T0'], t0t1, state['T1'],
                                                    state['T2'], state['T3'],
                                                    state['std1'], state['std2'],
                                                    state['avg1'], state['avg2']))
                        self.auto_test_logged.setdefault(dev_id, set()).add('complete')
                    self.status_label.setText(f"通道{dev_id+1}完成, T3={state['T3']:.1f}min, std2={state['std2']:.4f}")
                self._update_plot_annotations()

    def _update_plot_annotations(self):
        """在曲线图上标注 T0/T1/T2/T3 垂直线"""
        # 清除旧标注线
        for line in self.auto_test_lines:
            try:
                if self.primary_plot:
                    self.primary_plot.removeItem(line)
            except:
                pass
        self.auto_test_lines.clear()
        # 取第一个 auto_test 通道绘制标注（避免重复）
        if not self.auto_test_state:
            return
        first_dev = next(iter(self.auto_test_state))
        state0 = self.auto_test_state.get(first_dev, {})
        if state0.get('T0') is None:
            return
        labels = []
        if state0.get('T0') is not None:
            labels.append((state0['T0'], 'T0', (0.66, 0.27, 0.66)))
        if state0.get('T1') is not None:
            labels.append((state0['T1'], 'T1', (0.80, 0.40, 0.00)))
        if state0.get('T2') is not None:
            labels.append((state0['T2'], 'T2', (0.00, 0.60, 0.00)))
        if state0.get('T3') is not None:
            labels.append((state0['T3'], 'T3', (0.20, 0.40, 0.80)))
        for t_val, label, color in labels:
            line = pg.InfiniteLine(pos=t_val, angle=90, pen=pg.mkPen(color, width=2, style=Qt.PenStyle.DashLine),
                                   label=label, labelOpts={'color': color, 'position': 0.95})
            if self.primary_plot:
                self.primary_plot.addItem(line)
                self.auto_test_lines.append(line)

    def init_plots(self):
        self.plot_widget.clear()
        self.curves = [None] * self._dev_row_count
        self.device_viewboxes = {}
        self.extra_vb_list = []
        self.quantity_viewbox_map = {}
        self.device_quantity_info = {}
        self.primary_plot = None

        # 初始默认单轴图（开始采集时会根据命令重建多轴图）
        self.primary_plot = self.plot_widget.addPlot()
        self.primary_plot.setLabel('bottom', '时间(min)')
        self.primary_plot.setLabel('left', '温度(°C)')
        self.primary_plot.showGrid(x=True, y=True, alpha=0.3)

        for i in range(self._dev_row_count):
            color = self.device_colors[i]
            c = self.primary_plot.plot([], [], pen=pg.mkPen(color, width=2), symbol='o', symbolSize=2.5,
                                       symbolBrush=pg.mkBrush(color), symbolPen=pg.mkPen(color, width=0.5))
            self.curves[i] = c
            self.device_viewboxes[i] = self.primary_plot.vb
            _, _, unit = self.determine_quantity_type(self.devices[i].get('read_command', ''))
            self.legend_widget.add_device(i, self.devices[i]['name'], color, unit, self.devices[i].get('auto_test', False))

    def on_device_name_changed(self, dev_id, new_name):
        self.devices[dev_id]['name'] = new_name
        self.legend_widget.update_device_name(dev_id, new_name)
        self.save_config()

    def open_debug_window(self):
        if self.debug_dialog is None:
            self.debug_dialog = SerialDebugDialog(self)
        self.debug_dialog.show()
        self.debug_dialog.raise_()

    def on_debug_info(self, device_id, command, response, parsed_value, success, error_msg):
        # 问询（读取/查询温度）命令不记录、不打印
        is_query = ('?' in command) or command.strip().upper().startswith(('READ', 'MEAS', 'FETCH'))
        if is_query:
            return
        if self.debug_dialog and self.debug_dialog.isVisible():
            self.debug_dialog.append_comm_detail(device_id, command, response, parsed_value, success, error_msg)
        # 仅控制类命令失败时打印到终端区
        if hasattr(self, 'legend_widget') and not success:
            dev_name = self.devices[device_id].get('name', f'设备{device_id+1}') if device_id < len(self.devices) else f'设备{device_id+1}'
            self.legend_widget.append_terminal(f"{dev_name} 发送:{command.strip()} 失败: {error_msg}")

    def save_config_to_file(self):
        path, _ = QFileDialog.getSaveFileName(self, "保存配置", "serial_config.json", "JSON (*.json)")
        if not path:
            return
        try:
            cfg = {"devices": self.devices, "interval": self.interval_spin.value()}
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(cfg, f, indent=4, ensure_ascii=False)
            QMessageBox.information(self, "成功", f"配置已保存至：{path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存失败：{str(e)}")

    def load_config_from_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "导入配置", "", "JSON (*.json)")
        if not path:
            return
        try:
            with open(path, 'r', encoding='utf-8') as f:
                cfg = json.load(f)
            if 'devices' in cfg:
                for i, d in enumerate(cfg['devices']):
                    if i < 6:
                        self.devices[i].update(d)
                        self.device_widgets[i]['name'].setText(d.get('name', ''))
                        self.device_widgets[i]['enable'].setChecked(d.get('enabled', False))
                        conn = d.get('connection', 'serial')
                        self.device_widgets[i]['connection'].setCurrentText(conn)
                        if conn == 'serial':
                            self.device_widgets[i]['port_ip'].setCurrentIndex(0)
                            self._populate_port_combo(self.device_widgets[i]['port_combo'], d.get('port', ''))
                            self.device_widgets[i]['baud_port'].setText(d.get('baudrate', '9600'))
                        else:
                            self.device_widgets[i]['port_ip'].setCurrentIndex(1)
                            self.device_widgets[i]['ip_edit'].setText(d.get('host', ''))
                            lan_port = d.get('lan_port', '8000')
                            if not lan_port:
                                lan_port = '8000'
                            self.device_widgets[i]['baud_port'].setText(str(lan_port))
                        self.device_widgets[i]['cmd'].setText(d.get('read_command', ''))
                        if 'auto_test' in self.device_widgets[i]:
                            self.device_widgets[i]['auto_test'].setChecked(d.get('auto_test', False))
                        if 'curve_btn' in self.device_widgets[i]:
                            self.device_widgets[i]['curve_btn'].setChecked(d.get('curve_visible', True))
            if 'interval' in cfg:
                self.interval_spin.setValue(cfg['interval'])
            self.save_config()
            QMessageBox.information(self, "成功", "配置导入成功")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导入失败：{str(e)}")

    def load_config(self):
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)
                if 'devices' in cfg:
                    dev_list = cfg['devices']
                    # 先扩展设备列表到正确数量
                    while len(self.devices) < len(dev_list):
                        self.devices.append({'enabled': False, 'connection': 'serial', 'port': '', 'baudrate': '9600',
                                             'name': f'设备{len(self.devices)+1}', 'read_command': '', 'host': '', 'lan_port': ''})
                    self._dev_row_count = len(dev_list)
                    # 确保颜色列表足够（旧配置只有6种颜色，扩展后>6的设备需要颜色）
                    extra_palette = ['#17becf', '#bcbd22', '#7f7f7f', '#e377c2', '#8c564b', '#9467bd',
                                     '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b',
                                     '#393b79', '#637939', '#8c6d31', '#843c39', '#7b4173', '#5254a3']
                    while len(self.device_colors) < self._dev_row_count:
                        self.device_colors.append(extra_palette[len(self.device_colors) % len(extra_palette)])
                    self._build_device_rows()
                    for i, d in enumerate(dev_list):
                        if i < len(self.devices):
                            self.devices[i].update(d)
                        if i < len(self.device_widgets):
                            self.device_widgets[i]['name'].setText(d.get('name', ''))
                            self.device_widgets[i]['enable'].setChecked(d.get('enabled', False))
                            conn = d.get('connection', 'serial')
                            self.device_widgets[i]['connection'].setCurrentText(conn)
                            if conn == 'serial':
                                self.device_widgets[i]['port_ip'].setCurrentIndex(0)
                                self._populate_port_combo(self.device_widgets[i]['port_combo'], d.get('port', ''))
                                self.device_widgets[i]['baud_port'].setText(d.get('baudrate', '9600'))
                            else:
                                self.device_widgets[i]['port_ip'].setCurrentIndex(1)
                                self.device_widgets[i]['ip_edit'].setText(d.get('host', ''))
                                lan_port = d.get('lan_port', '8000')
                                if not lan_port:
                                    lan_port = '8000'
                                self.device_widgets[i]['baud_port'].setText(str(lan_port))
                            self.device_widgets[i]['cmd'].setText(d.get('read_command', ''))
                            if 'auto_test' in self.device_widgets[i]:
                                self.device_widgets[i]['auto_test'].setChecked(d.get('auto_test', False))
                            if 'curve_btn' in self.device_widgets[i]:
                                self.device_widgets[i]['curve_btn'].setChecked(d.get('curve_visible', True))
                    # 重新初始化图表和图例（设备数据已更新，使用正确的命令判断单位）
                    self.init_plots()
                if 'interval' in cfg:
                    self.interval_spin.setValue(cfg['interval'])
                if 'test_interval' in cfg:
                    self.test_interval_spin.setValue(cfg['test_interval'])
                if 'ts_row_count' in cfg and cfg['ts_row_count'] > 3:
                    needed = cfg['ts_row_count'] - 3
                    for _ in range(needed):
                        self._ts_add_row()
                if 'row_setpoints' in cfg:
                    for i, v in enumerate(cfg['row_setpoints']):
                        if i < self._ts_row_count:
                            self.row_setpoint_spins[i].setValue(v)
                if 'row_specs' in cfg:
                    for i, v in enumerate(cfg['row_specs']):
                        if i < self._ts_row_count:
                            self.row_setpoint_spec[i].setValue(v)
                if 'row_mains' in cfg:
                    for i, v in enumerate(cfg['row_mains']):
                        if i < self._ts_row_count:
                            self.row_main_spins[i].setValue(v)
                if 'row_secs' in cfg:
                    for i, v in enumerate(cfg['row_secs']):
                        if i < self._ts_row_count:
                            self.row_sec_spins[i].setValue(v)
                if 'row_checks' in cfg:
                    for i, v in enumerate(cfg['row_checks']):
                        if i < self._ts_row_count:
                            self.row_checks[i].setChecked(v)
                if 'temp_source_conn_prefs' in cfg:
                    self._ts_device_conn_prefs.update(cfg['temp_source_conn_prefs'])
                if 'temp_source_device_settings' in cfg:
                    for dev_type, settings in cfg['temp_source_device_settings'].items():
                        if dev_type in self._ts_device_settings:
                            self._ts_device_settings[dev_type].update(settings)
                # 每个设备类型独立保存的完整行参数 + 波动阈值
                if 'temp_source_full_settings' in cfg:
                    for dev_type, settings in cfg['temp_source_full_settings'].items():
                        if dev_type in self._ts_device_full_settings:
                            self._ts_device_full_settings[dev_type] = settings
                if 'temp_source_device' in cfg:
                    idx = self.temp_source_device_combo.findText(cfg['temp_source_device'])
                    if idx >= 0:
                        self.temp_source_device_combo.setCurrentIndex(idx)
                    else:
                        self.temp_source_device_combo.setCurrentText('Fluke 9250')
                if 'stability_threshold' in cfg:
                    self.stability_threshold_spin.setValue(cfg['stability_threshold'])
                    self.stability_threshold = cfg['stability_threshold']
                # 注意：temp_source_conn 保留为兼容旧配置，但设备切换时会用 _ts_device_conn_prefs 覆盖
                if 'temp_source_conn' in cfg:
                    saved_conn = cfg['temp_source_conn']
                    # 仅当对应设备偏好未设置时才使用旧格式的通用值
                    if self._ts_device_type not in self._ts_device_conn_prefs:
                        self._ts_device_conn_prefs[self._ts_device_type] = saved_conn
                if 'temp_source_port' in cfg:
                    port_text = cfg['temp_source_port']
                    idx = self.temp_source_port_combo.findText(port_text)
                    if idx >= 0:
                        self.temp_source_port_combo.setCurrentIndex(idx)
                    else:
                        if not self.temp_source_port_combo.isEditable() and port_text:
                            self.temp_source_port_combo.insertItem(0, port_text)
                            self.temp_source_port_combo.setCurrentIndex(0)
                        else:
                            self.temp_source_port_combo.setCurrentText(port_text)
                if 'temp_source_baud' in cfg:
                    self.temp_source_baud_spin.setValue(cfg['temp_source_baud'])
                if 'temp_source_ip' in cfg:
                    self.temp_source_ip_edit.setText(cfg['temp_source_ip'])
                if 'temp_source_lan_port' in cfg:
                    self.temp_source_lan_port_spin.setValue(cfg['temp_source_lan_port'])
                if 'row_main_pid_p' in cfg:
                    for i, v in enumerate(cfg['row_main_pid_p']):
                        if i < self._ts_row_count:
                            self.row_main_pid_p[i].setValue(v)
                if 'row_main_pid_i' in cfg:
                    for i, v in enumerate(cfg['row_main_pid_i']):
                        if i < self._ts_row_count:
                            self.row_main_pid_i[i].setValue(v)
                if 'row_main_pid_d' in cfg:
                    for i, v in enumerate(cfg['row_main_pid_d']):
                        if i < self._ts_row_count:
                            self.row_main_pid_d[i].setValue(v)
                if 'row_sec_pid_p' in cfg:
                    for i, v in enumerate(cfg['row_sec_pid_p']):
                        if i < self._ts_row_count:
                            self.row_sec_pid_p[i].setValue(v)
                if 'row_sec_pid_i' in cfg:
                    for i, v in enumerate(cfg['row_sec_pid_i']):
                        if i < self._ts_row_count:
                            self.row_sec_pid_i[i].setValue(v)
                if 'row_sec_pid_d' in cfg:
                    for i, v in enumerate(cfg['row_sec_pid_d']):
                        if i < self._ts_row_count:
                            self.row_sec_pid_d[i].setValue(v)
                if 'row_weights' in cfg:
                    for i, v in enumerate(cfg['row_weights']):
                        if i < self._ts_row_count:
                            parts = v.split(',')
                            for j in range(min(len(parts), len(self.row_weights[i]))):
                                self.row_weights[i][j].setText(parts[j])
                if 'manual_check_t3' in cfg:
                    self.manual_check_t3_cb.setChecked(bool(cfg['manual_check_t3']))

                # 加载完成后，根据当前设备的偏好恢复通讯方式
                # （若设备类型未变，currentTextChanged 信号不会触发，需手动恢复）
                saved_conn = self._ts_device_conn_prefs.get(self._ts_device_type, 'serial')
                if self.temp_source_conn_combo.currentText() != saved_conn:
                    self.temp_source_conn_combo.setCurrentText(saved_conn)
                if self._ts_device_type == 'Const 1210':
                    # Const 1210 强制 LAN 并禁用下拉框
                    self.temp_source_conn_combo.setEnabled(False)

                # 应用当前设备类型保存的完整行参数 + 波动阈值（每个设备独立记忆）
                cur_settings = self._ts_device_full_settings.get(self._ts_device_type)
                if cur_settings:
                    self._apply_ts_settings(cur_settings)
                # 应用后，把当前设备的连接参数也重新落回 UI（_apply_ts_settings 不改连接参数）
                cur_conn = self._ts_device_settings.get(self._ts_device_type, {})
                if 'port' in cur_conn:
                    port_text = cur_conn['port']
                    idx = self.temp_source_port_combo.findText(port_text)
                    if idx >= 0:
                        self.temp_source_port_combo.setCurrentIndex(idx)
                    else:
                        if not self.temp_source_port_combo.isEditable() and port_text:
                            self.temp_source_port_combo.insertItem(0, port_text)
                            self.temp_source_port_combo.setCurrentIndex(0)
                        else:
                            self.temp_source_port_combo.setCurrentText(port_text)
                if 'baud' in cur_conn:
                    self.temp_source_baud_spin.setValue(cur_conn['baud'])
                if 'ip' in cur_conn:
                    self.temp_source_ip_edit.setText(cur_conn['ip'])
                if 'lan_port' in cur_conn:
                    self.temp_source_lan_port_spin.setValue(cur_conn['lan_port'])
            except Exception as e:
                import traceback
                err_msg = f"加载配置失败: {e}\n{traceback.format_exc()}"
                print(err_msg)
                QMessageBox.warning(self, "配置加载提示", f"配置文件加载失败，已使用默认设置。\n错误: {e}")

    def save_config(self):
        if hasattr(self, '_loading') and self._loading:
            return
        # 同步当前UI参数到当前设备设置
        if self._ts_device_type:
            self._ts_device_settings[self._ts_device_type] = {
                'port': self.temp_source_port_combo.currentText(),
                'baud': self.temp_source_baud_spin.value(),
                'ip': self.temp_source_ip_edit.text(),
                'lan_port': self.temp_source_lan_port_spin.value(),
            }
            # 同步当前设备完整行参数 + 波动阈值
            self._save_current_ts_settings()
        cfg = {
            "devices": self.devices,
            "interval": self.interval_spin.value(),
            "test_interval": self.test_interval_spin.value(),
            "ts_row_count": self._ts_row_count,
            "temp_source_device": self._ts_device_type,
            "stability_threshold": self.stability_threshold_spin.value(),
            "row_setpoints": [self.row_setpoint_spins[i].value() for i in range(self._ts_row_count)],
            "row_specs": [self.row_setpoint_spec[i].value() for i in range(self._ts_row_count)],
            "row_mains": [self.row_main_spins[i].value() for i in range(self._ts_row_count)],
            "row_secs": [self.row_sec_spins[i].value() for i in range(self._ts_row_count)],
            "row_checks": [self.row_checks[i].isChecked() for i in range(self._ts_row_count)],
            "temp_source_conn_prefs": dict(self._ts_device_conn_prefs),
            "temp_source_device_settings": dict(self._ts_device_settings),
            "temp_source_full_settings": {k: v for k, v in self._ts_device_full_settings.items() if v is not None},
            "temp_source_conn": self.temp_source_conn_combo.currentText(),
            "temp_source_port": self.temp_source_port_combo.currentText(),
            "temp_source_baud": self.temp_source_baud_spin.value(),
            "temp_source_ip": self.temp_source_ip_edit.text(),
            "temp_source_lan_port": self.temp_source_lan_port_spin.value(),
            "row_main_pid_p": [self.row_main_pid_p[i].value() if self.row_main_pid_p[i].value() is not None else '' for i in range(self._ts_row_count)],
            "row_main_pid_i": [self.row_main_pid_i[i].value() if self.row_main_pid_i[i].value() is not None else '' for i in range(self._ts_row_count)],
            "row_main_pid_d": [self.row_main_pid_d[i].value() if self.row_main_pid_d[i].value() is not None else '' for i in range(self._ts_row_count)],
            "row_sec_pid_p": [self.row_sec_pid_p[i].value() if self.row_sec_pid_p[i].value() is not None else '' for i in range(self._ts_row_count)],
            "row_sec_pid_i": [self.row_sec_pid_i[i].value() if self.row_sec_pid_i[i].value() is not None else '' for i in range(self._ts_row_count)],
            "row_sec_pid_d": [self.row_sec_pid_d[i].value() if self.row_sec_pid_d[i].value() is not None else '' for i in range(self._ts_row_count)],
            "row_weights": [','.join(w.text() for w in self.row_weights[i]) for i in range(self._ts_row_count)],
            "manual_check_t3": self.manual_check_t3_cb.isChecked(),
        }
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(cfg, f, indent=4, ensure_ascii=False)
        except Exception as e:
            print(f"保存配置失败: {e}")


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    app.setStyleSheet("""
        QWidget { background-color: #ffffff; color: #000000; }
        QGroupBox { background-color: #ffffff; color: #000000; }
        QLabel { background-color: #ffffff; color: #000000; }
        QLineEdit { background-color: #ffffff; color: #000000; border: 1px solid #cccccc; border-radius: 3px; padding: 2px 5px; }
        QLineEdit::placeholder { color: #999999; }
        QComboBox { background-color: #ffffff; color: #000000; border: 1px solid #cccccc; border-radius: 3px; padding: 2px 5px; }
        QComboBox::drop-down { border: none; }
        QSpinBox { background-color: #ffffff; color: #000000; border: 1px solid #cccccc; border-radius: 3px; }
        QPushButton { background-color: #f5f5f5; color: #000000; }
        QTextEdit { background-color: #ffffff; color: #000000; }
        QCheckBox { background-color: #ffffff; color: #000000; }
        QCheckBox::indicator {
            width: 16px; height: 16px;
            border: 1.5px solid #555555; border-radius: 3px;
            background: #ffffff;
        }
        QCheckBox::indicator:checked {
            background: #2196F3; border: 1.5px solid #2196F3;
        }
        QCheckBox::indicator:unchecked {
            background: #ffffff; border: 1.5px solid #555555;
        }
        QStatusBar { background-color: #f0f0f0; color: #000000; }
    """)
    app.setFont(QFont(font_family, 9))
    window = DataCollectorApp()
    window.show()
    # 窗口显示后输出程序启动信息到终端区
    QTimer.singleShot(100, lambda: window._log_startup_info())
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
